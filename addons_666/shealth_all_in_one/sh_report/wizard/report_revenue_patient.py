# -*- coding: utf-8 -*-

from odoo import fields, api, models, _
from odoo.exceptions import ValidationError
from datetime import date, datetime, time, timedelta
from calendar import monthrange
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment, PatternFill
import base64
from io import BytesIO
from dateutil.relativedelta import relativedelta
from operator import itemgetter
from pytz import timezone

import logging

class SHReportRevenue(models.TransientModel):
    _name = 'sh.report.revenue'
    _description = 'Báo cáo doanh thu'

    start_date = fields.Date('Start date', default=date.today().replace(day=1))
    end_date = fields.Date('End date')

    def _get_report_type(self):
        report_list = [('patient', 'Doanh thu theo bệnh nhân'), ('service', 'Doanh thu theo dịch vụ')]

        return report_list

    report_type = fields.Selection(_get_report_type,
                                   'Report type', default='patient')

    # convert date to datetime for search domain, should be removed if using datetime directly
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')

    # convert date to datetime for search domain, should be removed if using datetime directly
    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        if self.start_date and self.end_date:
            fmt = "%Y-%m-%d %H:%M:%S"
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            now_utc = datetime.now(timezone('UTC'))
            local_tz = self.env.user.tz or 'Etc/GMT-7'
            now_timezone = now_utc.astimezone(timezone(local_tz))
            UTC_OFFSET_TIMEDELTA = datetime.strptime(now_utc.strftime(fmt), fmt) - datetime.strptime(
                now_timezone.strftime(fmt), fmt)
            utc_start_datetime = start_datetime + UTC_OFFSET_TIMEDELTA
            utc_end_datetime = end_datetime + UTC_OFFSET_TIMEDELTA
            self.start_datetime = utc_start_datetime
            self.end_datetime = utc_end_datetime
            # print(self.start_datetime)
            # print(self.end_datetime)

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    @api.multi
    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            if start_date > end_date:
                raise ValidationError(
                    _("End Date cannot be set before Start Date."))

    # BÁO CÁO DOANH THU THEO DỊCH VỤ
    def _get_revenue_service_data(self):
        ret_data = []
        revenue_data = self.env['account.payment'].search_read([('payment_date', '>=', self.start_date), ('payment_date', '<=', self.end_date), ('state','=','posted')], ['walkin'])
        w_ids = [w['walkin'][0] for w in revenue_data]
        service_data = self.env['sh.medical.appointment.register.walkin'].search_read([('id','in',w_ids)],['service','date'],order='date')

        start = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
        end = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 0, 0, 0)
        delta = timedelta(days=1)
        while start <= end:
            services_in_date = list(filter(lambda s: start <= s['date'] < start + delta, service_data))
            services = [service for service_list in services_in_date for service in service_list['service']]
            print(services_in_date)
            for service in set(services):
                service_detail = self.env['sh.medical.health.center.service'].browse(service)
                ret_data.append({'service_code': service_detail.default_code,'service_name': service_detail.name,'service_price': service_detail.list_price, 'total_count': services.count(service),'date': start.strftime('%d/%m/%Y')})
            start += delta

        return ret_data

        # [{'service_code': DV01,'service_name':'Tên dịch vụ 1', 'service_price':100000, 'total_count': 4, date': '2020/02/02 00:00:00'},{'service_code': DV02,'service_name':'Tên dịch vụ 2', 'service_price':200000, 'total_count': 1, date': '2020/02/02 00:00:00'}]

    def revenue_service_report(self):
        simple_revenue_service_attachment = self.env['ir.attachment'].browse(
            self.env.ref('shealth_all_in_one.report_revenue_service_template').id)
        decode = base64.b64decode(simple_revenue_service_attachment.datas)
        wb = load_workbook(BytesIO(decode))

        ws = wb.active
        institute = self.env['sh.medical.health.center'].search([('company_id', '=', self.env.user.company_id.id)],
                                                                limit=1)

        ws['A1'].value = institute.company_name
        ws['D4'].value = '%s' % (self.start_date.strftime('%d/%m/%Y'))
        ws['F4'].value = '%s' % (self.end_date.strftime('%d/%m/%Y'))
        thin = borders.Side(style='thin')
        all_border_thin = borders.Border(left=thin, right=thin, top=thin, bottom=thin)
        datas = self._get_revenue_service_data()
        # print(datas)
        line_font = Font(name='Times New Roman', size=12)

        key_col_list = [2, 3, 4, 5, 9]
        key_list = ['service_code', 'service_name', 'service_price', 'total_count', 'date']
        row = 8
        for line_data in datas:
            # for col, k in enumerate(key_list, 1):
            # index here
            ws.cell(row, 1).value = row - 7
            for col, k in zip(key_col_list, key_list):
                cell = ws.cell(row, col)
                cell.value, cell.font, cell.border = line_data[k], line_font, all_border_thin
                cell.alignment = Alignment(horizontal='center', vertical='center') if col != 3 else Alignment(
                    horizontal='left', vertical='center')

            row += 1
        row += 1
        ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=9)
        ws.cell(row, 7).value, ws.cell(row, 7).font, ws.cell(row, 7).alignment  = 'Người lập bảng',\
                                                                                  Font(name='Times New Roman', size=12, bold=True),\
                                                                                  Alignment(horizontal='center', vertical='center')

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({'name': 'MS report',
                                                              'datas_fname': 'bao_cao_doanh_thu_theo_dich_vu.xlsx',
                                                              'datas': report,
                                                              'res_model': 'temp.creation',
                                                              'public': True})

        return {'name': 'Báo cáo doanh thu theo dịch vụ',
                'type': 'ir.actions.act_window',
                'res_model': 'temp.wizard',
                'view_mode': 'form',
                'view_type': 'form',
                'target': 'inline',
                'view_id': self.env.ref('ms_templates.report_wizard').id,
                'context': {'attachment_id': attachment.id}}

    def report_revenue_patient_excel(self):
        template = self.env.ref('shealth_all_in_one.report_revenue_patient_template')
        domain_payment = [('payment_date', '>=', self.start_date), ('payment_date', '<=', self.end_date),('state','=','posted')]
        list_payment = self.env['account.payment'].search(domain_payment)
        # domain_result = [('payment_ids','in',list_payment.ids)]
        # records = self.env['account.payment'].search(domain_result)
        return {'name': (_('Báo cáo doanh thu theo bệnh nhân')),
            'type': 'ir.actions.act_window',
            'res_model': 'temp.wizard',
            'view_mode': 'form',
            'view_type': 'form',
            'target': 'inline',
            'view_id': self.env.ref('ms_templates.report_wizard').id,
            'context': {'default_template_id': template.id, 'active_ids': list_payment.ids,
                        'active_model': 'account.payment',
                        'external_keys': {'a1': list_payment[0].walkin.institution.name if len(list_payment) > 0 else 'BỆNH VIỆN THẨM MỸ KANGNAM',
                                          'e4': self.start_date.strftime('%d/%m/%Y'),
                                          'g4': self.end_date.strftime('%d/%m/%Y'),
                                          'j10007': self.env.user.name,'j10003': date.today().strftime('Ngày %d tháng %m năm %Y')}}}

    def report_revenue(self):
        if self.report_type in ('patient'):
            return self.report_revenue_patient_excel()
        else:
            return self.revenue_service_report()

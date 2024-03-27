from odoo import fields, api, models, _
from odoo.exceptions import ValidationError, UserError
from datetime import date, datetime, time
from datetime import timedelta
from calendar import monthrange
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment
import base64
from io import BytesIO
from openpyxl.styles import Font, borders, Alignment, PatternFill
from dateutil.relativedelta import relativedelta
from pytz import timezone

import logging

_logger = logging.getLogger(__name__)

thin = borders.Side(style='thin')
double = borders.Side(style='double')
all_border_thin = borders.Border(thin, thin, thin, thin)


class SoKhamBenhReport(models.TransientModel):
    _name = 'sokhambenh.report'
    _description = 'BÁO CÁO SỔ KHÁM BỆNH'

    department = fields.Many2one('sh.medical.health.center.ward', string='Khoa',
                                 help="Bộ phận của Trung tâm Y tế đã chọn")
    # department = fields.Selection([
    #     ('medical_examination', 'Khoa khám bệnh'), ('surgery', 'Khoa phẫu thuật thẩm mỹ'), ('anesthesia', 'Khoa gây mê hồi sức'), ('lab_image', 'Khoa cận lâm sàng'), ('dentomaxillofacial', 'Khoa răng-hàm-mặt')
    # ], default="medical_examination")
    start_date = fields.Date('Start date', default=date.today().replace(day=1))
    end_date = fields.Date('End date')
    # convert date to datetime for search domain, should be removed if using datetime directly
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')

    # convert date to datetime for search domain, should be removed if using datetime directly
    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        if self.start_date and self.end_date:
            fmt = "%Y-%m-%d %H:%M:%S"
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            now_utc = datetime.now(timezone('UTC'))
            local_tz = self.env.user.tz or 'Etc/GMT-7'
            now_timezone = now_utc.astimezone(timezone(local_tz))
            UTC_OFFSET_TIMEDELTA = datetime.strptime(now_utc.strftime(fmt), fmt) - datetime.strptime(
                now_timezone.strftime(fmt), fmt)
            utc_start_datetime = start_datetime + UTC_OFFSET_TIMEDELTA
            utc_end_datetime = end_datetime + UTC_OFFSET_TIMEDELTA
            self.start_datetime = utc_start_datetime
            self.end_datetime = utc_end_datetime

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    @api.multi
    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            if start_date > end_date:
                raise ValidationError(
                    _("End Date cannot be set before Start Date."))

    def _get_data_report(self):
    # def sokhambenh_report(self):
        ret_data = []
        """ Lấy ra các phiếu khám ở trạng thái hoàn thành """
        domain = [('evaluation_start_date', '>=', self.start_datetime), ('evaluation_end_date', '<=', self.end_datetime),
                  ('state', '=', 'Completed'), ('ward', '=', self.department.id)]
        evaluation = self.env['sh.medical.evaluation'].search(domain)
        a = self.start_date.day
        b = self.end_date.day
        check = (b - a) + 1
        for i in range(0, check):
            stt = 1
            ret_data.append([])
            data = self.start_date + timedelta(days=i)
            data_day = data.strftime('%d/%m/%Y')
            ret_data[i].append(data_day)
            partner_moves = evaluation.filtered(lambda m: m.evaluation_start_date.strftime('%d/%m/%Y') == data_day)
            for rec in partner_moves:
                value = [stt, rec.patient.identification_code, rec.patient.name,rec.patient.dob.strftime('%d/%m/%Y') if rec.patient.sex == "Male" else None, rec.patient.dob.strftime('%d/%m/%Y') if rec.patient.sex == "Female" else None,
                     rec.patient.street, None, rec.walkin.info_diagnosis, ','.join(rec.services.mapped('name')),'x', None, None, None, None, 'x', None, rec.doctor.name if rec.doctor.name else rec.walkin.doctor.name,'x', None]
                ret_data[i].append(value)
                stt += 1
        # return ret_data
            # ret_data.append({
            #     "stt": stt,
            #     "ma_benh_nhan": rec.patient.identification_code,
            #     "ho_ten_nguoi_benh": rec.patient.name,
            #     "nam": rec.patient.dob.strftime('%d/%m/%Y') if rec.patient.sex == "Male" else None,
            #     "nu": rec.patient.dob.strftime('%d/%m/%Y') if rec.patient.sex == "Female" else None,
            #     "dia_chi": rec.patient.street,
            #     "tuyen_duoi_chan_doan": None,
            #     "khoa_kham_benh": rec.walkin.info_diagnosis,
            #     "chi_dinh": ','.join(rec.services.mapped('name')),
            #     "vao_vien": 'x',
            #     "tuyen_tren": None,
            #     "tuyen_duoi": None,
            #     "ngoai_tru": None,
            #     "ve_nha": None,
            #     "tien_hanh_thu_thuat": 'x',
            #     "kham_chuyen_khoa": None,
            #     "bac_si": rec.doctor.name if rec.doctor.name else rec.walkin.doctor.name,
            #     "thu_phi": 'x',
            #     "cap_cuu": None,
            # })

    #     for rec in evaluation:
    #         value = [stt, rec.patient.identification_code,rec.patient.name,rec.patient.dob.strftime('%d/%m/%Y') if rec.patient.sex == "Male" else None,rec.patient.dob.strftime('%d/%m/%Y') if rec.patient.sex == "Female" else None,
    #             rec.patient.street,None,rec.walkin.info_diagnosis,','.join(rec.services.mapped('name')),'x',None,None,None,None,
    #             'x',None,rec.doctor.name if rec.doctor.name else rec.walkin.doctor.name,'x',None]
    #         ret_data.append(value)
    #         stt += 1
        return ret_data


        # xuất ra ex
    def sokhambenh_report(self):
        datas = self._get_data_report()
        # in du lieu
        report_brand_overview_attachment = self.env['ir.attachment'].browse(
            self.env.ref('shealth_all_in_one.bao_cao_so_kham_benh_attachment').id)
        decode = base64.b64decode(report_brand_overview_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        line_font = Font(name='Times New Roman', size=12)

        ws['F3'].value += self.start_date.strftime('%d/%m/%Y')
        ws['G3'].value += self.end_datetime.strftime('%d/%m/%Y')
        ws['A4'].value += self.department.name

        key_col_list = list(range(1, 20))
        key_list = [
            "stt",
            "ma_benh_nhan",
            "ho_ten_nguoi_benh",
            "nam",
            "nu",
            "dia_chi",
            "tuyen_duoi_chan_doan",
            "khoa_kham_benh",
            "chi_dinh",
            "vao_vien",
            "tuyen_tren",
            "tuyen_duoi",
            "ngoai_tru",
            "ve_nha",
            "tien_hanh_thu_thuat",
            "kham_chuyen_khoa",
            "bac_si",
            "thu_phi",
            "cap_cuu",
        ]
        row = 8
        for data in datas:
        #     print(data)
        #     for col, k in zip(key_col_list, key_list):
        #         beforeCell = ws.cell(5, col)
        #         beforeCell.font = Font(name='Times New Roman', size=12, color='FFFFFF')
        #         cell = ws.cell(row, col)
        #         cell.value = data[k]
        #         cell.font = line_font
        #         cell.border = all_border_thin
        #         cell.alignment = Alignment(horizontal='center', vertical='center')
        #     row += 1
            line_font = Font(name='Times New Roman', size=12)
            # for group in datas:
            ws.cell(row, 1).value, ws.cell(row, 1).font = data[0], line_font
            for col in range(1, 20):
                ws.cell(row, col).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                ws.cell(row, col).border = all_border_thin
            row += 1
            for i in range(1, len(data)):
                for col in range(1, 20):
                    # print('a')
                    ws.cell(row, col).value, ws.cell(row, col).font = data[i][col - 1], line_font
                for col in range(1, 20):
                    ws.cell(row, col).border = all_border_thin
                row += 1

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'Báo cáo sổ khám bệnh',
            'datas_fname': 'bao_cao_so_kham_benh.xlsx',
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=datas_fname&field=datas&download=true&filename=Báo cáo sổ khám bệnh.xlsx" \
              % (attachment.id)
        return {'name': 'Báo cáo sổ khám bệnh',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
                }

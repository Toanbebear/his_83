# -*- coding: utf-8 -*-

from odoo import fields, api, models, _
from odoo.exceptions import ValidationError, UserError
from datetime import date, datetime, time
from calendar import monthrange
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment
from pytz import timezone, utc
import base64
from io import BytesIO
from dateutil.relativedelta import relativedelta

import logging

_logger = logging.getLogger(__name__)

thin = borders.Side(style='thin')
double = borders.Side(style='double')
all_border_thin = borders.Border(thin, thin, thin, thin)


class DieutriReport(models.TransientModel):
    _name = 'dieutri.report'
    _description = 'Báo cáo điều trị'

    start_date = fields.Date('Start date', default=date.today().replace(day=1))
    end_date = fields.Date('End date')
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')
    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        self.start_datetime = False
        self.end_datetime = False
        if self.start_date and self.end_date:
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)
            self.start_datetime = start_datetime.astimezone(utc).replace(tzinfo=None)
            self.end_datetime = end_datetime.astimezone(utc).replace(tzinfo=None)

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    @api.multi
    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            if start_date > end_date:
                raise ValidationError(
                    _("End Date cannot be set before Start Date."))

    def report_dieutri(self):
        template = self.env['ir.attachment'].browse(self.env.ref('shealth_all_in_one.bao_cao_hoat_dong_dieu_tri_attachment').id)
        decode = base64.b64decode(template.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        thin = borders.Side(style='thin')
        all_border_thin = borders.Border(left=thin, right=thin, top=thin, bottom=thin)
        line_font = Font(name='Times New Roman', size=13)
        start_date = self.start_date
        end_date = self.end_date
        if start_date.month != end_date.month:
            raise UserError(_('Nhập ngày bắt đầu và kết thúc trong cùng một tháng'))
        if start_date.day != 1:
            raise UserError(_('Nhập ngày bắt đầu là mùng 1'))

        cell = ws.cell(3, 5)
        cell.value = 'Tháng %s năm %s' %(self.start_date.month, self.start_date.year)
        list_value = ['so_giuong_benh','so_nguoi_benh_dau_ky','tong_so_noi_tru','YHCT','tre_em_6','tre_em_15'
            ,'so_cap_cuu','so_ngay_dieu_tri','tong_so_tu_vong','tre_em_1_tv', 'tre_em_5_tv', 'tre_em_15_tv', 'truoc_24h',
            'bhyt','nguoi_benh_cuoi_ki']
        list_key = ['tong_so','tong_nu']
        config = self.env['ir.config_parameter'].sudo()
        da_lieu = config.get_param('da_lieu')
        gmhs = config.get_param('gmhs')
        pttm = config.get_param('pttm')
        rhm = config.get_param('rhm')
        list_key_2 = [int(da_lieu), int(gmhs), int(pttm), int(rhm)]
        for i in list_key_2:
            list_key.append(i)
        big_data = {}
        for key in list_key:
            big_data[key] = {}
        select_dau_ky_all = """
            select count(id)
            from sh_medical_appointment_register_walkin
            where date <= %s and date_out is NULL
        """
        select_cuoi_ky_all = """
                    select count(id)
                    from sh_medical_appointment_register_walkin
                    where date >= %s and date_out is NULL
                """
        select_dau_ky_nu = """
                    select count(id)
                    from sh_medical_appointment_register_walkin
                    where date <= %s and date_out is NULL and sex = 'Female'
                """
        select_cuoi_ky_nu = """
                            select count(id)
                            from sh_medical_appointment_register_walkin
                            where date >= %s and date_out is NULL and sex = 'Female'
                        """
        select_dau_ky_dk = """
                    select count(sma.id)
                    from sh_medical_appointment_register_walkin sma
                    left join sh_medical_health_center_ward smh on smh.id = sma.department
                    where date <= %s and date_out is NULL and sex = 'Female' and smh.id = %s
                """
        select_cuoi_ky_dk = """
                    select count(sma.id)
                    from sh_medical_appointment_register_walkin sma
                    left join sh_medical_health_center_ward smh on smh.id = sma.department
                    where date >= %s and date_out is NULL and sex = 'Female' and smh.id = %s
                """
        select_noi_tru = """
                    select count(smi.id)
                    from sh_medical_inpatient smi
                    where admission_date >= %s and admission_date <= %s
        """
        select_noi_tru_khoa = """
                            select count(smi.id)
                            from sh_medical_inpatient smi
                            left join sh_medical_health_center_ward smh on smh.id = smi.ward
                            where smi.admission_date >= %s and smi.admission_date <= %s and smh.id = %s
                """
        select_noi_tru_nu = """
                            select count(smi.id)
                            from sh_medical_inpatient smi
                            left join sh_medical_patient smp on smp.id = smi.patient
                            where smi.admission_date >= %s and smi.admission_date <= %s and smp.sex = 'Female' 
                """
        select_id_noi_tru = """
                select smi.id
                from sh_medical_inpatient smi
                where admission_date >= %s and admission_date <= %s
        """
        select_id_noi_tru_khoa = """
                        select smi.id
                        from sh_medical_inpatient smi
                        left join sh_medical_health_center_ward smh on smh.id = smi.ward
                        where smi.admission_date >= %s and smi.admission_date <= %s and smh.id = %s
                """
        select_id_noi_tru_nu = """
                        select smi.id
                        from sh_medical_inpatient smi
                        left join sh_medical_patient smp on smp.id = smi.patient
                        where smi.admission_date >= %s and smi.admission_date <= %s and smp.sex = 'Female' 
                """
        select_giuong_tong = """
        select count(DISTINCT hcb.id)
        from sh_medical_inpatient smi
        left join sh_medical_health_center_beds hcb on hcb.id = smi.bed
        where smi.admission_date >= %s and smi.admission_date <= %s 
        """
        select_giuong_nu = """
        select count(DISTINCT hcb.id)
        from sh_medical_inpatient smi
        left join sh_medical_health_center_beds hcb on hcb.id = smi.bed
        left join sh_medical_patient smp on smp.id = smi.patient
        where smp.sex = 'Female' and smi.admission_date >= %s and smi.admission_date <= %s 
        """
        select_giuong_khoa = """
        select count(DISTINCT hcb.id)
        from sh_medical_inpatient smi
        left join sh_medical_health_center_beds hcb on hcb.id = smi.bed
        left join sh_medical_health_center_ward hcw on hcw.id = smi.ward
        where smi.admission_date >= %s and smi.admission_date <= %s  and hcw.id = %s
        """
        ### TỔNG
        # Tính tổng số lượng đầu kỳ
        self.env.cr.execute(select_dau_ky_all,
                            (self.start_datetime.strftime('%Y-%m-%d %H:%M:%S'),))
        result = self.env.cr.fetchall()
        for c in result:
            big_data['tong_so']['so_nguoi_benh_dau_ky'] = c[0]
        # Tính tổng số lượng giường bệnh
        self.env.cr.execute(select_giuong_tong,
                            (self.start_datetime.strftime('%Y-%m-%d %H:%M:%S'), self.end_datetime.strftime('%Y-%m-%d %H:%M:%S')))
        result = self.env.cr.fetchall()
        for c in result:
            big_data['tong_so']['so_giuong_benh'] = c[0]
        # Tính tổng số người bệnh điều trị nội trú
        self.env.cr.execute(select_id_noi_tru,
                            (self.start_datetime.strftime('%Y-%m-%d %H:%M:%S'),self.end_datetime.strftime('%Y-%m-%d %H:%M:%S')))
        result = self.env.cr.fetchall()
        tong_1 = 0
        for c in result:
            inpatient = self.env['sh.medical.inpatient'].sudo().browse(int(c[0]))
            ngay_ra = datetime(inpatient.discharge_date.year, inpatient.discharge_date.month, inpatient.discharge_date.day) if inpatient.discharge_date else datetime(date.today().year,date.today().month,date.today().day)
            so_ngay = ngay_ra - datetime(inpatient.admission_date.year, inpatient.admission_date.month, inpatient.admission_date.day) if inpatient.admission_date.day else 0
            tong_1 += so_ngay.days+1
        big_data['tong_so']['so_ngay_dieu_tri'] = tong_1
        # Tính tổng số ngày điều trị nội trú
        self.env.cr.execute(select_noi_tru,
                            (self.start_datetime.strftime('%Y-%m-%d %H:%M:%S'), self.end_datetime.strftime('%Y-%m-%d %H:%M:%S')))
        result = self.env.cr.fetchall()
        for c in result:
            big_data['tong_so']['tong_so_noi_tru'] = c[0]
        # Tính tổng số người bệnh còn lại cuối kỳ
        self.env.cr.execute(select_cuoi_ky_all,
                            (self.end_datetime.strftime('%Y-%m-%d %H:%M:%S'),))
        result = self.env.cr.fetchall()
        for c in result:
            big_data['tong_so']['nguoi_benh_cuoi_ki'] = c[0]

        ### Nữ
        # Tính tổng số lượng đầu kỳ nữ
        self.env.cr.execute(select_dau_ky_nu,
                            (self.start_datetime.strftime('%Y-%m-%d %H:%M:%S'),))
        result = self.env.cr.fetchall()
        for c in result:
            big_data['tong_nu']['so_nguoi_benh_dau_ky'] = c[0]
        # Tính tổng số lượng giường bệnh
        self.env.cr.execute(select_giuong_nu, (self.start_datetime.strftime('%Y-%m-%d %H:%M:%S'), self.end_datetime.strftime('%Y-%m-%d %H:%M:%S')))
        result = self.env.cr.fetchall()
        for c in result:
            big_data['tong_nu']['so_giuong_benh'] = c[0]
        # Tính tổng số người bệnh điều trị nội trú
        self.env.cr.execute(select_id_noi_tru_nu,
                            (self.start_datetime.strftime('%Y-%m-%d %H:%M:%S'), self.end_datetime.strftime('%Y-%m-%d %H:%M:%S')))
        result = self.env.cr.fetchall()
        tong_2 = 0
        for c in result:
            inpatient = self.env['sh.medical.inpatient'].sudo().browse(int(c[0]))
            ngay_ra = datetime(inpatient.discharge_date.year, inpatient.discharge_date.month,
                               inpatient.discharge_date.day) if inpatient.discharge_date else datetime(
                date.today().year, date.today().month, date.today().day)
            so_ngay = ngay_ra - datetime(inpatient.admission_date.year, inpatient.admission_date.month,
                                         inpatient.admission_date.day) if inpatient.admission_date.day else 0
            tong_2 += so_ngay.days + 1
        big_data['tong_nu']['so_ngay_dieu_tri'] = tong_2
        # Tính tổng số ngày điều trị nội trú
        self.env.cr.execute(select_noi_tru_nu,
                            (self.start_datetime.strftime('%Y-%m-%d %H:%M:%S'), self.end_datetime.strftime('%Y-%m-%d %H:%M:%S')))
        result = self.env.cr.fetchall()
        for c in result:
            big_data['tong_nu']['tong_so_noi_tru'] = c[0]
        # Tính tổng số người bệnh còn lại cuối kỳ
        self.env.cr.execute(select_cuoi_ky_nu,
                            (self.end_datetime.strftime('%Y-%m-%d %H:%M:%S'),))
        result = self.env.cr.fetchall()
        for c in result:
            big_data['tong_nu']['nguoi_benh_cuoi_ki'] = c[0]

        ### Theo từng khoa
        for id_khoa in list_key_2:
            # Tính tổng số lượng đầu kỳ nữ
            self.env.cr.execute(select_dau_ky_dk,
                                (self.start_datetime, int(id_khoa)))
            result = self.env.cr.fetchall()
            for c in result:
                big_data[int(id_khoa)]['so_nguoi_benh_dau_ky'] = c[0]
            # Tính tổng số lượng giường bệnh
            self.env.cr.execute(select_giuong_khoa,(self.start_datetime.strftime('%Y-%m-%d %H:%M:%S'), self.end_datetime.strftime('%Y-%m-%d %H:%M:%S'),int(id_khoa)))
            result = self.env.cr.fetchall()
            for c in result:
                big_data[int(id_khoa)]['so_giuong_benh'] = c[0]
            # Tính tổng số người bệnh điều trị nội trú
            self.env.cr.execute(select_id_noi_tru_khoa,
                                (self.start_datetime.strftime('%Y-%m-%d %H:%M:%S'), self.end_datetime.strftime('%Y-%m-%d %H:%M:%S'), int(id_khoa)))
            result = self.env.cr.fetchall()
            tong_3 = 0
            for c in result:
                inpatient = self.env['sh.medical.inpatient'].sudo().browse(int(c[0]))
                ngay_ra = datetime(inpatient.discharge_date.year, inpatient.discharge_date.month,
                                   inpatient.discharge_date.day) if inpatient.discharge_date else datetime(
                    date.today().year, date.today().month, date.today().day)
                so_ngay = ngay_ra - datetime(inpatient.admission_date.year, inpatient.admission_date.month,
                                             inpatient.admission_date.day) if inpatient.admission_date.day else 0
                tong_3 += so_ngay.days + 1
            big_data[int(id_khoa)]['so_ngay_dieu_tri'] = tong_3
            # Tính tổng số ngày điều trị nội trú
            self.env.cr.execute(select_noi_tru_khoa,
                                (self.start_datetime.strftime('%Y-%m-%d %H:%M:%S'), self.end_datetime.strftime('%Y-%m-%d %H:%M:%S'), int(id_khoa)))
            result = self.env.cr.fetchall()
            for c in result:
                big_data[int(id_khoa)]['tong_so_noi_tru'] = c[0]
            # Tính tổng số người bệnh còn lại cuối kỳ
            self.env.cr.execute(select_cuoi_ky_dk,
                                (self.end_datetime.strftime('%Y-%m-%d %H:%M:%S'), int(id_khoa)))
            result = self.env.cr.fetchall()
            for c in result:
                big_data[int(id_khoa)]['nguoi_benh_cuoi_ki'] = c[0]
        row = 10
        for key in list_key:
            col = 4
            for value in list_value:
                cell = ws.cell(row, col)
                cell.value = big_data[key][value] if value in big_data[key] else 0
                cell.font = line_font
                cell.border = all_border_thin
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                col += 1
            row += 1
        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'Báo cáo hoạt động điều trị.xlsx',
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=name&field=datas&download=true" \
              % attachment.id
        return {'name': 'Báo cáo hoạt động điều trị',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
                }

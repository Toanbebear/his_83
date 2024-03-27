from odoo import fields, api, models, _
from odoo.exceptions import ValidationError, UserError
from datetime import date, datetime, time, timedelta
from calendar import monthrange
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment
import base64
from io import BytesIO
from dateutil.relativedelta import relativedelta
from pytz import timezone

import logging

_logger = logging.getLogger(__name__)

thin = borders.Side(style='thin')
double = borders.Side(style='double')
all_border_thin = borders.Border(thin, thin, thin, thin)

class SoRaVaoVienReport(models.TransientModel):
    _name = 'soravaovien.report'
    _description = 'BÁO CÁO SỔ RA VÀO VIỆN'

    department = fields.Selection([('surgery', 'Khoa phẫu thuật thẩm mỹ'), ('anesthesia', 'Khoa gây mê hồi sức'), ('spa', 'Spa'), ('dentomaxillofacial', 'Khoa răng-hàm-mặt')], default="surgery", string='Khoa', help="Bộ phận của Trung tâm Y tế đã chọn")
    start_date = fields.Date('Start date', default=date.today().replace(day=1))
    end_date = fields.Date('End date')
    # convert date to datetime for search domain, should be removed if using datetime directly
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')

    # convert date to datetime for search domain, should be removed if using datetime directly
    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        if self.start_date and self.end_date:
            fmt = "%Y-%m-%d %H:%M:%S"
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            now_utc = datetime.now(timezone('UTC'))
            local_tz = self.env.user.tz or 'Etc/GMT-7'
            now_timezone = now_utc.astimezone(timezone(local_tz))
            UTC_OFFSET_TIMEDELTA = datetime.strptime(now_utc.strftime(fmt), fmt) - datetime.strptime(
                now_timezone.strftime(fmt), fmt)
            utc_start_datetime = start_datetime + UTC_OFFSET_TIMEDELTA
            utc_end_datetime = end_datetime + UTC_OFFSET_TIMEDELTA
            self.start_datetime = utc_start_datetime
            self.end_datetime = utc_end_datetime

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    @api.multi
    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            if start_date > end_date:
                raise ValidationError(
                    _("End Date cannot be set before Start Date."))

    # Lấy dữ liệu
    def _get_data_report(self):
        ret_data = []
        if self.department == 'surgery':
            domain_surgery = [('surgery_date', '>=', self.start_datetime), ('surgery_date', '<=', self.end_datetime),
                      ('state', '=', 'Done'), ('operating_room', '=', self.env.ref('shealth_all_in_one.sh_supersonic_daiphau_room_knhn').id)]
            walkin = self.env['sh.medical.surgery'].search(domain_surgery)
        elif self.department == 'anesthesia':
            domain_anesthesia = [('surgery_date', '>=', self.start_datetime), ('surgery_date', '<=', self.end_datetime),
                              ('state', '=', 'Done'), ('operating_room', '=', self.env.ref('shealth_all_in_one.sh_supersonic_tieuphau_room_knhn').id)]
            walkin = self.env['sh.medical.surgery'].search(domain_anesthesia)
        elif self.department == 'spa':
            domain_anesthesia = [('services_date', '>=', self.start_datetime), ('services_date', '<=', self.end_datetime),
                                 ('state', '=', 'Done'), ('department', '=', self.env.ref(
                    'shealth_all_in_one.sh_dalieu_dep_knhn').id)]
            walkin = self.env['sh.medical.specialty'].search(domain_anesthesia)
        elif self.department == 'dentomaxillofacial':
            domain_anesthesia = [('services_date', '>=', self.start_datetime), ('services_date', '<=', self.end_datetime),
                                 ('state', '=', 'Done'), ('department', '=', self.env.ref(
                    'shealth_all_in_one.sh_rhm_dep_knhn').id)]
            walkin = self.env['sh.medical.specialty'].search(domain_anesthesia)
        datas_name = walkin.mapped('name')
        stt = 1
        for rec in walkin:
            d1 = rec.walkin.date + timedelta(hours=7)
            d2 = rec.walkin.date_out + timedelta(hours=7) if rec.walkin.date_out else ''
            ret_data.append({
                "stt": stt,
                "so_vao_vien": rec.walkin.hospital_number or None,
                "ho_ten_nguoi_benh":rec.patient.name,
                "nam": rec.walkin.dob.strftime('%d/%m/%Y') if rec.walkin.sex == "Male" else None,
                "nu": rec.walkin.dob.strftime('%d/%m/%Y') if rec.walkin.sex == "Female" else None,
                "cong_nhan_vc": None,
                "bhyt": None,
                "tt": 'x',
                "nt": None,
                "<12t": None,
                "1-15t": None,
                "nghe_nghiep": rec.patient.function or None,
                "dia_chi": rec.walkin.patient.street or None,
                "noi_gioi_thieu": 'Tự Đến',
                "vao_vien":d1,
                "chuyen_vien": None,
                "ra_vien":d2,
                "tv": None,
                "tv_trong_24h": None,
                "tuyen_duoi": None,
                "phong_kham": rec.walkin.pathological_process or None,
                "khoa_dieu_tri": rec.walkin.info_diagnosis or None,
                "khoa_gpb": None,
                "khoi": None,
                "do": 'x',
                "nang_hon": None,
                "khong_thay_doi": None,
                "ngay_dt": (rec.walkin.date_out-rec.walkin.date).days +1 if rec.walkin.date_out else '',

            })
            stt += 1
        print(ret_data)
        return ret_data
    # xuất ra ex
    def soravaovien_report(self):
        dict_patient_level = {
            'surgery': 'Khoa phẫu thuật thẩm mỹ',
            'anesthesia': 'Khoa gây mê hồi sức',
            'spa': 'Spa',
            'dentomaxillofacial': 'Khoa răng-hàm-mặt',
        }

        datas = self._get_data_report()
        # in du lieu
        report_brand_overview_attachment = self.env['ir.attachment'].browse(
            self.env.ref('shealth_all_in_one.bao_cao_so_ra_vao_vien_attachment').id)
        decode = base64.b64decode(report_brand_overview_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        line_font = Font(name='Times New Roman', size=12)

        ws['A2'].value += dict_patient_level['%s' % self.department]
        ws['A3'].value += self.start_date.strftime('%d/%m/%Y')
        ws['A4'].value += self.end_datetime.strftime('%d/%m/%Y')

        key_col_list = list(range(1, 29))
        key_list = [
            "stt",
            "so_vao_vien",
            "ho_ten_nguoi_benh",
            "nam",
            "nu",
            "cong_nhan_vc",
            "bhyt",
            "tt",
            "nt",
            "<12t",
            "1-15t",
            "nghe_nghiep",
            "dia_chi",
            "noi_gioi_thieu",
            "vao_vien",
            "chuyen_vien",
            "ra_vien",
            "tv",
            "tv_trong_24h",
            "tuyen_duoi",
            "phong_kham",
            "khoa_dieu_tri",
            "khoa_gpb",
            "khoi",
            "do",
            "nang_hon",
            "khong_thay_doi",
            "ngay_dt",
        ]
        row = 9
        for data in datas:
            for col, k in zip(key_col_list, key_list):
                beforeCell = ws.cell(8, col)
                beforeCell.font = Font(name='Times New Roman', size=12, color='000000')
                cell = ws.cell(row, col)
                cell.value = data[k]
                cell.font = line_font
                cell.border = all_border_thin
                cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 1
        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'Báo cáo sổ ra vào viện',
            'datas_fname': 'bao_cao_so_ra_vao_vien.xlsx',
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=datas_fname&field=datas&download=true&filename=Báo cáo sổ ra vào viện.xlsx" \
              % (attachment.id)
        return {'name': 'Báo cáo sổ ra vào viện',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
                }

class SHealthAppointmentWalkin(models.Model):
    _inherit = 'sh.medical.appointment.register.walkin'

    hospital_number = fields.Char('Số vào viện',)
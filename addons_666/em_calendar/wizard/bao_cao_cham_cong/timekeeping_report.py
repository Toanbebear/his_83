from odoo import fields, api, models, _
from odoo.exceptions import ValidationError, UserError
from datetime import date, datetime, time, timedelta
from calendar import monthrange
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment
import base64
from openpyxl.styles import Font, borders, Alignment, PatternFill
from io import BytesIO
from dateutil.relativedelta import relativedelta
from pytz import timezone, utc

thin = borders.Side(style='thin')
double = borders.Side(style='double')
all_border_thin = borders.Border(thin, thin, thin, thin)

MONTHS = [('1', 'Tháng 1'), ('2', 'Tháng 2'), ('3', 'Tháng 3'), ('4', 'Tháng 4'), ('5', 'Tháng 5'), ('6', 'Tháng 6'),
          ('7', 'Tháng 7'), ('8', 'Tháng 8'), ('9', 'Tháng 9'), ('10', 'Tháng 10'), ('11', 'Tháng 11'),
          ('12', 'Tháng 12')]

YEARS = [('2021', '2021'), ('2022', '2022'), ('2023', '2023'), ('2024', '2024'), ('2025', '2025'), ('2026', '2026'),
         ('2027', '2027'), ('2028', '2028'), ('2029', '2029'), ('2030', '2030')]


class TimeKeeping(models.TransientModel):
    _name = 'timekeeping.report'
    _description = 'Báo cáo chấm công'

    department_id = fields.Many2one('sh.medical.health.center.ward', 'Khoa/Phòng')

    start_year = fields.Selection(YEARS, 'Năm', default=lambda *a: str(date.today().year))
    start_month = fields.Selection(MONTHS, 'Tháng', default=lambda *a: '1')
    end_year = fields.Selection(YEARS, 'Tới năm', default=lambda *a: str(date.today().year))
    end_month = fields.Selection(MONTHS, 'Tới tháng', default=lambda *a: '1')

    @api.onchange('start_year')
    def _onchange_start_year(self):
        if self.start_year:
            self.end_year = self.start_year

    @api.onchange('start_month')
    def _onchange_end_year(self):
        if self.start_month:
            self.end_month = self.start_month

    @api.constrains('start_year', 'end_year')
    def check_year(self):
        for record in self:
            start_year = int(record.start_year)
            end_year = int(record.end_year)

            if start_year != end_year:
                raise ValidationError("Hệ thống hỗ trợ báo cáo trong cùng 1 năm. "
                                      "Vui lòng chọn thời gian bắt đầu và thời gian kết thúc trong 1 năm.")

    @api.constrains('start_month', 'end_month')
    def check_month(self):
        for record in self:
            start_month = int(record.start_month)
            end_month = int(record.end_month)

            if start_month > end_month:
                raise ValidationError("Thời gian bắt đầu phải trước hoặc bằng thời gian kết thúc.")

    def _get_time(self):
        start_year = int(self.start_year)
        start_month = int(self.start_month)
        end_year = int(self.end_year)
        end_month = int(self.end_month)
        for month in range(start_month, end_month + 1):
            today_last_day_in_start = monthrange(start_year, month)
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')

            start_datetime = datetime(start_year, month, 1, 0, 0, 0)
            end_datetime = datetime(end_year, month, today_last_day_in_start[1], 23, 59, 59)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)

            start_time_month = start_datetime.astimezone(utc).replace(tzinfo=None)
            end_time_month = end_datetime.astimezone(utc).replace(tzinfo=None)

            """Lấy ngày thứ trong tháng"""
            today = start_datetime
            first_day = datetime(today.year, today.month, 1)

            weekdays = ("T2", "T3", "T4", "T5", "T6", "T7", "CN")
            dates = []
            for i in range(31):
                date = first_day + timedelta(days=i)
                if date.month == first_day.month:
                    weekdays_idx = date.weekday()
                    dates.append((date.strftime('%d'), weekdays[weekdays_idx]))
            return dates
    def _get_data(self):
    # def create_report(self):
        start_year = int(self.start_year)
        start_month = int(self.start_month)
        end_year = int(self.end_year)
        end_month = int(self.end_month)
        for month in range(start_month, end_month + 1):
            today_last_day_in_start = monthrange(start_year, month)
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')

            start_datetime = datetime(start_year, month, 1, 0, 0, 0)
            end_datetime = datetime(end_year, month, today_last_day_in_start[1], 23, 59, 59)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)

            start_time_month = start_datetime.astimezone(utc).replace(tzinfo=None)
            end_time_month = end_datetime.astimezone(utc).replace(tzinfo=None)

            """Lấy ngày thứ trong tháng"""
            today = start_datetime
            first_day = datetime(today.year, today.month, 1)

            weekdays = ("T2", "T3", "T4", "T5", "T6", "T7", "CN")
            dates = []
            for i in range(31):
                date = first_day + timedelta(days=i)
                if date.month == first_day.month:
                    weekdays_idx = date.weekday()
                    dates.append((date.strftime('%d'), weekdays[weekdays_idx]))

            """" Lấy tất cả nhân sự làm việc tại khoa phòng  """

            querry = """ select hr.id, hr.name, j.name,
                         CASE
                             WHEN es.status = '1' THEN 'Off'
                             WHEN es.status = '2' THEN 'X'
                             WHEN es.status = '3' THEN 'KL'
                             WHEN es.status = '4' THEN 'X/12'
                             WHEN es.status = '5' THEN 'CT'
                             WHEN es.status = '6' THEN 'CT'
                             WHEN es.status = '7' THEN 'RT'
                             WHEN es.status = '8' THEN 'L'
                             WHEN es.status = '9' THEN 'TS'
                             WHEN es.status = '10' THEN 'KL'
                             WHEN es.status = '11' THEN 'O'
                             WHEN es.status = '12' THEN 'T'
                             ELSE ''
                       END AS status,
                        to_char(es.start_date, 'DD')
              from hr_employee hr
                left join em_schedule_hr_employee_rel eer on eer.hr_employee_id = hr.id
                left join em_schedule es on es.id = eer.em_schedule_id
                left join hr_job j on hr.job_id = j.id
             where es.department_id = %s and es.state in ('2', '3') and es.start_date >= %s and es.start_date <= %s"""
            self.env.cr.execute(querry, (self.department_id.id, start_time_month, end_time_month))
            result_ = self.env.cr.fetchall()
            dict = {}
            for rec in result_:
                id = rec[0]
                date = rec[4]
                status = rec[3]
                if id not in dict:
                    dict[id] = {}
                dict[id][date] = status

            """" Lấy tất cả nhân sự làm việc tại khoa phòng  """
            querry = """ select distinct hr.id, hr.name, j.name  from hr_employee hr
                left join em_schedule_hr_employee_rel eer on eer.hr_employee_id = hr.id
                left join em_schedule es on es.id = eer.em_schedule_id
                left join hr_job j on hr.job_id = j.id
             where es.department_id = %s and es.state in ('2', '3') and es.start_date >= %s and es.start_date <= %s"""
            self.env.cr.execute(querry, (self.department_id.id, start_time_month, end_time_month))
            result = self.env.cr.fetchall()
            dict_result = {}
            for rec in result:
                id = rec[0]
                name = rec[1]
                department = rec[2]
                dict_result[id] = {'code': None, 'name': name, 'department': department}
                for date in dates:
                    dict_result[id][date[0]] = ''

            for key, value in dict.items():
                if key in dict_result:
                    dict_result[key].update(value)

        return dict_result


    def create_report(self):
        inventory_attachment = self.env['ir.attachment'].sudo().browse(
            self.env.ref('em_calendar.bao_cao_cham_cong_attachment').id)
        decode = base64.b64decode(inventory_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        datas = self._get_data()
        datas_time = self._get_time()
        ws['C3'].value += self.start_year
        ws['B3'].value += self.start_month
        ws['A2'].value = self.department_id.name

        line_font = Font(name='Times New Roman', size=12, )
        line_font_st = Font(name='Times New Roman', size=11, bold=True)
        center_alm = Alignment(horizontal='center', vertical='center')

        col_time = 5
        end_row = 7
        for group in datas_time:
            for i in range(0, len(group)):
                ws.cell(5+i, col_time).value, ws.cell(5+i, col_time).font = group[i], line_font_st
                # bắt đầu từ hàng số 5, cột số 6
                # for row in range(5, end_row):
                    # ws.cell(5, col).value, ws.cell(5, col).font = group[i], line_font_st
                    # ws.cell(6, col).value, ws.cell(6, col).font = group[i], line_font_st
                for row in range(5, end_row):
                    ws.cell(row, col_time).border = all_border_thin
                for row in range(5, end_row):
                    ws.cell(row, col_time).alignment = center_alm
            col_time += 1
        stt = 1
        row = 7
        for key, value in datas.items():
            # In các giá trị cột 1 đến cột 4
            ws.cell(row, 1).value = stt
            ws.cell(row, 2).value = value['code']
            ws.cell(row, 3).value = value['name']
            ws.cell(row, 4).value = value['department']
            stt += 1

            max_day = None
            if any(x.isdigit() and value[x].isdigit() for x in value):
                max_day = max(int(value[x]) for x in value if x.isdigit() and value[x].isdigit())

            # In các giá trị của các cột '01' đến '31'
            for col in range(5, 36):
                col_name = '{:02d}'.format(col - 4)
                if col_name in value:
                    ws.cell(row, col).value = value[col_name]

            # if len(datas_time) == 28:
            for col in range(1, len(datas_time) + 5):
                ws.cell(row, col).font = line_font
                ws.cell(row, col).border = all_border_thin
            for col in range(1, len(datas_time)):
                ws.cell(row, col).alignment = center_alm

            row += 1

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'Báo cáo danh sách nhân sự',
            'datas_fname': 'bao_cao_cham_cong.xlsx',
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })

        url = "/web/content/?model=ir.attachment&id=%s&filename_field=datas_fname&field=datas&download=true&filename=Báo cáo chấm công.xlsx" \
              % (attachment.id)
        return {
            'name': 'BÁO CÁO CHẤM CÔNG',
            'type': 'ir.actions.act_url',
            'url': url,
            'target': 'self',
        }
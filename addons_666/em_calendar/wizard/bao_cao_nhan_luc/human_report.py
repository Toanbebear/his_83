from odoo import fields, api, models, _
from odoo.exceptions import ValidationError, UserError
from datetime import date, datetime, time
from calendar import monthrange
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment
import base64
from io import BytesIO
from dateutil.relativedelta import relativedelta
from pytz import timezone, utc
from collections import defaultdict

# khoa_kham_benh = ['kh_cu', 'kh_moi', 'kh_xuat', 'kh_con', 'ns', 'ns_di_lam', 'ns_off', 'ns_nghi', 'ns_di_hoc','ns_tang_cuong']
# khoa_gay_me_hoi_suc = {'kh_cu', 'kh_moi', 'kh_xuat', 'kh_con', 'ns', 'ns_di_lam', 'ns_off', 'ns_nghi', 'ns_di_hoc', 'ns_tang_cuong'}
# khoa_pttm = {'kh_cu', 'kh_moi', 'kh_xuat', 'kh_con', 'ns', 'ns_di_lam', 'ns_off', 'ns_nghi', 'ns_di_hoc', 'ns_tang_cuong'}
# khoa_rhm = {'kh_cu', 'kh_moi', 'kh_xuat', 'kh_con', 'ns', 'ns_di_lam', 'ns_off', 'ns_nghi', 'ns_di_hoc', 'ns_tang_cuong'}
# khoa_da_lieu = {'kh_cu', 'kh_moi', 'kh_xuat', 'kh_con', 'ns', 'ns_di_lam', 'ns_off', 'ns_nghi', 'ns_di_hoc', 'ns_tang_cuong'}
# khoa_can_lam_san = {'kh_cu', 'kh_moi', 'kh_xuat', 'kh_con', 'ns', 'ns_di_lam', 'ns_off', 'ns_nghi', 'ns_di_hoc', 'ns_tang_cuong'}
thin = borders.Side(style='thin')
double = borders.Side(style='double')
all_border_thin = borders.Border(thin, thin, thin, thin)

class HumanReport(models.TransientModel):
    _name = 'human.report'
    _description = 'Báo cáo nhân lực'

    start_date = fields.Date('Ngày báo cáo', default=date.today().replace(day=1))
    end_date = fields.Date('Ngày kết thúc')
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')

    # check ngày
    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        self.start_datetime = False
        self.end_datetime = False
        if self.start_date and self.end_date:
            local_tz = timezone(self.env.user.tz or 'Etc/GMT+7')
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            start_datetime = local_tz.localize(start_datetime, is_dst=None)
            end_datetime = local_tz.localize(end_datetime, is_dst=None)
            self.start_datetime = start_datetime.astimezone(utc).replace(tzinfo=None)
            self.end_datetime = end_datetime.astimezone(utc).replace(tzinfo=None)

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = self.start_date
            else:
                self.end_date = self.start_date

    @api.multi
    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            if start_date > end_date:
                raise ValidationError(
                    _("Ngày kết thúc không thể ở trước ngày bắt đầu khi xuất báo cáo!!!"))

    #lấy dữ liệu khoa khám bệnh
    def _get_data_khoa_kham_benh(self):
        ret_data = []
        # """ Đứng ở bệnh nhân lưu tìm ra tất cả bản ghi có trạng thái đang nhập viện, lấy ra bệnh nhân cũ """
        # inpatient_cu = self.env['sh.medical.evaluation'].search([('evaluation_start_date', '<', self.start_date), ('state', '=', 'InProgress'), ('ward', '=', self.env.ref('shealth_all_in_one.sh_kb_dep_knhn').id)])

        """ Đứng ở bệnh nhân lưu tìm ra tất cả bản ghi có trạng thái đang nhập viện, lấy ra bệnh nhân mới """
        inpatient_moi = self.env['sh.medical.evaluation'].search([('evaluation_start_date', '>=', self.start_datetime), ('evaluation_start_date', '<=', self.end_datetime), ('ward', '=', self.env.ref('shealth_all_in_one.sh_kb_dep_knhn').id)])

        """ Đứng ở bệnh nhân lưu tìm ra tất cả bản ghi có trạng thái đang nhập viện, đã nhập viện, lấy ra bệnh nhân xuất viện """
        inpatient_xuat = self.env['sh.medical.evaluation'].search(
            [('evaluation_end_date', '>=', self.start_datetime), ('evaluation_end_date', '<=', self.end_datetime), ('ward', '=', self.env.ref('shealth_all_in_one.sh_kb_dep_knhn').id)])

        """ Đứng ở Thông tin bác sĩ tìm ra tất cả bản ghi """
        physician = self.env['sh.medical.physician'].search([('speciality', 'in', [4, 21]), ('department', '=', self.env.ref('shealth_all_in_one.sh_kb_dep_knhn').id), ('status', '=', False)])

        count_rested = 0
        """ đứng ở lịch làm việc lấy ra tổng nhân sự đi làm """
        schedule_rested = self.env["em.schedule"].search([('status', 'in', ['2', '4', '7']), ('state', 'in', ['2', '3']),
                ('department_id', '=', self.env.ref('shealth_all_in_one.sh_kb_dep_knhn').id), ('start_date', '<=', self.start_date), ('end_date', '>=', self.end_date)])
        seen_employee_ids = set()
        for rec in schedule_rested:
            for recc in rec.employee:
                if recc.id not in seen_employee_ids:
                    seen_employee_ids.add(recc.id)
                    for nv in physician:
                        if recc in nv.employee_id:
                            count_rested += 1

        count_nghi = 0
        """ đứng ở lịch làm việc lấy ra tổng nhân sự nghỉ dài hạn """
        schedule_nghi = self.env["em.schedule"].search([('status', 'in', ['3', '9']), ('state', 'in', ['2', '3']), ('start_date', '<=', self.start_date),
                ('end_date', '>=', self.end_date), ('department_id', '=', self.env.ref('shealth_all_in_one.sh_kb_dep_knhn').id)])
        employee_nghi_ids = set()
        for rec in schedule_nghi:
            for recc in rec.employee:
                if recc.id not in employee_nghi_ids:
                    employee_nghi_ids.add(recc.id)
                    for nv in physician:
                        if recc in nv.employee_id:
                            count_nghi += 1


        count_di_hoc = 0
        """ đứng ở lịch làm việc lấy ra tổng nhân sự nghỉ đi học """
        schedule_di_hoc = self.env["em.schedule"].search(
            [('status', 'in', ['5', '6']), ('state', 'in', ['2', '3']), ('start_date', '<=', self.start_date),
             ('end_date', '>=', self.end_date),
             ('department_id', '=', self.env.ref('shealth_all_in_one.sh_kb_dep_knhn').id)])
        employee_di_hoc_ids = set()
        for rec in schedule_di_hoc:
            for recc in rec.employee:
                if recc.id not in employee_di_hoc_ids:
                    employee_di_hoc_ids.add(recc.id)
                    for nv in physician:
                        if recc in nv.employee_id:
                            count_di_hoc += 1

        count_tang_cuong = 0
        """ đứng ở lịch làm việc lấy ra tổng nhân sự tăng cường """
        schedule_tang_cuong = self.env["em.schedule"].search(
            [('status', 'in', ['7']), ('state', 'in', ['2', '3']), ('start_date', '<=', self.start_date),
             ('end_date', '>=', self.end_date),
             ('department_id', '=', self.env.ref('shealth_all_in_one.sh_kb_dep_knhn').id)])
        employee_tang_cuong_ids = set()
        for rec in schedule_tang_cuong:
            for recc in rec.employee:
                if recc.id not in employee_tang_cuong_ids:
                    employee_tang_cuong_ids.add(recc.id)
                    for nv in physician:
                        if recc in nv.employee_id:
                            count_tang_cuong += 1

        ret_data.append({
            'kh_cu': 0,
            'kh_moi': len(inpatient_moi),
            'kh_xuat': len(inpatient_xuat),
            'kh_con': len(inpatient_moi) - len(inpatient_xuat),
            'ns': len(physician),
            'ns_di_lam': count_rested,
            'ns_off': len(physician) - count_rested,
            'ns_nghi': count_nghi,
            'ns_di_hoc': count_di_hoc,
            'ns_tang_cuong': count_tang_cuong,
            'ghi_chu': None,
        })
        return ret_data
    #lấy dữ liệu khoa gây mê hồi sức
    def _get_data_khoa_gay_me_hoi_suc(self):
        ret_data = []
        """ Đứng ở phiếu PTTT tìm ra tất cả bản ghi có trạng thái đang nhập viện, lấy ra bệnh nhân cũ """
        inpatient_cu = self.env['sh.medical.surgery'].search([('surgery_date', '<', self.start_datetime), ('state', 'in', ['Confirmed', 'In Progress']), ('department', '=', self.env.ref('shealth_all_in_one.sh_surgeries_dep_knhn').id)])


        """ Đứng ở phiếu PTTT tìm ra tất cả bản ghi có trạng thái đang nhập viện, lấy ra bệnh nhân mới """
        inpatient_moi = self.env['sh.medical.surgery'].search([('surgery_date', '>=', self.start_datetime), ('surgery_date', '<=', self.end_datetime), ('department', '=', self.env.ref('shealth_all_in_one.sh_surgeries_dep_knhn').id)])


        """ Đứng ở phiếu PTTT tìm ra tất cả bản ghi có trạng thái đang nhập viện, đã nhập viện, lấy ra bệnh nhân xuất viện """
        inpatient_xuat = self.env['sh.medical.surgery'].search(
            [('surgery_end_date', '>=', self.start_datetime), ('surgery_end_date', '<=', self.end_datetime), ('department', '=', self.env.ref('shealth_all_in_one.sh_surgeries_dep_knhn').id)])

        """ Đứng ở Thông tin bác sĩ tìm ra tất cả bản ghi """
        physician = self.env['sh.medical.physician'].search([('speciality', 'in', [4, 21]), ('department', '=', self.env.ref('shealth_all_in_one.sh_surgeries_dep_knhn').id), ('status', '=', False)])


        count_rested = 0
        """ đứng ở lịch làm việc lấy ra tổng nhân sự đi làm """
        schedule_rested = self.env["em.schedule"].search([('status', 'in', ['2', '4', '7']), ('state', 'in', ['2', '3']),
                ('department_id', '=', self.env.ref('shealth_all_in_one.sh_surgeries_dep_knhn').id), ('start_date', '<=', self.start_date), ('end_date', '>=', self.end_date)])
        seen_employee_ids = set()
        for rec in schedule_rested:
            for recc in rec.employee:
                if recc.id not in seen_employee_ids:
                    seen_employee_ids.add(recc.id)
                    for nv in physician:
                        if recc in nv.employee_id:
                            count_rested +=1

        count_nghi = 0
        """ đứng ở lịch làm việc lấy ra tổng nhân sự nghỉ dài hạn """
        schedule_nghi = self.env["em.schedule"].search([('status', 'in', ['3', '9']), ('state', 'in', ['2', '3']), ('start_date', '<=', self.start_date),
                ('end_date', '>=', self.end_date), ('department_id', '=', self.env.ref('shealth_all_in_one.sh_surgeries_dep_knhn').id)])
        employee_nghi_ids = set()
        for rec in schedule_nghi:
            for recc in rec.employee:
                if recc.id not in employee_nghi_ids:
                    employee_nghi_ids.add(recc.id)
                    for nv in physician:
                        if recc in nv.employee_id:
                            count_nghi += 1

        count_di_hoc = 0
        """ đứng ở lịch làm việc lấy ra tổng nhân sự nghỉ đi học """
        schedule_di_hoc = self.env["em.schedule"].search(
            [('status', 'in', ['5', '6']), ('state', 'in', ['2', '3']), ('start_date', '<=', self.start_date),
             ('end_date', '>=', self.end_date),
             ('department_id', '=', self.env.ref('shealth_all_in_one.sh_surgeries_dep_knhn').id)])
        employee_di_hoc_ids = set()
        for rec in schedule_di_hoc:
            for recc in rec.employee:
                if recc.id not in employee_di_hoc_ids:
                    employee_di_hoc_ids.add(recc.id)
                    for nv in physician:
                        if recc in nv.employee_id:
                            count_di_hoc += 1

        count_tang_cuong = 0
        """ đứng ở lịch làm việc lấy ra tổng nhân sự tăng cường """
        schedule_tang_cuong = self.env["em.schedule"].search(
            [('status', 'in', ['7']), ('state', 'in', ['2', '3']), ('start_date', '<=', self.start_date),
             ('end_date', '>=', self.end_date),
             ('department_id', '=', self.env.ref('shealth_all_in_one.sh_surgeries_dep_knhn').id)])
        employee_tang_cuong_ids = set()
        for rec in schedule_tang_cuong:
            for recc in rec.employee:
                if recc.id not in employee_tang_cuong_ids:
                    employee_tang_cuong_ids.add(recc.id)
                    for nv in physician:
                        if recc in nv.employee_id:
                            count_tang_cuong += 1

        ret_data.append({
            'kh_cu': len(inpatient_cu),
            'kh_moi': len(inpatient_moi),
            'kh_xuat': len(inpatient_xuat),
            'kh_con': len(inpatient_cu) + len(inpatient_moi) - len(inpatient_xuat),
            'ns': len(physician),
            'ns_di_lam': count_rested,
            'ns_off': len(physician) - count_rested,
            'ns_nghi': count_nghi,
            'ns_di_hoc': count_di_hoc,
            'ns_tang_cuong': count_tang_cuong,
            'ghi_chu': None,
        })
        return ret_data

    #lấy dữ liệu khoa PTTM
    def _get_data_khoa_pttm(self):
        ret_data = []

        """ Đứng ở bệnh nhân lưu tìm ra tất cả bản ghi có trạng thái đang nhập viện, lấy ra bệnh nhân cũ """
        inpatient_cu = self.env['sh.medical.inpatient'].search([('admission_date', '<', self.start_datetime), ('discharge_date', '>=', self.end_datetime), ('ward', '=', 6)])

        """ Đứng ở bệnh nhân lưu tìm ra tất cả bản ghi có trạng thái đang nhập viện, lấy ra bệnh nhân mới """
        inpatient_moi = self.env['sh.medical.inpatient'].search([('admission_date', '>=', self.start_datetime), ('admission_date', '<=', self.end_datetime), ('ward', '=', 6)])


        """ Đứng ở bệnh nhân lưu tìm ra tất cả bản ghi có trạng thái đang nhập viện, đã nhập viện, lấy ra bệnh nhân xuất viện """
        inpatient_xuat = self.env['sh.medical.inpatient'].search(
            [('discharge_date', '>=', self.start_datetime), ('discharge_date', '<=', self.end_datetime), ('ward', '=', 6)])


        """ Đứng ở Thông tin bác sĩ tìm ra tất cả bản ghi """
        physician = self.env['sh.medical.physician'].search([('speciality', 'in', [4, 21]), ('department', '=', 6), ('status', '=', False)])

        count_rested = 0
        """ đứng ở lịch làm việc lấy ra tổng nhân sự đi làm """
        schedule_rested = self.env["em.schedule"].search([('status', 'in', ['2', '4', '7']), ('state', 'in', ['2', '3']),
                ('department_id', '=', 6), ('start_date', '<=', self.start_date), ('end_date', '>=', self.end_date)])
        seen_employee_ids = set()
        for rec in schedule_rested:
            for recc in rec.employee:
                if recc.id not in seen_employee_ids:
                    seen_employee_ids.add(recc.id)
                    for nv in physician:
                        if recc in nv.employee_id:
                            count_rested +=1

        count_nghi = 0
        """ đứng ở lịch làm việc lấy ra tổng nhân sự nghỉ dài hạn """
        schedule_nghi = self.env["em.schedule"].search([('status', 'in', ['3', '9']), ('state', 'in', ['2', '3']), ('start_date', '<=', self.start_date),
                ('end_date', '>=', self.end_date), ('department_id', '=', 6)])
        employee_nghi_ids = set()
        for rec in schedule_nghi:
            for recc in rec.employee:
                if recc.id not in employee_nghi_ids:
                    employee_nghi_ids.add(recc.id)
                    for nv in physician:
                        if recc in nv.employee_id:
                            count_nghi += 1

        count_di_hoc = 0
        """ đứng ở lịch làm việc lấy ra tổng nhân sự nghỉ đi học """
        schedule_di_hoc = self.env["em.schedule"].search(
            [('status', 'in', ['5', '6']), ('state', 'in', ['2', '3']), ('start_date', '<=', self.start_date),
             ('end_date', '>=', self.end_date),('department_id', '=', 6)])
        employee_di_hoc_ids = set()
        for rec in schedule_di_hoc:
            for recc in rec.employee:
                if recc.id not in employee_di_hoc_ids:
                    employee_di_hoc_ids.add(recc.id)
                    for nv in physician:
                        if recc in nv.employee_id:
                            count_di_hoc += 1

        count_tang_cuong = 0
        """ đứng ở lịch làm việc lấy ra tổng nhân sự tăng cường """
        schedule_tang_cuong = self.env["em.schedule"].search(
            [('status', 'in', ['7']), ('state', 'in', ['2', '3']), ('start_date', '<=', self.start_date),
             ('end_date', '>=', self.end_date),
             ('department_id', '=', 6)])
        employee_tang_cuong_ids = set()
        for rec in schedule_tang_cuong:
            for recc in rec.employee:
                if recc.id not in employee_tang_cuong_ids:
                    employee_tang_cuong_ids.add(recc.id)
                    for nv in physician:
                        if recc in nv.employee_id:
                            count_tang_cuong += 1

        ret_data.append({
            'kh_cu': len(inpatient_cu),
            'kh_moi': len(inpatient_moi),
            'kh_xuat': len(inpatient_xuat),
            'kh_con': len(inpatient_cu) + len(inpatient_moi) - len(inpatient_xuat),
            'ns': len(physician),
            'ns_di_lam': count_rested,
            'ns_off': len(physician) - count_rested,
            'ns_nghi': count_nghi,
            'ns_di_hoc': count_di_hoc,
            'ns_tang_cuong': count_tang_cuong,
            'ghi_chu': None,
        })
        return ret_data

    # lấy dữ liệu khoa răng hàm mặt
    def _get_data_khoa_rang_ham_mat(self):
        ret_data = []
        # """ Đứng ở bệnh nhân lưu tìm ra tất cả bản ghi có trạng thái đang nhập viện, lấy ra bệnh nhân cũ """
        # inpatient_cu = self.env['sh.medical.specialty'].search([('services_date', '<', self.start_date), ('state', 'in', ['Confirmed', 'In Progress']), ('department', '=', self.env.ref('shealth_all_in_one.sh_rhm_dep_knhn').id)])

        """ Đứng ở bệnh nhân lưu tìm ra tất cả bản ghi có trạng thái đang nhập viện, lấy ra bệnh nhân mới """
        inpatient_moi = self.env['sh.medical.specialty'].search([('services_date', '>=', self.start_datetime), ('services_date', '<=', self.end_datetime), ('department', '=', self.env.ref('shealth_all_in_one.sh_rhm_dep_knhn').id)])

        """ Đứng ở bệnh nhân lưu tìm ra tất cả bản ghi có trạng thái đang nhập viện, đã nhập viện, lấy ra bệnh nhân xuất viện """
        inpatient_xuat = self.env['sh.medical.specialty'].search(
            [('services_end_date', '>=', self.start_datetime), ('services_end_date', '<=', self.end_datetime), ('department', '=', self.env.ref('shealth_all_in_one.sh_rhm_dep_knhn').id)])

        """ Đứng ở Thông tin bác sĩ tìm ra tất cả bản ghi """
        physician = self.env['sh.medical.physician'].search([('speciality', 'in', [4, 21]), ('department', '=', self.env.ref('shealth_all_in_one.sh_rhm_dep_knhn').id), ('status', '=', False)])

        count_rested = 0
        """ đứng ở lịch làm việc lấy ra tổng nhân sự đi làm """
        schedule_rested = self.env["em.schedule"].search([('status', 'in', ['2', '4', '7']), ('state', 'in', ['2', '3']),
                ('department_id', '=', self.env.ref('shealth_all_in_one.sh_rhm_dep_knhn').id), ('start_date', '<=', self.start_date), ('end_date', '>=', self.end_date)])
        seen_employee_ids = set()
        for rec in schedule_rested:
            for recc in rec.employee:
                if recc.id not in seen_employee_ids:
                    seen_employee_ids.add(recc.id)
                    for nv in physician:
                        if recc in nv.employee_id:
                            count_rested +=1

        count_nghi = 0
        """ đứng ở lịch làm việc lấy ra tổng nhân sự nghỉ dài hạn """
        schedule_nghi = self.env["em.schedule"].search([('status', 'in', ['3', '9']), ('state', 'in', ['2', '3']), ('start_date', '<=', self.start_date),
                ('end_date', '>=', self.end_date), ('department_id', '=', self.env.ref('shealth_all_in_one.sh_rhm_dep_knhn').id)])
        employee_nghi_ids = set()
        for rec in schedule_nghi:
            for recc in rec.employee:
                if recc.id not in employee_nghi_ids:
                    employee_nghi_ids.add(recc.id)
                    for nv in physician:
                        if recc in nv.employee_id:
                            count_nghi += 1

        count_di_hoc = 0
        """ đứng ở lịch làm việc lấy ra tổng nhân sự nghỉ đi học """
        schedule_di_hoc = self.env["em.schedule"].search(
            [('status', 'in', ['5', '6']), ('state', 'in', ['2', '3']), ('start_date', '<=', self.start_date),
             ('end_date', '>=', self.end_date),
             ('department_id', '=', self.env.ref('shealth_all_in_one.sh_rhm_dep_knhn').id)])
        employee_di_hoc_ids = set()
        for rec in schedule_di_hoc:
            for recc in rec.employee:
                if recc.id not in employee_di_hoc_ids:
                    employee_di_hoc_ids.add(recc.id)
                    for nv in physician:
                        if recc in nv.employee_id:
                            count_di_hoc += 1

        count_tang_cuong = 0
        """ đứng ở lịch làm việc lấy ra tổng nhân sự tăng cường """
        schedule_tang_cuong = self.env["em.schedule"].search(
            [('status', 'in', ['7']), ('state', 'in', ['2', '3']), ('start_date', '<=', self.start_date),
             ('end_date', '>=', self.end_date),
             ('department_id', '=', self.env.ref('shealth_all_in_one.sh_rhm_dep_knhn').id)])
        employee_tang_cuong_ids = set()
        for rec in schedule_tang_cuong:
            for recc in rec.employee:
                if recc.id not in employee_tang_cuong_ids:
                    employee_tang_cuong_ids.add(recc.id)
                    for nv in physician:
                        if recc in nv.employee_id:
                            count_tang_cuong += 1

        ret_data.append({
            'kh_cu': 0,
            'kh_moi': len(inpatient_moi),
            'kh_xuat': len(inpatient_xuat),
            'kh_con': len(inpatient_moi) - len(inpatient_xuat),
            'ns': len(physician),
            'ns_di_lam': count_rested,
            'ns_off': len(physician) - count_rested,
            'ns_nghi': count_nghi,
            'ns_di_hoc': count_di_hoc,
            'ns_tang_cuong': count_tang_cuong,
            'ghi_chu': None,
        })
        return ret_data

    #lấy dữ liệu khoa da liễu
    def _get_data_khoa_da_lieu(self):
        ret_data = []
        # """ Đứng ở bệnh nhân lưu tìm ra tất cả bản ghi có trạng thái đang nhập viện, lấy ra bệnh nhân cũ """
        # inpatient_cu = self.env['sh.medical.specialty'].search([('services_date', '<', self.start_date), ('state', 'in', ['Confirmed', 'In Progress']), ('department', '=', self.env.ref('shealth_all_in_one.sh_dalieu_dep_knhn').id)])

        """ Đứng ở bệnh nhân lưu tìm ra tất cả bản ghi có trạng thái đang nhập viện, lấy ra bệnh nhân mới """
        inpatient_moi = self.env['sh.medical.specialty'].search([('services_date', '>=', self.start_datetime), ('services_date', '<=', self.end_datetime), ('department', '=', self.env.ref('shealth_all_in_one.sh_dalieu_dep_knhn').id)])

        """ Đứng ở bệnh nhân lưu tìm ra tất cả bản ghi có trạng thái đang nhập viện, đã nhập viện, lấy ra bệnh nhân xuất viện """
        inpatient_xuat = self.env['sh.medical.specialty'].search(
            [('services_end_date', '>=', self.start_datetime), ('services_end_date', '<=', self.end_datetime), ('department', '=', self.env.ref('shealth_all_in_one.sh_dalieu_dep_knhn').id)])


        """ Đứng ở Thông tin bác sĩ tìm ra tất cả bản ghi """
        physician = self.env['sh.medical.physician'].search([('speciality', 'in', [4, 21]), ('department', '=', self.env.ref('shealth_all_in_one.sh_dalieu_dep_knhn').id), ('status', '=', False)])


        count_rested = 0
        """ đứng ở lịch làm việc lấy ra tổng nhân sự đi làm """
        schedule_rested = self.env["em.schedule"].search([('status', 'in', ['2', '4', '7']), ('state', 'in', ['2', '3']),
                ('department_id', '=', self.env.ref('shealth_all_in_one.sh_dalieu_dep_knhn').id), ('start_date', '<=', self.start_date), ('end_date', '>=', self.end_date)])
        seen_employee_ids = set()
        for rec in schedule_rested:
            for recc in rec.employee:
                if recc.id not in seen_employee_ids:
                    seen_employee_ids.add(recc.id)
                    for nv in physician:
                        if recc in nv.employee_id:
                            count_rested +=1

        count_nghi = 0
        """ đứng ở lịch làm việc lấy ra tổng nhân sự nghỉ dài hạn """
        schedule_nghi = self.env["em.schedule"].search([('status', 'in', ['3', '9']), ('state', 'in', ['2', '3']), ('start_date', '<=', self.start_date),
                ('end_date', '>=', self.end_date), ('department_id', '=', self.env.ref('shealth_all_in_one.sh_dalieu_dep_knhn').id)])
        employee_nghi_ids = set()
        for rec in schedule_nghi:
            for recc in rec.employee:
                if recc.id not in employee_nghi_ids:
                    employee_nghi_ids.add(recc.id)
                    for nv in physician:
                        if recc in nv.employee_id:
                            count_nghi += 1

        count_di_hoc = 0
        """ đứng ở lịch làm việc lấy ra tổng nhân sự nghỉ đi học """
        schedule_di_hoc = self.env["em.schedule"].search(
            [('status', 'in', ['5', '6']), ('state', 'in', ['2', '3']), ('start_date', '<=', self.start_date),
             ('end_date', '>=', self.end_date),
             ('department_id', '=', self.env.ref('shealth_all_in_one.sh_dalieu_dep_knhn').id)])
        employee_di_hoc_ids = set()
        for rec in schedule_di_hoc:
            for recc in rec.employee:
                if recc.id not in employee_di_hoc_ids:
                    employee_di_hoc_ids.add(recc.id)
                    for nv in physician:
                        if recc in nv.employee_id:
                            count_di_hoc += 1

        count_tang_cuong = 0
        """ đứng ở lịch làm việc lấy ra tổng nhân sự tăng cường """
        schedule_tang_cuong = self.env["em.schedule"].search(
            [('status', 'in', ['7']), ('state', 'in', ['2', '3']), ('start_date', '<=', self.start_date),
             ('end_date', '>=', self.end_date),
             ('department_id', '=', self.env.ref('shealth_all_in_one.sh_dalieu_dep_knhn').id)])
        employee_tang_cuong_ids = set()
        for rec in schedule_tang_cuong:
            for recc in rec.employee:
                if recc.id not in employee_tang_cuong_ids:
                    employee_tang_cuong_ids.add(recc.id)
                    for nv in physician:
                        if recc in nv.employee_id:
                            count_tang_cuong += 1

        ret_data.append({
            'kh_cu': 0,
            'kh_moi': len(inpatient_moi),
            'kh_xuat': len(inpatient_xuat),
            'kh_con': len(inpatient_moi) - len(inpatient_xuat),
            'ns': len(physician),
            'ns_di_lam': count_rested,
            'ns_off': len(physician) - count_rested,
            'ns_nghi': count_nghi,
            'ns_di_hoc': count_di_hoc,
            'ns_tang_cuong': count_tang_cuong,
            'ghi_chu': None,
        })
        return ret_data

    # lấy dữ liệu khoa cận lâm sàn
    def _get_data_khoa_can_lam_san(self):
        ret_data = []
        """ Đứng ở bệnh nhân lưu tìm ra tất cả bản ghi có trạng thái đang nhập viện, lấy ra bệnh nhân mới """
        inpatient_moi = self.env['sh.medical.lab.test'].search([('date_analysis', '>=', self.start_datetime), ('date_analysis', '<=', self.end_datetime), ('department', '=', self.env.ref('shealth_all_in_one.sh_labtest_dep_knhn').id)])

        """ Đứng ở bệnh nhân lưu tìm ra tất cả bản ghi có trạng thái đang nhập viện, đã nhập viện, lấy ra bệnh nhân xuất viện """
        inpatient_xuat = self.env['sh.medical.lab.test'].search(
            [('date_done', '>=', self.start_datetime), ('date_done', '<=', self.end_datetime), ('department', '=', self.env.ref('shealth_all_in_one.sh_labtest_dep_knhn').id)])

        """ Đứng ở Thông tin bác sĩ tìm ra tất cả bản ghi """
        physician = self.env['sh.medical.physician'].search([('speciality', 'in', [4, 21]), ('department', '=', self.env.ref('shealth_all_in_one.sh_labtest_dep_knhn').id), ('status', '=', False)])

        count_rested = 0
        """ đứng ở lịch làm việc lấy ra tổng nhân sự đi làm """
        schedule_rested = self.env["em.schedule"].search([('status', 'in', ['2', '4', '7']), ('state', 'in', ['2', '3']),
                ('department_id', '=', self.env.ref('shealth_all_in_one.sh_labtest_dep_knhn').id), ('start_date', '<=', self.start_date), ('end_date', '>=', self.end_date)])
        seen_employee_ids = set()
        for rec in schedule_rested:
            for recc in rec.employee:
                if recc.id not in seen_employee_ids:
                    seen_employee_ids.add(recc.id)
                    for nv in physician:
                        if recc in nv.employee_id:
                            count_rested +=1

        count_nghi = 0
        """ đứng ở lịch làm việc lấy ra tổng nhân sự nghỉ dài hạn """
        schedule_nghi = self.env["em.schedule"].search([('status', 'in', ['3', '9']), ('state', 'in', ['2', '3']), ('start_date', '<=', self.start_date),
                ('end_date', '>=', self.end_date), ('department_id', '=', self.env.ref('shealth_all_in_one.sh_labtest_dep_knhn').id)])
        employee_nghi_ids = set()
        for rec in schedule_nghi:
            for recc in rec.employee:
                if recc.id not in employee_nghi_ids:
                    employee_nghi_ids.add(recc.id)
                    for nv in physician:
                        if recc in nv.employee_id:
                            count_nghi += 1

        count_di_hoc = 0
        """ đứng ở lịch làm việc lấy ra tổng nhân sự nghỉ đi học """
        schedule_di_hoc = self.env["em.schedule"].search(
            [('status', 'in', ['5', '6']), ('state', 'in', ['2', '3']), ('start_date', '<=', self.start_date),
             ('end_date', '>=', self.end_date),
             ('department_id', '=', self.env.ref('shealth_all_in_one.sh_labtest_dep_knhn').id)])
        employee_di_hoc_ids = set()
        for rec in schedule_di_hoc:
            for recc in rec.employee:
                if recc.id not in employee_di_hoc_ids:
                    employee_di_hoc_ids.add(recc.id)
                    for nv in physician:
                        if recc in nv.employee_id:
                            count_di_hoc += 1

        count_tang_cuong = 0
        """ đứng ở lịch làm việc lấy ra tổng nhân sự tăng cường """
        schedule_tang_cuong = self.env["em.schedule"].search(
            [('status', 'in', ['7']), ('state', 'in', ['2', '3']), ('start_date', '<=', self.start_date),
             ('end_date', '>=', self.end_date),
             ('department_id', '=', self.env.ref('shealth_all_in_one.sh_labtest_dep_knhn').id)])
        employee_tang_cuong_ids = set()
        for rec in schedule_tang_cuong:
            for recc in rec.employee:
                if recc.id not in employee_tang_cuong_ids:
                    employee_tang_cuong_ids.add(recc.id)
                    for nv in physician:
                        if recc in nv.employee_id:
                            count_tang_cuong += 1

        ret_data.append({
            'kh_cu': 0,
            'kh_moi': len(inpatient_moi),
            'kh_xuat': len(inpatient_xuat),
            'kh_con': len(inpatient_moi) - len(inpatient_xuat),
            'ns': len(physician),
            'ns_di_lam': count_rested,
            'ns_off': len(physician) - count_rested,
            'ns_nghi': count_nghi,
            'ns_di_hoc': count_di_hoc,
            'ns_tang_cuong': count_tang_cuong,
            'ghi_chu': None,
        })
        return ret_data

    def create_report(self):

        datas = self._get_data_khoa_kham_benh()
        datas_1 = self._get_data_khoa_gay_me_hoi_suc()
        datas_2 = self._get_data_khoa_pttm()
        datas_3 = self._get_data_khoa_rang_ham_mat()
        datas_4 = self._get_data_khoa_da_lieu()
        datas_5 = self._get_data_khoa_can_lam_san()

        # in du lieu
        report_brand_overview_attachment = self.env['ir.attachment'].browse(
            self.env.ref('em_calendar.bao_cao_nhan_luc_attachment').id)
        decode = base64.b64decode(report_brand_overview_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        line_font = Font(name='Times New Roman', size=12)

        # ws['I3'].value = self.end_datetime.strftime('Ngày %d tháng %m năm %Y')
        ws['I3'].value = self.end_datetime.strftime('Ngay %d thang %m nam %Y')

        key_col_list = list(range(1, 12))
        key_list = [
            'kh_cu',
            'kh_moi',
            'kh_xuat',
            'kh_con',
            'ns',
            'ns_di_lam',
            'ns_off',
            'ns_nghi',
            'ns_di_hoc',
            'ns_tang_cuong'
        ]
        row = 6
        for data in datas:
            for col, k in zip(key_col_list, key_list):
                beforeCell = ws.cell(4, col)
                beforeCell.font = Font(name='Times New Roman', size=12, color='000000')
                cell = ws.cell(row, col + 3)
                cell.value = data[k]
                cell.font = line_font
                cell.border = all_border_thin
                cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 1
        row = 7
        for data in datas_1:
            for col, k in zip(key_col_list, key_list):
                cell = ws.cell(row, col + 3)
                cell.value = data[k]
                cell.font = line_font
                cell.border = all_border_thin
                cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 1
        row = 8
        for data in datas_2:
            for col, k in zip(key_col_list, key_list):
                cell = ws.cell(row, col + 3)
                cell.value = data[k]
                cell.font = line_font
                cell.border = all_border_thin
                cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 1
        row = 9
        for data in datas_3:
            for col, k in zip(key_col_list, key_list):
                cell = ws.cell(row, col + 3)
                cell.value = data[k]
                cell.font = line_font
                cell.border = all_border_thin
                cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 1
        row = 10
        for data in datas_4:
            for col, k in zip(key_col_list, key_list):
                cell = ws.cell(row, col + 3)
                cell.value = data[k]
                cell.font = line_font
                cell.border = all_border_thin
                cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 1
        row = 11
        for data in datas_5:
            for col, k in zip(key_col_list, key_list):
                cell = ws.cell(row, col + 3)
                cell.value = data[k]
                cell.font = line_font
                cell.border = all_border_thin
                cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 1

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'Báo cáo nhân lực',
            'datas_fname': 'bao_cao_nhan_luc.xlsx',
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })
        url = "/web/content/?model=ir.attachment&id=%s&filename_field=datas_fname&field=datas&download=true&filename=Báo cáo nhân lực.xlsx" \
              % (attachment.id)
        return {'name': 'Báo cáo nhân lực',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
                }


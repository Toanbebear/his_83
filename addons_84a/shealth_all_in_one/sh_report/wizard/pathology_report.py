# -*- coding: utf-8 -*-

from odoo import fields, api, models, _
from odoo.exceptions import ValidationError, UserError
from datetime import date, datetime, time
from calendar import monthrange
from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment
import base64
from io import BytesIO
from dateutil.relativedelta import relativedelta
from pytz import timezone

import logging

_logger = logging.getLogger(__name__)

thin = borders.Side(style='thin')
double = borders.Side(style='double')
all_border_thin = borders.Border(thin, thin, thin, thin)


class PathologyReport(models.TransientModel):
    _name = 'pathology.report'
    _description = 'Báo cáo tình hình bệnh tật'

    start_date = fields.Date('Start date', default=date.today().replace(day=1))
    end_date = fields.Date('End date')
    # convert date to datetime for search domain, should be removed if using datetime directly
    start_datetime = fields.Datetime('Start datetime', compute='_compute_datetime')
    end_datetime = fields.Datetime('End datetime', compute='_compute_datetime')

    # convert date to datetime for search domain, should be removed if using datetime directly
    @api.depends('start_date', 'end_date')
    def _compute_datetime(self):
        if self.start_date and self.end_date:
            fmt = "%Y-%m-%d %H:%M:%S"
            start_datetime = datetime(self.start_date.year, self.start_date.month, self.start_date.day, 0, 0, 0)
            end_datetime = datetime(self.end_date.year, self.end_date.month, self.end_date.day, 23, 59, 59)
            now_utc = datetime.now(timezone('UTC'))
            local_tz = self.env.user.tz or 'Etc/GMT-7'
            now_timezone = now_utc.astimezone(timezone(local_tz))
            UTC_OFFSET_TIMEDELTA = datetime.strptime(now_utc.strftime(fmt), fmt) - datetime.strptime(
                now_timezone.strftime(fmt), fmt)
            utc_start_datetime = start_datetime + UTC_OFFSET_TIMEDELTA
            utc_end_datetime = end_datetime + UTC_OFFSET_TIMEDELTA
            self.start_datetime = utc_start_datetime
            self.end_datetime = utc_end_datetime

    @api.onchange('start_date')
    def _onchange_start_date(self):
        if self.start_date:
            if self.start_date.month == fields.date.today().month:
                self.end_date = fields.date.today()
            else:
                self.end_date = date(self.start_date.year, self.start_date.month,
                                     monthrange(self.start_date.year, self.start_date.month)[1])

    @api.multi
    @api.constrains('start_date', 'end_date')
    def check_dates(self):
        for record in self:
            start_date = fields.Date.from_string(record.start_date)
            end_date = fields.Date.from_string(record.end_date)
            if start_date > end_date:
                raise ValidationError(
                    _("End Date cannot be set before Start Date."))

    def death_situation_report(self):
        datas = {}

        walkin_data = self.env['sh.medical.appointment.register.walkin'].read_group([('date', '>=', self.start_datetime),
                                                               ('date', '<=', self.end_datetime),('pathology','!=',False)], fields=['pathology'], groupby=['pathology'])

        for walkin in walkin_data:
            pathology_detail = self.env['sh.medical.pathology'].browse(walkin['pathology'][0])

            walkin_data_female = self.env['sh.medical.appointment.register.walkin'].search_count(walkin['__domain']+[('sex','=','Female')])

            if not datas.get(str(pathology_detail.index_pathology)):
                datas[str(pathology_detail.index_pathology)] = {'total': walkin['pathology_count'],'female': walkin_data_female}
            else:
                datas[str(pathology_detail.index_pathology)]['total'] += walkin['pathology_count']
                datas[str(pathology_detail.index_pathology)]['female'] += walkin_data_female

        pathology_attachment = self.env['ir.attachment'].browse(
            self.env.ref('shealth_all_in_one.bao_cao_benh_icd10').id)

        decode = base64.b64decode(pathology_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        # ws['b2'].value = self.env.user.company_id.name
        ws['b4'].value = self.end_date.strftime('Tháng %m Năm %Y')
        ws['l347'].value = datetime.now().strftime('Ngày %d Tháng %m Năm %Y')

        for row in ws.iter_rows(13, 346, 2, 2):
            for cell in row:
                if cell.value != '' and datas.get(str(cell.value)):
                    r = cell.row
                    ws.cell(r, 5).value = datas.get(str(cell.value))['total']
                    ws.cell(r, 6).value = datas.get(str(cell.value))['female']

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({'name': 'MS report',
                                                              'datas_fname': 'BC_TINH_HINH_BENH_TAT_THEO_MA_ICD10.xlsx',
                                                              'datas': report,
                                                              'res_model': 'temp.creation',
                                                              'public': True})

        return {'name': 'Báo cáo Tình hình bệnh tật theo mã icd10',
                'type': 'ir.actions.act_window',
                'res_model': 'temp.wizard',
                'view_mode': 'form',
                'view_type': 'form',
                'target': 'inline',
                'view_id': self.env.ref('ms_templates.report_wizard').id,
                'context': {'attachment_id': attachment.id}}

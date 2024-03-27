##############################################################################
#    Copyright (C) 2018 shealth (<http://scigroup.com.vn/>). All Rights Reserved
#    shealth, Hospital Management Solutions

# Odoo Proprietary License v1.0
#
# This software and associated files (the "Software") may only be used (executed,
# modified, executed after modifications) if you have purchased a valid license
# from the authors, typically via Odoo Apps, shealth.in, openerpestore.com, or if you have received a written
# agreement from the authors of the Software.
#
# You may develop Odoo modules that use the Software as a library (typically
# by depending on it, importing it and using its resources), but without copying
# any source code or material from the Software. You may distribute those
# modules under the license of your choice, provided that this license is
# compatible with the terms of the Odoo Proprietary License (For example:
# LGPL, MIT, or proprietary licenses similar to this one).
#
# It is forbidden to publish, distribute, sublicense, or sell copies of the Software
# or modified copies of the Software.
#
# The above copyright notice and this permission notice must be included in all
# copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT.
# IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM,
# DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE,
# ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER
# DEALINGS IN THE SOFTWARE.

##############################################################################

import logging
import pytz
from odoo import api, fields, models, _
from odoo.exceptions import UserError
from odoo.tools.translate import _
from odoo.exceptions import UserError, AccessError, ValidationError, Warning
import time
import datetime
from datetime import timedelta
from odoo.osv import expression
from odoo.tools import pycompat

from openpyxl import load_workbook
from openpyxl.styles import Font, borders, Alignment, PatternFill
import base64
from io import BytesIO

_logger = logging.getLogger(__name__)

thin = borders.Side(style='thin')
thick = borders.Side(style='medium')
dotted = borders.Side(style='hair')
gray_thin = borders.Side(style='thin', color='808080')
all_border_thin = borders.Border(left=thin, right=thin, top=thin, bottom=thin)
all_border_gray = borders.Border(left=gray_thin, right=gray_thin, top=gray_thin, bottom=gray_thin)
dotted_top_bot = borders.Border(left=thin, right=thin, top=dotted, bottom=dotted)
thick_dotted_top_bot = borders.Border(left=thick, right=thick, top=dotted, bottom=dotted)
all_border_thick = borders.Border(left=thick, right=thick, top=thick, bottom=thick)
center_alm = Alignment(horizontal='center', vertical='center')

# Foreign Management
class SHealthForeign(models.Model):
    _name = 'sh.foreign'
    _description = 'Thông tin ngoại kiều'

    name = fields.Char(size=256, string='Tên', required=True, help='Tên ngoại kiều')

# Family Management
class SHealthFamily(models.Model):
    _name = 'sh.medical.patient.family'
    _description = 'Information about family of patient'

    FAMILY_RELATION = [
                ('Father', 'Father'),
                ('Mother', 'Mother'),
                ('Brother', 'Brother'),
                ('Sister', 'Sister'),
                ('Wife', 'Wife'),
                ('Husband', 'Husband'),
                ('Aunt', 'Aunt'),
                ('Uncle', 'Uncle'),
                ('Nephew', 'Nephew'),
                ('Niece', 'Niece'),
                ('Cousin', 'Cousin'),
                ('Relative', 'Relative'),
                ('Other', 'Other'),
    ]

    name = fields.Char(size=256, string='Name', required=True, help='Family Member Name')
    phone = fields.Char(string='Phone', help='Family Member Phone')
    relation = fields.Selection(FAMILY_RELATION, string='Relation', help="Family Relation", index=True)
    age = fields.Selection([(num, str(num)) for num in reversed(range((datetime.datetime.now().year)-80, datetime.datetime.now().year))],'Năm sinh')
    address = fields.Text(string='Địa chỉ liên hệ', help='Địa chỉ liên hệ')
    deceased = fields.Boolean(string='Deceased?',help="Mark if the family member has died")
    patient_id = fields.Many2one('sh.medical.patient', 'Patient', required=True, ondelete='cascade', index=True)

# Patient Management

class SHealthPatient(models.Model):
    _name = 'sh.medical.patient'
    _description = 'Information of patient'
    _inherits = {
        'res.partner': 'partner_id',
    }

    _order = "identification_code desc"

    MARITAL_STATUS = [
        ('Single', 'Single'),
        ('Married', 'Married'),
        ('Widowed', 'Widowed'),
        ('Divorced', 'Divorced'),
        ('Separated', 'Separated'),
    ]

    SEX = [
        ('Male', 'Male'),
        ('Female', 'Female'),
    ]

    BLOOD_TYPE = [
        ('A', 'A'),
        ('B', 'B'),
        ('AB', 'AB'),
        ('O', 'O'),
    ]

    RH = [
        ('+','+'),
        ('-','-'),
    ]

    @api.multi
    def _app_count(self):
        oe_apps = self.env['sh.medical.appointment']
        for pa in self:
            domain = [('patient', '=', pa.id)]
            app_ids = oe_apps.search(domain)
            apps = oe_apps.browse(app_ids)
            app_count = 0
            for ap in apps:
                app_count+=1
            pa.app_count = app_count
        return True

    @api.multi
    def _prescription_count(self):
        oe_pres = self.env['sh.medical.prescription']
        for pa in self:
            domain = [('patient', '=', pa.id)]
            pres_ids = oe_pres.search(domain)
            pres = oe_pres.browse(pres_ids)
            pres_count = 0
            for pr in pres:
                pres_count+=1
            pa.prescription_count = pres_count
        return True

    @api.multi
    def _admission_count(self):
        oe_admission = self.env['sh.medical.inpatient']
        for adm in self:
            domain = [('patient', '=', adm.id)]
            admission_ids = oe_admission.search(domain)
            admissions = oe_admission.browse(admission_ids)
            admission_count = 0
            for ad in admissions:
                admission_count+=1
            adm.admission_count = admission_count
        return True

    @api.multi
    def _vaccine_count(self):
        oe_vac = self.env['sh.medical.vaccines']
        for va in self:
            domain = [('patient', '=', va.id)]
            vec_ids = oe_vac.search(domain)
            vecs = oe_vac.browse(vec_ids)
            vecs_count = 0
            for vac in vecs:
                vecs_count+=1
            va.vaccine_count = vecs_count
        return True

    @api.multi
    def _invoice_count(self):
        oe_invoice = self.env['account.invoice']
        for inv in self:
            invoice_ids = self.env['account.invoice'].search([('patient', '=', inv.id)])
            invoices = oe_invoice.browse(invoice_ids)
            invoice_count = 0
            for inv_id in invoices:
                invoice_count+=1
            inv.invoice_count = invoice_count
        return True

    @api.multi
    def _patient_age(self):
        def compute_age_from_dates(patient_dob, patient_deceased, patient_dod):
            now = datetime.datetime.now()
            if (patient_dob):
                dob = datetime.datetime.strptime(patient_dob.strftime('%Y-%m-%d'), '%Y-%m-%d')
                if patient_deceased:
                    dod = datetime.datetime.strptime(patient_dod.strftime('%Y-%m-%d'), '%Y-%m-%d')
                    delta = dod - dob
                    deceased = " (deceased)"
                    # years_months_days = _(str(delta.days // 365) + " years " + str(delta.days % 365) + " days" + deceased)
                    # years_months_days = _("%s tuổi %s ngày %s"%(str(delta.days // 365),str(delta.days%365),deceased))
                    years_months_days = _("%s tuổi "%(str(delta.days // 365)))
                else:
                    delta = now - dob
                    # years_months_days = _(str(delta.days // 365) + " years " + str(delta.days % 365) + " days")
                    # years_months_days = _("%s tuổi %s ngày"%(str(delta.days // 365),str(delta.days%365)))
                    years_months_days = _("%s tuổi"%(str(delta.days // 365)))
            else:
                years_months_days = _("No DoB !")

            return years_months_days

        for patient_data in self:
            patient_data.age = compute_age_from_dates(patient_data.dob, patient_data.deceased, patient_data.dod)
        return True

    partner_id = fields.Many2one('res.partner', string='Related Partner', required=True, ondelete='cascade', help='Partner-related data of the patient')
    family = fields.One2many('sh.medical.patient.family', 'patient_id', string='Family')
    ssn = fields.Char(size=256, string='SSN')
    current_insurance = fields.Many2one('sh.medical.insurance', string="Insurance", domain="[('patient','=', active_id),('state','=','Active')]", help="Insurance information. You may choose from the different insurances belonging to the patient")
    doctor = fields.Many2one('sh.medical.physician', string='Family Physician', help="Current primary care physician / family doctor", domain=[('is_pharmacist','=',False)])
    dob = fields.Date(string='Date of Birth')
    age = fields.Char(compute=_patient_age, size=32, string='Patient Age', help="It shows the age of the patient in years(y), months(m) and days(d).\nIf the patient has died, the age shown is the age at time of death, the age corresponding to the date on the death certificate. It will show also \"deceased\" on the field")
    # age = fields.Char(string='Patient Age', help="It shows the age of the patient in years(y), months(m) and days(d).\nIf the patient has died, the age shown is the age at time of death, the age corresponding to the date on the death certificate. It will show also \"deceased\" on the field")
    sex = fields.Selection(SEX, string='Sex', index=True)
    marital_status = fields.Selection(MARITAL_STATUS, string='Marital Status')
    blood_type = fields.Selection(BLOOD_TYPE, string='Blood Type')
    rh = fields.Selection(RH, string='Rh')
    identification_code = fields.Char(string='Patient ID', size=256, help='Patient Identifier provided by the Health Center', copy=False)
    ethnic_group = fields.Many2one('sh.medical.ethnicity','Dân tộc', default=lambda self: self.env.ref('shealth_all_in_one.sheth_kinh').id)
    foreign = fields.Many2one('sh.foreign','Ngoại kiều')
    critical_info = fields.Text(string='Important disease, allergy or procedures information', help="Write any important information on the patient's disease, surgeries, allergies, ...")
    general_info = fields.Text(string='General Information', help="General information about the patient")
    genetic_risks = fields.Many2many('sh.medical.genetics', 'sh_genetic_risks_rel', 'patient_id', 'genetic_risk_id', string='Genetic Risks')
    deceased = fields.Boolean(string='Patient Deceased ?', help="Mark if the patient has died")
    dod = fields.Date(string='Date of Death')
    cod = fields.Many2one('sh.medical.pathology', string='Cause of Death')
    app_count = fields.Integer(compute=_app_count, string="Appointments")
    prescription_count = fields.Integer(compute=_prescription_count, string="Prescriptions")
    admission_count = fields.Integer(compute=_admission_count, string="Admission / Discharge")
    vaccine_count = fields.Integer(compute=_vaccine_count, string="Vaccines")
    invoice_count = fields.Integer(compute=_invoice_count, string="Invoices")
    # invoice_count = fields.Integer(string="Invoices")
    sh_patient_user_id = fields.Many2one('res.users', string='Responsible Odoo User')
    prescription_line = fields.One2many('sh.medical.prescription.line', 'patient', string='Medicines', readonly=True)
    prescription_ids = fields.One2many('sh.medical.prescription', 'patient', string='Prescriptions', readonly=True)

    country_id = fields.Many2one('res.country', string='Country', ondelete='restrict', default=lambda self: self.env.ref('base.vn').id)

    id_card = fields.Char('CMND/Hộ chiếu')
    function = fields.Char(string='Nghề nghiệp', default='Tự do')

    _sql_constraints = [
        ('code_sh_patient_userid_uniq', 'unique (sh_patient_user_id)', "Selected 'Responsible' user is already assigned to another patient !")
    ]

    @api.multi
    @api.returns('self', lambda value: value.id)
    def copy(self, default=None):
        if default is None:
            default = {}
            default['name'] = self.name + " (Copy)"

        return super(SHealthPatient, self).copy(default=default)

    @api.model
    def create(self, vals):
        vals['is_patient'] = True
        health_patient = super(SHealthPatient, self).create(vals)
        #TỰ ĐỘNG CẬP NHẬT MÃ TỰ TĂNG NẾU KO CÓ
        if not health_patient.identification_code:
            sequence = self.env['ir.sequence'].next_by_code('sh.medical.patient')
            health_patient.identification_code = sequence

        return health_patient

    # @api.multi
    # @api.constrains('dob')
    # def _check_dob(self):
    #     for record in self:
    #         if record.dob > fields.Date.today():
    #             raise ValidationError(_(
    #                 "Ngày sinh không thể lớn hơn ngày hiện tại!"))

    @api.onchange('dob')
    def onchange_dob(self):
        if self.dob and self.dob > fields.Date.today():
            self.dob = False
            raise ValidationError(_(
                    "Ngày sinh không thể lớn hơn ngày hiện tại!"))

    @api.onchange('state_id')
    def onchange_state(self):
        if self.state_id:
            self.country_id = self.state_id.country_id.id

    @api.onchange('foreign')
    def onchange_state(self):
        if self.foreign:
            self.ethnic_group = False

    @api.multi
    def print_patient_label(self):
        return self.env.ref('shealth_all_in_one.action_report_patient_label').report_action(self)

    @api.multi
    def name_get(self):
        res = []
        for category in self:
            res.append((category.id, _('[%s] %s - %s') % (category.identification_code, category.name[0:50], category.dob.strftime("%Y"))))
        return res

    @api.model
    def _name_search(self, name, args=None, operator='ilike', limit=100, name_get_uid=None):
        args = args or []
        domain = []
        if name:
            domain = ['|', ('name', operator, name), ('identification_code', operator, name)]
        patient = self._search(expression.AND([domain, args]), limit=limit, access_rights_uid=name_get_uid)
        return self.browse(patient).name_get()

# Physician Management

class SHealthPhysicianSpeciality(models.Model):
    _name = "sh.medical.speciality"
    _description = "Physician Speciality"

    name = fields.Char(string='Description', size=128, help="ie, Addiction Psychiatry", translate=True, required=True)
    code = fields.Char(string='Code', size=128, help="ie, ADP")

    _order = 'name'
    _sql_constraints = [
        ('code_uniq', 'unique (name)', 'The Medical Speciality code must be unique')]

class SHealthTeamRole(models.Model):
    _name = "sh.medical.team.role"
    _description = "Vai trò trong nhóm thực hiện dịch vụ"

    ROLE_TYPE = [
        ('spa', 'Spa'),
        ('laser', 'Laser'),
        ('surgery', 'Phẫu thuật'),
        ('odontology', 'Nha khoa')
    ]

    name = fields.Char(string='Description', size=128, translate=True, required=True)
    type = fields.Selection(ROLE_TYPE, string='Loại')

    _order = 'name'
    _sql_constraints = [
        ('name_uniq', 'unique (name)', 'Tên vai trò phải là duy nhất!')]

class SHealthPhysicianDegree(models.Model):
    _name = "sh.medical.degrees"
    _description = "Physicians Degrees"

    name = fields.Char(string='Degree', size=128, required=True)
    keyword = fields.Char(string='Keyword', size=128)
    full_name = fields.Char(string='Full Name', translate=True, size=128)
    education_level = fields.Selection([('saudaihoc','Sau đại học'),('daihoc','Đại học'),('trungcap','Trung Cấp'),('socap','Sơ Cấp')],string='Cấp học', translate=True,default='trungcap')
    physician_ids = fields.Many2many('sh.medical.physician', id1='degree_id', id2='physician_id', string='Physicians')

    _sql_constraints = [
        ('full_name_uniq', 'unique (name,education_level)', 'The Medical Degree must be unique')]

    @api.multi
    def name_get(self):
        res = []
        for degree in self:
            res.append((degree.id, _('%s (%s)') % (degree.name, dict(degree._fields['education_level'].selection).get(degree.education_level))))
        return res

class SHealthPhysician(models.Model):
    _name = "sh.medical.physician"
    _description = "Information about the doctor"
    _inherits={
        'hr.employee': 'employee_id',
    }

    CONSULTATION_TYPE = [
        ('Residential', 'Residential'),
        ('Visiting', 'Visiting'),
        ('Other', 'Other'),
    ]

    APPOINTMENT_TYPE = [
        ('Not on Weekly Schedule', 'Not on Weekly Schedule'),
        ('On Weekly Schedule', 'On Weekly Schedule'),
    ]

    @api.multi
    def _app_count(self):
        oe_apps = self.env['sh.medical.appointment']
        for pa in self:
            domain = [('doctor', '=', pa.id)]
            app_ids = oe_apps.search(domain)
            apps = oe_apps.browse(app_ids)
            app_count = 0
            for ap in apps:
                app_count+=1
            pa.app_count = app_count
        return True

    @api.multi
    def _prescription_count(self):
        oe_pres = self.env['sh.medical.prescription']
        for pa in self:
            domain = [('doctor', '=', pa.id)]
            pres_ids = oe_pres.search(domain)
            pres = oe_pres.browse(pres_ids)
            pres_count = 0
            for pr in pres:
                pres_count+=1
            pa.prescription_count = pres_count
        return True

    employee_id = fields.Many2one('hr.employee', string='Related Employee', required=True, ondelete='cascade', help='Employee-related data of the physician')
    institution = fields.Many2one('sh.medical.health.center', string='Institution', help="Institution where doctor works")
    department = fields.Many2many('sh.medical.health.center.ward','sh_medical_physician_department_rel','physician_id','dep_id' ,string='Department',domain="[('institution', '=', institution)]", help="Department where doctor works")
    code = fields.Char(string='Licence ID', size=128, help="Physician's License ID")
    speciality = fields.Many2one('sh.medical.speciality', string='Speciality', help="Speciality Code")
    consultancy_type = fields.Selection(CONSULTATION_TYPE, string='Consultancy Type', help="Type of Doctor's Consultancy", default=lambda *a: 'Residential')
    consultancy_price = fields.Integer(string='Consultancy Charge', help="Physician's Consultancy price")
    available_lines = fields.One2many('sh.medical.physician.line', 'physician_id', string='Physician Availability')
    degree_id = fields.Many2many('sh.medical.degrees', id1='physician_id', id2='degree_id', string='Degrees')
    app_count = fields.Integer(compute=_app_count, string="Appointments")
    prescription_count = fields.Integer(compute=_prescription_count, string="Prescriptions")
    is_pharmacist = fields.Boolean(string='Pharmacist?', default=lambda *a: False)
    sh_user_id = fields.Many2one('res.users', string='Responsible Odoo User')
    appointment_type = fields.Selection(APPOINTMENT_TYPE, string='Allow Appointment on?', default=lambda *a: 'Not on Weekly Schedule')
    contract_type = fields.Selection([('fulltime','Fulltime'),('parttime','Parttime')], string='Hợp đồng', default=lambda *a: 'fulltime')

    _sql_constraints = [
        ('code_sh_physician_userid_uniq', 'unique(sh_user_id)', "Selected 'Responsible' user is already assigned to another physician !")
    ]

    @api.multi
    def name_get(self):
        res = []
        for physician in self:
            res.append((physician.id, _('%s - %s') % (physician.name, physician.speciality.name)))
        return res

    @api.onchange('state_id')
    def onchange_state(self):
        if self.state_id:
            self.country_id = self.state_id.country_id.id

    @api.onchange('address_id')
    def _onchange_address(self):
        self.work_phone = self.address_id.phone
        self.mobile_phone = self.address_id.mobile

    @api.onchange('company_id')
    def _onchange_company(self):
        address = self.company_id.partner_id.address_get(['default'])
        self.address_id = address['default'] if address else False

    @api.onchange('user_id')
    def _onchange_user(self):
        self.work_email = self.user_id.email
        self.name = self.user_id.name
        self.image = self.user_id.image

    @api.multi
    def write(self, vals):
        if 'name' in vals:
           vals['name_related'] = vals['name']
        return super(SHealthPhysician, self).write(vals)

    #BÁO CÁO CÁN BỘ CÔNG CHỨC VIÊN CHỨC
    def report_physician(self):
        physician_attachment = self.env['ir.attachment'].browse(self.env.ref('shealth_all_in_one.bao_cao_can_bo').id)
        decode = base64.b64decode(physician_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active

        #get all Khoa phòng
        department = self.env['sh.medical.health.center.ward'].search([])
        data = []
        data_line_tong_CBCNV = []
        data_line_tong_CBCNV_nu = []
        tong_CBCNV = 0
        tong_CBCNV_nu = 0

        TONG_GIAO_SU = 0
        TONG_TIEN_SI = 0
        TONG_THAC_SI = 0
        TONG_CKII = 0
        TONG_CKI = 0
        TONG_BAC_SI = 0
        TONG_DH_DUOCSI = 0
        TONG_DH_YTA = 0
        TONG_TH_YSI = 0
        TONG_TH_KTV = 0
        TONG_TH_DUOCSI = 0
        TONG_TH_YTA = 0
        TONG_TH_NHS = 0
        TONG_SH_YTA = 0
        TONG_SH_NHS = 0
        TONG_SH_DUOCTA = 0
        TONG_HD_TRONG = 0
        TONG_HD_NGOAI = 0

        TONG_GIAO_SU_NU = 0
        TONG_TIEN_SI_NU = 0
        TONG_THAC_SI_NU = 0
        TONG_CKII_NU = 0
        TONG_CKI_NU = 0
        TONG_BAC_SI_NU = 0
        TONG_DH_DUOCSI_NU = 0
        TONG_DH_YTA_NU = 0
        TONG_TH_YSI_NU = 0
        TONG_TH_KTV_NU = 0
        TONG_TH_DUOCSI_NU = 0
        TONG_TH_YTA_NU = 0
        TONG_TH_NHS_NU = 0
        TONG_SH_YTA_NU = 0
        TONG_SH_NHS_NU = 0
        TONG_SH_DUOCTA_NU = 0
        TONG_HD_TRONG_NU = 0
        TONG_HD_NGOAI_NU = 0

        for dep in department:
            TONG_SO = self.env['sh.medical.physician'].search_count([('department','in',dep.id)])
            GT_NU = self.env['sh.medical.physician'].search_count([('department','in',dep.id),('gender','=','female')])
            HD_TRONG = self.env['sh.medical.physician'].search_count([('department','in',dep.id),('contract_type','=','fulltime')])
            HD_NGOAI = self.env['sh.medical.physician'].search_count([('department','in',dep.id),('contract_type','=','parttime')])

            id_giao_su = self.env['sh.medical.degrees'].search([('keyword','=','GIAO_SU')],limit=1)
            id_tien_si = self.env['sh.medical.degrees'].search([('keyword','=','TIEN_SI')],limit=1)
            id_thac_si = self.env['sh.medical.degrees'].search([('keyword','=','THAC_SI')],limit=1)
            id_ckii = self.env['sh.medical.degrees'].search([('keyword','=','CKII')],limit=1)
            id_cki = self.env['sh.medical.degrees'].search([('keyword','=','CKI')],limit=1)
            id_bac_si = self.env['sh.medical.degrees'].search([('keyword','=','BAC_SI')],limit=1)
            id_dh_duocsi = self.env['sh.medical.degrees'].search([('keyword','=','DH_DUOCSI')],limit=1)
            id_dh_yta = self.env['sh.medical.degrees'].search([('keyword','=','DH_YTA')],limit=1)
            id_th_ysi = self.env['sh.medical.degrees'].search([('keyword','=','TH_YSI')],limit=1)
            id_th_ktv = self.env['sh.medical.degrees'].search([('keyword','=','TH_KTV')],limit=1)
            id_th_duocsi = self.env['sh.medical.degrees'].search([('keyword','=','TH_DUOCSI')],limit=1)
            id_th_yta = self.env['sh.medical.degrees'].search([('keyword','=','TH_YTA')],limit=1)
            id_th_nhs = self.env['sh.medical.degrees'].search([('keyword','=','TH_NHS')],limit=1)
            id_sh_yta = self.env['sh.medical.degrees'].search([('keyword','=','SH_YTA')],limit=1)
            id_sh_nhs = self.env['sh.medical.degrees'].search([('keyword','=','SH_NHS')],limit=1)
            id_sh_duocta = self.env['sh.medical.degrees'].search([('keyword','=','SH_DUOCTA')],limit=1)

            # CBCNV Nu
            GIAO_SU_NU = self.env['sh.medical.physician'].search_count([('department','in',dep.id),('degree_id', 'in', id_giao_su.id),('gender','=','female')])
            TIEN_SI_NU = self.env['sh.medical.physician'].search_count([('department','in',dep.id),('degree_id', 'in', id_tien_si.id),('gender','=','female')])
            THAC_SI_NU = self.env['sh.medical.physician'].search_count([('department','in',dep.id),('degree_id', 'in', id_thac_si.id),('gender','=','female')])
            CKII_NU = self.env['sh.medical.physician'].search_count([('department','in',dep.id),('degree_id','in',id_ckii.id),('gender','=','female')])
            CKI_NU = self.env['sh.medical.physician'].search_count([('department','in',dep.id),('degree_id','in',id_cki.id),('gender','=','female')])
            BAC_SI_NU = self.env['sh.medical.physician'].search_count([('department','in',dep.id),('degree_id','in',id_bac_si.id),('gender','=','female')])
            DH_DUOCSI_NU = self.env['sh.medical.physician'].search_count([('department','in',dep.id),('degree_id','in',id_dh_duocsi.id),('gender','=','female')])
            DH_YTA_NU = self.env['sh.medical.physician'].search_count([('department','in',dep.id),('degree_id','in',id_dh_yta.id),('gender','=','female')])
            TH_YSI_NU = self.env['sh.medical.physician'].search_count([('department','in',dep.id),('degree_id','in',id_th_ysi.id),('gender','=','female')])
            TH_KTV_NU = self.env['sh.medical.physician'].search_count([('department','in',dep.id),('degree_id','in',id_th_ktv.id),('gender','=','female')])
            TH_DUOCSI_NU = self.env['sh.medical.physician'].search_count([('department','in',dep.id),('degree_id','in',id_th_duocsi.id),('gender','=','female')])
            TH_YTA_NU = self.env['sh.medical.physician'].search_count([('department','in',dep.id),('degree_id','in',id_th_yta.id),('gender','=','female')])
            TH_NHS_NU = self.env['sh.medical.physician'].search_count([('department','in',dep.id),('degree_id','in',id_th_nhs.id),('gender','=','female')])
            SH_YTA_NU = self.env['sh.medical.physician'].search_count([('department','in',dep.id),('degree_id','in',id_sh_yta.id),('gender','=','female')])
            SH_NHS_NU = self.env['sh.medical.physician'].search_count([('department','in',dep.id),('degree_id','in',id_sh_nhs.id),('gender','=','female')])
            SH_DUOCTA_NU = self.env['sh.medical.physician'].search_count([('department','in',dep.id),('degree_id','in',id_sh_duocta.id),('gender','=','female')])
            HD_TRONG_NU = self.env['sh.medical.physician'].search_count([('department','in',dep.id),('contract_type','=','fulltime'),('gender','=','female')])
            HD_NGOAI_NU = self.env['sh.medical.physician'].search_count([('department','in',dep.id),('contract_type','=','parttime'),('gender','=','female')])

            GIAO_SU = self.env['sh.medical.physician'].search_count([('department','in',dep.id),('degree_id','in',id_giao_su.id)])
            TIEN_SI = self.env['sh.medical.physician'].search_count([('department','in',dep.id),('degree_id','in',id_tien_si.id)])
            THAC_SI = self.env['sh.medical.physician'].search_count([('department','in',dep.id),('degree_id','in',id_thac_si.id)])
            CKII = self.env['sh.medical.physician'].search_count([('department','in',dep.id),('degree_id','in',id_ckii.id)])
            CKI = self.env['sh.medical.physician'].search_count([('department','in',dep.id),('degree_id','in',id_cki.id)])
            BAC_SI = self.env['sh.medical.physician'].search_count([('department','in',dep.id),('degree_id','in',id_bac_si.id)])
            DH_DUOCSI = self.env['sh.medical.physician'].search_count([('department','in',dep.id),('degree_id','in',id_dh_duocsi.id)])
            DH_YTA = self.env['sh.medical.physician'].search_count([('department','in',dep.id),('degree_id','in',id_dh_yta.id)])
            TH_YSI = self.env['sh.medical.physician'].search_count([('department','in',dep.id),('degree_id','in',id_th_ysi.id)])
            TH_KTV = self.env['sh.medical.physician'].search_count([('department','in',dep.id),('degree_id','in',id_th_ktv.id)])
            TH_DUOCSI = self.env['sh.medical.physician'].search_count([('department','in',dep.id),('degree_id','in',id_th_duocsi.id)])
            TH_YTA = self.env['sh.medical.physician'].search_count([('department','in',dep.id),('degree_id','in',id_th_yta.id)])
            TH_NHS = self.env['sh.medical.physician'].search_count([('department','in',dep.id),('degree_id','in',id_th_nhs.id)])
            SH_YTA = self.env['sh.medical.physician'].search_count([('department','in',dep.id),('degree_id','in',id_sh_yta.id)])
            SH_NHS = self.env['sh.medical.physician'].search_count([('department','in',dep.id),('degree_id','in',id_sh_nhs.id)])
            SH_DUOCTA = self.env['sh.medical.physician'].search_count([('department','in',dep.id),('degree_id','in',id_sh_duocta.id)])

            tong_CBCNV += TONG_SO
            tong_CBCNV_nu += GT_NU

            TONG_GIAO_SU += GIAO_SU
            TONG_TIEN_SI += TIEN_SI
            TONG_THAC_SI +=THAC_SI
            TONG_CKII +=CKII
            TONG_CKI += CKI
            TONG_BAC_SI += BAC_SI
            TONG_DH_DUOCSI += DH_DUOCSI
            TONG_DH_YTA += DH_YTA
            TONG_TH_YSI += TH_YSI
            TONG_TH_KTV += TH_KTV
            TONG_TH_DUOCSI += TH_DUOCSI
            TONG_TH_YTA += TH_YTA
            TONG_TH_NHS += TH_NHS
            TONG_SH_YTA += SH_YTA
            TONG_SH_NHS += SH_NHS
            TONG_SH_DUOCTA += SH_DUOCTA
            TONG_HD_TRONG += HD_TRONG
            TONG_HD_NGOAI += HD_NGOAI

            TONG_GIAO_SU_NU += GIAO_SU_NU
            TONG_TIEN_SI_NU += TIEN_SI_NU
            TONG_THAC_SI_NU += THAC_SI_NU
            TONG_CKII_NU += CKII_NU
            TONG_CKI_NU += CKI_NU
            TONG_BAC_SI_NU += BAC_SI_NU
            TONG_DH_DUOCSI_NU += DH_DUOCSI_NU
            TONG_DH_YTA_NU += DH_YTA_NU
            TONG_TH_YSI_NU += TH_YSI_NU
            TONG_TH_KTV_NU += TH_KTV_NU
            TONG_TH_DUOCSI_NU += TH_DUOCSI_NU
            TONG_TH_YTA_NU += TH_YTA_NU
            TONG_TH_NHS_NU += TH_NHS_NU
            TONG_SH_YTA_NU += SH_YTA_NU
            TONG_SH_NHS_NU += SH_NHS_NU
            TONG_SH_DUOCTA_NU += SH_DUOCTA_NU
            TONG_HD_TRONG_NU += HD_TRONG_NU
            TONG_HD_NGOAI_NU += HD_NGOAI_NU

            data.append([('KHOA_PHONG', dep.name),('TONG_SO', TONG_SO),('GT_NU',GT_NU),
                 ('GIAO_SU', GIAO_SU),('TIEN_SI',TIEN_SI),('TIEN_SI_DUOC',0),('THAC_SI',THAC_SI),('THACSI_DUOC',0),('THACSI_YTCC',0),('CKII',CKII),('CKII_DUOC',0),('CKI',CKI),('CKI_DUOC',0),
                 ('BAC_SI', BAC_SI),('DH_DUOCSI',DH_DUOCSI),('YTCC',0),('KTVY',0),('DH_KHAC',0),('DH_YTA',DH_YTA),
                         ('CD_KTVY',0),('CD_NUHS',0),('CD_YTDD',0),('YTCC',0),
                 ('TH_YSI', TH_YSI),('TH_KTV',TH_KTV),('TH_DUOCSI',TH_DUOCSI),('TH_YTA',TH_YTA),('TH_NHS',TH_NHS),('TH_KHAC',0),
                 ('SH_YTA', SH_YTA),('SH_NHS',SH_NHS),('SH_DUOCTA',SH_DUOCTA),('SH_KTVY',0),
                         ('LUONG', 0),('CAN_BO_KHAC',0),
                 ('HD_TRONG',HD_TRONG),('HD_NGOAI',HD_NGOAI)])
            data_line_tong_CBCNV.append([('TONG_CBCNV', tong_CBCNV),('TRONG_DO_NU',0),
                                         ('TONG_GIAO_SU', TONG_GIAO_SU), ('TONG_TIEN_SI', TONG_TIEN_SI),
                                         ('TONG_TIEN_SI_DUOC', 0), ('TONG_THAC_SI', TONG_THAC_SI),('TONG_THAC_SI_DUOC',0),
                                         ('TONG_THACSI_YTCC', 0), ('TONG_CKII', TONG_CKII), ('TONG_CKII_DUOC', 0),
                                         ('TONG_CKI', TONG_CKI),
                                         ('TONG_CKI_DUOC', 0),
                                         ('TONG_BAC_SI', TONG_BAC_SI), ('TONG_DH_DUOCSI', TONG_DH_DUOCSI),
                                         ('TONG_YTCC', 0), ('TONG_KTVY', 0), ('TONG_DH_KHAC', 0),
                                         ('TONG_DH_YTA', TONG_DH_YTA),
                                         ('TONG_CD_KTVY', 0), ('TONG_CD_NUHS', 0), ('TONG_CD_YTDD', 0),
                                         ('TONG_YTCC', 0),
                                         ('TONG_TH_YSI', TONG_TH_YSI), ('TONG_TH_KTV', TONG_TH_KTV),
                                         ('TONG_TH_DUOCSI', TONG_TH_DUOCSI), ('TONG_TH_YTA', TONG_TH_YTA),
                                         ('TONG_TH_NHS', TONG_TH_NHS), ('TONG_TH_KHAC', 0),
                                         ('TONG_SH_YTA', TONG_SH_YTA), ('TONG_SH_NHS', TONG_SH_NHS),
                                         ('TONG_SH_DUOCTA', TONG_SH_DUOCTA), ('TONG_SH_KTVY', 0),
                                         ('TONG_LUONG',0), ('TONG_CAN_BO_KHAC', 0),
                                         ('HD_TRONG', TONG_HD_TRONG), ('HD_NGOAI', TONG_HD_NGOAI)])
            data_line_tong_CBCNV_nu.append([('TONG_CBCNV_NU', tong_CBCNV_nu), ('TRONG_DO_NU', 0),
                                         ('TONG_GIAO_SU_NU', TONG_GIAO_SU_NU), ('TONG_TIEN_SI_NU', TONG_TIEN_SI_NU),
                                         ('TONG_TIEN_SI_DUOC_NU', 0), ('TONG_THAC_SI_NU', TONG_THAC_SI_NU),
                                         ('TONG_THAC_SI_DUOC_NU', 0),
                                         ('TONG_THACSI_YTCC_NU', 0), ('TONG_CKII_NU', TONG_CKII_NU), ('TONG_CKII_DUOC_NU', 0),
                                         ('TONG_CKI_NU', TONG_CKI_NU),
                                         ('TONG_CKI_DUOC_NU', 0),
                                         ('TONG_BAC_SI_NU', TONG_BAC_SI_NU), ('TONG_DH_DUOCSI_NU', TONG_DH_DUOCSI_NU),
                                         ('TONG_YTCC_NU', 0), ('TONG_KTVY_NU', 0), ('TONG_DH_KHAC_NU', 0),
                                         ('TONG_DH_YTA_NU', TONG_DH_YTA_NU),
                                         ('TONG_CD_KTVY_NU', 0), ('TONG_CD_NUHS_NU', 0), ('TONG_CD_YTDD_NU', 0),
                                         ('TONG_YTCC_NU', 0),
                                         ('TONG_TH_YSI_NU', TONG_TH_YSI_NU), ('TONG_TH_KTV_NU', TONG_TH_KTV_NU),
                                         ('TONG_TH_DUOCSI_NU', TONG_TH_DUOCSI_NU), ('TONG_TH_YTA_NU', TONG_TH_YTA_NU),
                                         ('TONG_TH_NHS_NU', TONG_TH_NHS_NU), ('TONG_TH_KHAC_NU', 0),
                                         ('TONG_SH_YTA_NU', TONG_SH_YTA_NU), ('TONG_SH_NHS_NU', TONG_SH_NHS_NU),
                                         ('TONG_SH_DUOCTA_NU', TONG_SH_DUOCTA_NU), ('TONG_SH_KTVY_NU', 0),
                                         ('TONG_LUONG_NU',0), ('TONG_CAN_BO_KHAC_NU', 0),
                                         ('HD_TRONG_NU', TONG_HD_TRONG_NU), ('HD_NGOAI_NU', TONG_HD_NGOAI_NU)])

        # ws['b2'].value = self.env.user.company_id.name
        ws['a2'].value = datetime.datetime.today().strftime('Ngày %d Tháng %m Năm %Y')
        row = 11
        line_font = Font(name='Times New Roman', size=12)
        for line_tong in data_line_tong_CBCNV:
            print(line_tong)
            ws.cell(row=8, column=3).border, ws.cell(row=8, column=3).value, ws.cell(row=8,column=3).alignment, ws.cell(row=8, column=3).font = all_border_thin, 8, Alignment(horizontal='center'), line_font
            for n, value in enumerate(line_tong,3):
                cell = ws.cell(row=8,column=n)
                if n == 3:
                    cell.border, cell.alignment, cell.font = all_border_thin, Alignment(horizontal='center'), line_font
                else:
                    cell.number_format = '0;-0;"-";@'
                    cell.border, cell.alignment, cell.font = all_border_thin, Alignment(horizontal='center'), Font(name='Times New Roman', size=12, bold=True)
                cell.value = value[1]


        for line_tong_nu in data_line_tong_CBCNV_nu:
            print(line_tong_nu)
            ws.cell(row=9, column=3).border, ws.cell(row=9, column=3).value, ws.cell(row=9,column=3).alignment, ws.cell(row=9, column=3).font = all_border_thin, 9, Alignment(horizontal='center'), line_font
            for n, value in enumerate(line_tong_nu,3):
                cell = ws.cell(row=9, column=n)
                if n == 3:
                    cell.border, cell.alignment, cell.font = all_border_thin, Alignment(horizontal='center'), line_font
                else:
                    cell.number_format = '0;-0;"-";@'
                    cell.border, cell.alignment, cell.font = all_border_thin, Alignment(horizontal='center'), Font(name='Times New Roman', size=12, bold=True)
                cell.value = value[1]

        for line in data:
            ws.cell(row=row, column=1).border, ws.cell(row=row, column=1).value, ws.cell(row=row, column=1).alignment, ws.cell(row=row, column=1).font = all_border_thin, row-8, Alignment(horizontal='center'),line_font
            for n, value in enumerate(line,2):
                cell = ws.cell(row=row, column=n)
                if n == 2:
                    cell.border, cell.alignment, cell.font = all_border_thin, Alignment(horizontal='left'), line_font
                else:
                    cell.number_format = '0;-0;"-";@'
                    cell.border, cell.alignment, cell.font = all_border_thin, Alignment(horizontal='center'), Font(name='Times New Roman', size=12, bold=True)

                cell = ws.cell(row=row, column=n)
                cell.value = value[1]
            row += 1

        #footer sign
        ws.cell(row=row+1, column=5).value,ws.cell(row=row+1, column=5).font = 'NGƯỜI LẬP BIỂU',Font(name='Times New Roman', size=10, bold=True)
        ws.cell(row=row+1, column=12).value,ws.cell(row=row+1, column=12).font = 'TRƯỞNG PHÒNG KHTH',Font(name='Times New Roman', size=10, bold=True)
        ws.cell(row=row+1, column=21).value,ws.cell(row=row+1, column=21).font = datetime.datetime.today().strftime('Ngày %d tháng %m năm %Y'), Font(name='Times New Roman', size=10, italic=True)

        ws.cell(row=row+2, column=5).value,ws.cell(row=row+2, column=5).font = "(Chức danh, ký tên)", Font(name='Times New Roman', size=10, italic=True)
        ws.cell(row=row+2, column=12).value,ws.cell(row=row+2, column=12).font = "(Chức danh, ký tên)", Font(name='Times New Roman', size=10, italic=True)
        ws.cell(row=row+2, column=21).value,ws.cell(row=row+2, column=21).font = "GIÁM ĐỐC", Font(name='Times New Roman', size=10, bold=True)
        ws.cell(row=row+3, column=21).value, ws.cell(row=row+3, column=21).font = "(Ký tên đóng dấu)", Font(name='Times New Roman', size=10, italic=True)

        ws.print_area = 'A1:AA%s'% str(row+8)

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({'name': 'MS report',
                                                              'datas_fname': 'BC_tinh_hinh_nhan_vien.xlsx',
                                                              'datas': report,
                                                              'res_model': 'temp.creation',
                                                              'public': True})

        print(attachment.id)

        return {'name': (_('Báo cáo tình hình cán bộ, công chức')),
                'type': 'ir.actions.act_window',
                'res_model': 'temp.wizard',
                'view_mode': 'form',
                'view_type': 'form',
                'target': 'inline',
                'view_id': self.env.ref('ms_templates.report_wizard').id,
                'context': {'attachment_id': attachment.id}}

        # return {'name': (_('Báo cáo tình hình cán bộ, công chức')),
        #         'type': 'ir.actions.act_window',
        #         'res_model': 'temp.wizard',
        #         'view_mode': 'form',
        #         'view_type': 'form',
        #         'target': 'inline',
        #         'view_id': self.env.ref('ms_templates.report_wizard').id,
        #         'context': {'default_template_id': template.id, 'external_keys': external_keys,
        #                     'active_model': 'sh.medical.physician'}}
#
class ResUsers(models.Model):
    _inherit = "res.users"

    physician_ids = fields.One2many('sh.medical.physician','sh_user_id', string='Physicians', auto_join=True)

class SHealthCentersWards(models.Model):
    _inherit = "sh.medical.health.center.ward"

    physician = fields.Many2many('sh.medical.physician', 'sh_medical_physician_department_rel',
                                  'dep_id', 'physician_id', string='Physicians', help="Physicians in Department")

class SHealthPhysicianLine(models.Model):

    # Array containing different days name
    PHY_DAY = [
        ('Monday', 'Monday'),
        ('Tuesday', 'Tuesday'),
        ('Wednesday', 'Wednesday'),
        ('Thursday', 'Thursday'),
        ('Friday', 'Friday'),
        ('Saturday', 'Saturday'),
        ('Sunday', 'Sunday'),
    ]

    _name = "sh.medical.physician.line"
    _description = "Information about doctor availability"

    name = fields.Selection(PHY_DAY, string='Available Day(s)', required=True)
    start_time = fields.Float(string='Start Time (24h format)')
    end_time = fields.Float(string='End Time (24h format)')
    physician_id = fields.Many2one('sh.medical.physician', string='Physician', index=True, ondelete='cascade')

# Appointment Management

class SHealthAppointment(models.Model):
    _name = 'sh.medical.appointment'
    _description = 'Appointment'
    _inherit = ['mail.thread']

    URGENCY_LEVEL = [
                ('Normal', 'Normal'),
                ('Urgent', 'Urgent'),
                ('Medical Emergency', 'Medical Emergency'),
            ]

    PATIENT_STATUS = [
                ('Ambulatory', 'Ambulatory'),
                ('Outpatient', 'Outpatient'),
                ('Inpatient', 'Inpatient'),
            ]

    APPOINTMENT_STATUS = [
            ('Scheduled', 'Scheduled'),
            ('Completed', 'Completed'),
            ('Invoiced', 'Invoiced'),
        ]

    # Automatically detect logged in physician
    @api.multi
    def _get_physician(self):
        """Return default physician value"""
        therapist_obj = self.env['sh.medical.physician']
        domain = [('sh_user_id', '=', self.env.uid)]
        user_ids = therapist_obj.search(domain, limit=1)
        if user_ids:
            return user_ids.id or False
        else:
            return False

    # Calculating Appointment End date
    @api.multi
    def _get_appointment_end(self):
        for apm in self:
            end_date = False
            duration = 1
            if apm.duration:
                duration = apm.duration
            if apm.appointment_date:
                end_date = datetime.datetime.strptime(apm.appointment_date.strftime("%Y-%m-%d %H:%M:%S"), "%Y-%m-%d %H:%M:%S") + timedelta(hours=duration)
            apm.appointment_end = end_date
        return True

    name = fields.Char(string='Appointment #', size=64, default=lambda *a: '/')
    patient = fields.Many2one('sh.medical.patient', string='Patient', help="Patient Name", required=True, readonly=True, states={'Scheduled': [('readonly', False)]})
    doctor = fields.Many2one('sh.medical.physician', string='Physician', help="Current primary care / family doctor", domain=[('is_pharmacist','=',False)], required=True, readonly=True,states={'Scheduled': [('readonly', False)]}, default=_get_physician)
    appointment_date = fields.Datetime(string='Appointment Date', required=True, readonly=True,states={'Scheduled': [('readonly', False)]}, default=datetime.datetime.now())
    appointment_end = fields.Datetime(compute=_get_appointment_end, string='Appointment End Date', readonly=True, states={'Scheduled': [('readonly', False)]})
    duration = fields.Integer(string='Duration (Hours)', readonly=True, states={'Scheduled': [('readonly', False)]}, default=lambda *a: 1)
    institution = fields.Many2one('sh.medical.health.center', string='Health Center', help="Medical Center", readonly=True, states={'Scheduled': [('readonly', False)]})
    urgency_level = fields.Selection(URGENCY_LEVEL, string='Urgency Level', readonly=True, states={'Scheduled': [('readonly', False)]}, default=lambda *a: 'Normal')
    comments = fields.Text(string='Comments', readonly=True, states={'Scheduled': [('readonly', False)]})
    patient_status = fields.Selection(PATIENT_STATUS, string='Patient Status', readonly=True, states={'Scheduled': [('readonly', False)]}, default=lambda *a: 'Inpatient')
    state = fields.Selection(APPOINTMENT_STATUS, string='State', readonly=True, default=lambda *a: 'Scheduled')

    _order = "appointment_date desc"

    @api.model
    def create(self, vals):
        if vals.get('doctor') and vals.get('appointment_date'):
            self.check_physician_availability(vals.get('doctor'),vals.get('appointment_date'))

        sequence = self.env['ir.sequence'].next_by_code('sh.medical.appointment')
        vals['name'] = sequence
        health_appointment = super(SHealthAppointment, self).create(vals)
        return health_appointment

    @api.multi
    def check_physician_availability(self, doctor, appointment_date):
        available = False
        DATETIME_FORMAT = "%Y-%m-%d %H:%M:%S"
        patient_line_obj = self.env['sh.medical.physician.line']
        need_to_check_availability = False

        query_doctor_availability = _("select appointment_type from sh_medical_physician where id=%s") % (doctor)
        self.env.cr.execute(query_doctor_availability)
        val = self.env.cr.fetchone()
        if val and val[0]:
            if val[0] == "On Weekly Schedule":
                need_to_check_availability = True

        # check if doctor is working on selected day of the week
        if need_to_check_availability:
            selected_day = datetime.datetime.strptime(appointment_date, DATETIME_FORMAT).strftime('%A')

            if selected_day:
                avail_days = patient_line_obj.search([('name', '=', str(selected_day)), ('physician_id', '=', doctor)],
                                                     limit=1)

                if not avail_days:
                    raise UserError(_('Physician is not available on selected day!'))
                else:
                    # get selected day's start and end time

                    phy_start_time = self.get_time_string(avail_days.start_time).split(':')
                    phy_end_time = self.get_time_string(avail_days.end_time).split(':')

                    user_pool = self.env['res.users']
                    user = user_pool.browse(self.env.uid)
                    tz = pytz.timezone(user.partner_id.tz) or pytz.utc

                    # get localized dates
                    appointment_date = pytz.utc.localize(
                        datetime.datetime.strptime(appointment_date, DATETIME_FORMAT)).astimezone(tz)

                    t1 = datetime.time(int(phy_start_time[0]), int(phy_start_time[1]), 0)
                    t3 = datetime.time(int(phy_end_time[0]), int(phy_end_time[1]), 0)

                    # get appointment hour and minute
                    t2 = datetime.time(appointment_date.hour, appointment_date.minute, 0)

                    if not (t2 > t1 and t2 < t3):
                        raise UserError(_('Physician is not available on selected time!'))
                    else:
                        available = True
        return available


    def get_time_string(self,duration):
        result =''
        currentHours = int(duration // 1)
        currentMinutes =int(round(duration % 1 * 60))
        if(currentHours <= 9):
            currentHours = "0" + str(currentHours)
        if(currentMinutes <= 9):
            currentMinutes = "0" + str(currentMinutes)
        result = str(currentHours)+":"+str(currentMinutes)
        return result

    @api.multi
    def _default_account(self):
        journal = self.env['account.journal'].search([('type', '=', 'sale')], limit=1)
        return journal.default_credit_account_id.id

    def action_appointment_invoice_create(self):
        invoice_obj = self.env["account.invoice"]
        invoice_line_obj = self.env["account.invoice.line"]
        inv_ids = []

        for acc in self:
            # Create Invoice
            if acc.patient:
                curr_invoice = {
                    'partner_id': acc.patient.partner_id.id,
                    'account_id': acc.patient.partner_id.property_account_receivable_id.id,
                    'patient': acc.patient.id,
                    'state': 'draft',
                    'type':'out_invoice',
                    'date_invoice': acc.appointment_date,
                    'origin': "Appointment # : " + acc.name,
                    'sequence_number_next_prefix': False
                }

                inv_ids = invoice_obj.create(curr_invoice)
                inv_id = inv_ids.id

                if inv_ids:
                    prd_account_id = self._default_account()
                    # Create Invoice line
                    curr_invoice_line = {
                        'name':"Consultancy invoice for " + acc.name,
                        'price_unit': acc.doctor.consultancy_price,
                        'quantity': 1,
                        'account_id': prd_account_id,
                        'invoice_id': inv_id,
                    }

                    inv_line_ids = invoice_line_obj.create(curr_invoice_line)

                self.write({'state': 'Invoiced'})
        return {
                'domain': "[('id','=', " + str(inv_id) + ")]",
                'name': _('Appointment Invoice'),
                'view_type': 'form',
                'view_mode': 'tree,form',
                'res_model': 'account.invoice',
                'type': 'ir.actions.act_window'
        }

    @api.multi
    def set_to_completed(self):
        return self.write({'state': 'Completed'})
    
    @api.multi
    def unlink(self):
        for appointment in self.filtered(lambda appointment: appointment.state not in ['Draft']):
            raise UserError(_('You can not delete an appointment which is not in "Draft" state !!'))
        return super(SHealthAppointment, self).unlink()

class SHealthReExamService(models.Model):
    _name = 'sh.medical.prescription.service.reexam'
    _description = "Lịch tái khám theo đơn thuốc"

    NOTE = [
        ('Check', 'Khám'),
        ('ReCheck', 'Khám định kỳ')
    ]

    name = fields.Char("Tên", required=True)
    prescription_id = fields.Many2one('sh.medical.prescription', 'Service', ondelete='cascade')
    service_date = fields.Datetime('Ngày làm dịch vụ', related='prescription_id.date')
    date_recheck = fields.Date('Ngày khám', compute="_compute_date", readonly=0)
    after_service_date = fields.Integer('Sau ngày làm dịch vụ (ngày)', default=1, required=True)
    type = fields.Selection(NOTE, 'Loại', required=True)
    for_service = fields.Text('Cho dịch vụ')

    @api.depends('after_service_date')
    def _compute_date(self):
        for record in self:
            if record.after_service_date and record.service_date:
                record.date_recheck = (datetime.datetime.strptime(record.service_date.strftime("%Y-%m-%d %H:%M:%S"), "%Y-%m-%d %H:%M:%S") + timedelta(days=record.after_service_date)).strftime("%Y-%m-%d")
                # record.date_recheck = fields.Datetime.context_timestamp(datetime.datetime.now()+ relativedelta(days=l.after_service_date)).strftime('%d/%m/%Y')

# Prescription Management
class SHealthPrescriptions(models.Model):
    _name = 'sh.medical.prescription'
    _description = 'Prescriptions'

    _order = "name"

    STATES = [
        ('Draft', 'Draft'),
        # ('Invoiced', 'Invoiced'),
        # ('Sent to Pharmacy', 'Sent to Pharmacy'),
        ('Đã xuất thuốc', 'Đã xuất thuốc'),
    ]

    # Automatically detect logged in physician
    @api.multi
    def _get_physician(self):
        """Return default physician value"""
        therapist_obj = self.env['sh.medical.physician']
        domain = [('sh_user_id', '=', self.env.uid)]
        user_ids = therapist_obj.search(domain, limit=1)
        if user_ids:
            return user_ids.id or False
        else:
            return False

    name = fields.Char(string='Prescription #', size=64, readonly=True, required=True, default=lambda *a: '/')
    patient = fields.Many2one('sh.medical.patient', string='Patient', help="Patient Name", required=True, readonly=True, states={'Draft': [('readonly', False)]})
    doctor = fields.Many2one('sh.medical.physician', string='Physician', domain=[('is_pharmacist','=',False)], help="Current primary care / family doctor", required=True, readonly=True, states={'Draft': [('readonly', False)]}, default=_get_physician)
    pharmacy = fields.Many2one('sh.medical.health.center.pharmacy', 'Pharmacy', readonly=True, states={'Draft': [('readonly', False)]})
    location_id = fields.Many2one('stock.location', 'Medicine stock', readonly=True, states={'Draft': [('readonly', False)]}, default=lambda self: self.env.ref('shealth_all_in_one.sh_location_medicine_prescription_knhn', raise_if_not_found=False))
    date = fields.Datetime(string='Ngày làm dịch vụ', readonly=True, states={'Draft': [('readonly', False)]}, default=datetime.datetime.now())
    date_out = fields.Datetime(string='Prescription Date out', readonly=True, states={'Draft': [('readonly', False)]}, default=datetime.datetime.now())
    info = fields.Text(string='Prescription Notes', readonly=True, states={'Draft': [('readonly', False)]})
    prescription_line = fields.One2many('sh.medical.prescription.line', 'prescription_id', string='Prescription Lines', readonly=True, states={'Draft': [('readonly', False)]})
    state = fields.Selection(STATES, 'State', readonly=True, default=lambda *a: 'Draft')

    services = fields.Many2many('sh.medical.health.center.service', 'sh_prescription_service_rel', 'prescription_id',
                               'service_id',track_visibility='onchange', string='Services', readonly=True, states={'Draft': [('readonly', False)]})

    days_reexam = fields.One2many('sh.medical.prescription.service.reexam', 'prescription_id', string='Lịch tái khám', default=False)

    @api.onchange('date', 'date_out')
    def _onchange_date(self):
        if self.date and self.date_out:
            if self.date > self.date_out:
                raise UserError(
                    _('Thông tin không hợp lệ! Ngày làm dịch vụ phải trước ngày xuất thuốc!'))

    @api.multi
    def write(self, vals):
        if vals.get('date_out') or vals.get('date'):
            for record in self:
                date = vals.get('date') or record.date
                date_out = vals.get('date_out') or record.date_out
                if date and date_out and (date > date_out):
                    raise UserError(
                        _(
                            'Thông tin không hợp lệ! Ngày làm dịch vụ phải trước ngày xuất thuốc!'))

        return super(SHealthPrescriptions, self).write(vals)

    @api.multi
    def view_detail_prescriptions(self):
        return {
            'name': _('Chi tiết Đơn thuốc'),  # label
            'type': 'ir.actions.act_window',
            'view_mode': 'form',
            'view_id': self.env.ref('shealth_all_in_one.sh_medical_prescription_view').id,
            'res_model': 'sh.medical.prescription',  # model want to display
            'target': 'new',  # if you want popup,
            'context': {'form_view_initial_mode': 'edit'},
            'res_id': self.id
        }

    @api.onchange('date', 'date_out')
    def _onchange_date(self):
        if not self.date_out:
            self.date_out = self.date

        if self.date and self.date_out:
            if self.date > self.date_out:
                raise UserError(
                    _('Thông tin không hợp lệ! Ngày kê đơn phải trước ngày xuất thuốc!'))

    @api.model
    def create(self, vals):
        if vals.get('prescription_line'):
            # Lấy ngày hiện tại
            current_date = datetime.datetime.now()
            # Lấy 2 số cuối của năm
            year_last_two_digits = str(current_date.year)[-2:] + str(current_date.month).zfill(2)
            sequence = self.env['ir.sequence'].next_by_code('sh.medical.prescription')
            vals['name'] ='00885' + year_last_two_digits + sequence + '-' + 'c'
            # vals['name'] = '92959MT' + sequence + 'HT-' + 'c'
            health_prescription = super(SHealthPrescriptions, self).create(vals)
            return health_prescription
        else:
            raise UserError(_('Bạn phải nhập ít nhất 1 thuốc cho đơn thuốc!'))

    @api.multi
    def write(self, vals):
        if vals.get('date_out') or vals.get('date'):
            for record in self:
                date = vals.get('date') or record.date
                date_out = vals.get('date_out') or record.date_out

                if isinstance(date, pycompat.string_types):
                    date = datetime.datetime.strptime(date, '%Y-%m-%d %H:%M:%S')
                if isinstance(date_out, pycompat.string_types):
                    date_out = datetime.datetime.strptime(date_out, '%Y-%m-%d %H:%M:%S')

                if date and date_out and (date > date_out):
                    raise UserError(
                        _(
                            'Thông tin không hợp lệ! Ngày kê đơn phải trước ngày xuất thuốc!'))

        return super(SHealthPrescriptions, self).write(vals)

    @api.onchange('services')
    def _onchange_services_prescription(self):
        self.prescription_line = False
        self.days_reexam = False

        if self.services:
            prescription_line = []
            seq = 0
            info_data = ''
            id_prescription_line = {}
            for service_done in self.services:
                if service_done.prescription_ids:
                    #check đã có thì cộng dồn
                    for prescription in service_done.prescription_ids:
                        med_dict_key = str(prescription.product_id.id)
                        med_init_qty_line = prescription.dose_unit._compute_quantity(prescription.qty, prescription.product_id.uom_id)
                        med_qty_line = prescription.dose_unit._compute_quantity(prescription.qty, prescription.product_id.uom_id)

                        if str(prescription.product_id.id) not in id_prescription_line:
                            seq += 1
                            id_prescription_line[med_dict_key] = seq
                            prescription_line.append((0, 0, {
                                'name': prescription.product_id.id,
                                'patient': self.patient,
                                'init_qty': med_init_qty_line,
                                'qty': med_qty_line,
                                'dose': prescription.dose,
                                'dose_unit_related': prescription.product_id.uom_id.id,
                                'common_dosage': prescription.common_dosage.id,
                                'duration': prescription.duration,
                                'duration_period': prescription.duration_period,
                                'is_buy_out': prescription.is_buy_out,
                                'info': prescription.note}))
                        else:
                            prescription_line[id_prescription_line[med_dict_key] - 1][2]['init_qty'] += med_init_qty_line
                            prescription_line[id_prescription_line[med_dict_key] - 1][2]['qty'] += med_qty_line

                #có hướng dẫn chăm sóc và lịch tái khám
                if service_done.info:
                    # info_data +='<p style="page-break-before:always;"> </p>' + service_done.info
                    info_data += service_done.info

                data_recheck = []
                if service_done.days_reexam:
                    for rc in service_done.days_reexam:
                        data_recheck.append((0, 0, {
                            'name': rc.name,
                            'after_service_date': rc.after_service_date,
                            'type': rc.type,
                            'service_date': self.date,
                            'date_recheck': (datetime.datetime.strptime(self.date.strftime("%Y-%m-%d %H:%M:%S"), "%Y-%m-%d %H:%M:%S") + timedelta(days=rc.after_service_date)).strftime("%Y-%m-%d") or fields.Date.today(),
                            'for_service': service_done.name
                        }))

            #đổ data đơn thuốc
            self.prescription_line = prescription_line
            #đổ data hướng dẫn theo dịch vụ
            self.info = info_data
            # đổ data lịch tái khám
            self.days_reexam = data_recheck

    @api.multi
    def _default_account(self):
        journal = self.env['account.journal'].search([('type', '=', 'sale')], limit=1)
        return journal.default_credit_account_id.id

    def action_prescription_invoice_create(self):
        invoice_obj = self.env["account.invoice"]
        invoice_line_obj = self.env["account.invoice.line"]
        inv_ids = []

        for pres in self:
            # Create Invoice
            if pres.patient:
                curr_invoice = {
                    'partner_id': pres.patient.partner_id.id,
                    'account_id': pres.patient.partner_id.property_account_receivable_id.id,
                    'patient': pres.patient.id,
                    'state': 'draft',
                    'type':'out_invoice',
                    'date_invoice': pres.date.strftime('%Y-%m-%d'),
                    'origin': "Prescription# : " + pres.name,
                    'sequence_number_next_prefix': False
                }

                inv_ids = invoice_obj.create(curr_invoice)
                inv_id = inv_ids.id

                if inv_ids:
                    prd_account_id = self._default_account()
                    if pres.prescription_line:
                        for ps in pres.prescription_line:

                            # Create Invoice line
                            curr_invoice_line = {
                                'name': ps.name.product_id.name,
                                'product_id': ps.name.product_id.id,
                                'price_unit': ps.name.product_id.list_price,
                                'quantity': ps.qty,
                                'account_id': prd_account_id,
                                'invoice_id': inv_id,
                            }

                            inv_line_ids = invoice_line_obj.create(curr_invoice_line)

                self.write({'state': 'Invoiced'})

        return {
                'domain': "[('id','=', " + str(inv_id) + ")]",
                'name': 'Prescription Invoice',
                'view_type': 'form',
                'view_mode': 'tree,form',
                'res_model': 'account.invoice',
                'type': 'ir.actions.act_window'
        }

    @api.multi
    def unlink(self):
        for priscription in self.filtered(lambda priscription: priscription.state not in ['Draft']):
            raise UserError(_('You can not delete a prescription which is not in "Draft" state !!'))
        return super(SHealthPrescriptions, self).unlink()

    def action_prescription_send_to_pharmacy(self):
        pharmacy_obj = self.env["sh.medical.health.center.pharmacy.line"]
        pharmacy_line_obj = self.env["sh.medical.health.center.pharmacy.prescription.line"]
        res = {}
        for pres in self:
            if not pres.pharmacy:
                raise UserError(_('No pharmacy selected !!'))
            else:
                curr_pres = {
                    'name': pres.id,
                    'patient': pres.patient.id,
                    'doctor': pres.doctor.id,
                    'pharmacy_id': pres.pharmacy.id,
                }
                phy_ids = pharmacy_obj.create(curr_pres)

                if phy_ids:
                    if pres.prescription_line:
                        for ps in pres.prescription_line:

                            # Create Prescription line
                            curr_pres_line = {
                                'name': ps.name.id,
                                'indication': ps.indication.id,
                                'price_unit': ps.name.product_id.list_price,
                                'qty': ps.qty,
                                'actual_qty': ps.qty,
                                'prescription_id': phy_ids.id,
                                'state': 'Draft',
                            }

                            phy_line_ids = pharmacy_line_obj.create(curr_pres_line)

                res = self.write({'state': 'Sent to Pharmacy'})

        return True

    def set_to_draft(self):
        #mở lại phiếu
        self.reverse_prescription()
        return self.write({'state': 'Draft'})

    def reverse_prescription(self):
        # num_of_location = len(self.prescription_line.mapped('location_id'))
        num_of_location = 1
        pick_need_reverses = self.env['stock.picking'].search([('origin', 'ilike', 'THBN - THUỐC KÊ ĐƠN - %s - %s' % (self.name,self.walkin.name))], order='create_date DESC', limit=num_of_location)
        if pick_need_reverses:
            for pick_need_reverse in pick_need_reverses:
                date_done = pick_need_reverse.date_done

                fail_pick_count = self.env['stock.picking'].search_count([('name', 'ilike', pick_need_reverse.name)])
                pick_need_reverse.name += '-FP%s' % fail_pick_count
                pick_need_reverse.move_ids_without_package.write(
                    {'reference': pick_need_reverse.name})  # sửa cả trường tham chiếu của move.line (Dịch chuyển kho)
                wizard = self.env['stock.return.picking'].with_context(active_id=pick_need_reverse.id, reopen_flag=True).create({'picking_id': pick_need_reverse.id})
                new_picking_id, pick_type_id = wizard._create_returns()
                new_picking = self.env['stock.picking'].browse(new_picking_id)
                new_picking.with_context(exact_location=True).action_assign()
                for move_line in new_picking.move_ids_without_package:
                    for move_live_detail in move_line.move_line_ids:
                        move_live_detail.qty_done = move_live_detail.product_uom_qty
                    # move_line.quantity_done = move_line.product_uom_qty
                new_picking.button_validate()

                # sua ngay hoan thanh
                for move_line in new_picking.move_ids_without_package:
                    move_line.move_line_ids.write(
                        {'date': date_done})  # sửa ngày hoàn thành ở stock move line
                new_picking.move_ids_without_package.write(
                    {'date': date_done})  # sửa ngày hoàn thành ở stock move

                new_picking.date_done = date_done
                new_picking.sci_date_done = date_done

    def check_have_record(self,record):
        if isinstance(record.id, models.NewId):
            res = super(SHealthPrescriptions, self).create({
                'patient': record.patient.id,
                'doctor': record.doctor.id,
                'walkin': record.walkin.id,
                'services': record.services.ids,
                'date': record.date,
                'date_out': record.date_out,
                'prescription_line': record.prescription_line.ids,
                'info': record.info,
            })
            return res
        else:
            return record

    def action_prescription_out_stock(self):
        #neu data đơn thuốc chưa được tạo thì gọi hàm tạo
        res = self.check_have_record(self)

        #trừ kho theo đơn thuốc xuất
        if res.date_out:
            date_out = res.date_out
        else:
            date_out = fields.Datetime.now()

        vals = {}
        validate_str = ''
        for medicine in res.prescription_line:
            #nếu thuốc này ko trừ trong kho
            if not medicine.is_buy_out:
                if medicine.qty > 0:  # CHECK SO LUONG SU DUNG > 0
                    quantity_on_hand = self.env['stock.quant']._get_available_quantity(medicine.name.product_id,
                                                                                       res.location_id)  # check quantity trong location
                    if medicine.dose_unit_related != medicine.name.product_id.uom_id:
                        medicine.write({'qty': medicine.dose_unit_related._compute_quantity(medicine.qty, medicine.name.product_id.uom_id),
                                   'dose_unit_related': medicine.name.product_id.uom_id})  # quy so suong su dung ve don vi chinh cua san pham

                    if quantity_on_hand < medicine.qty:
                        validate_str += "+ ""[%s]%s"": Còn %s %s tại ""%s"" \n" % (
                                medicine.name.product_id.default_code, medicine.name.product_id.name, str(quantity_on_hand), str(medicine.dose_unit_related.name), res.location_id.name)
                    else:  # truong one2many trong stock picking de tru cac product trong inventory
                        sub_vals = (0, 0, {
                            'name': 'THBN - THUỐC KÊ ĐƠN: ' + medicine.name.product_id.name,
                            'origin': res.name,
                            'date': date_out,
                            'company_id': self.env.user.company_id.id,
                            'date_expected': date_out,
                            'product_id': medicine.name.product_id.id,
                            'product_uom_qty': medicine.qty,
                            'product_uom': medicine.dose_unit_related.id,
                            'location_id': res.location_id.id,
                            'location_dest_id': res.patient.partner_id.property_stock_customer.id,
                            'partner_id': res.patient.partner_id.id,
                            # xuat cho khach hang/benh nhan nao
                        })
                        if not vals.get(str(res.location_id.id)):
                            vals[str(res.location_id.id)] = [sub_vals]
                        else:
                            vals[str(res.location_id.id)].append(sub_vals)

        # neu co data thuoc
        if vals and validate_str == '':
            # tao phieu xuat kho
            picking_type = self.env['stock.picking.type'].search([('code', '=', 'outgoing'),
                                                                  ('warehouse_id', '=',
                                                                   self.env.ref('stock.warehouse0').id)],
                                                                 limit=1).id
            for location_key in vals:
                pick_note = 'THBN - THUỐC KÊ ĐƠN - %s - %s - %s' % (
                res.name, res.walkin.name, location_key)
                pick_vals = {'note': pick_note,
                             'origin': pick_note,
                             'partner_id': res.patient.partner_id.id,
                             'patient_id': res.patient.id,
                             'picking_type_id': picking_type,
                             'location_id': int(location_key),
                             'location_dest_id': res.patient.partner_id.property_stock_customer.id,# xuat cho khach hang/benh nhan nao
                             'date_done': date_out,
                             'immediate_transfer': True,
                             'move_ids_without_package': vals[location_key]}
                fail_pick_name = self.env['stock.picking'].search(
                    [('origin', 'ilike', 'THBN - THUỐC KÊ ĐƠN - %s - %s - %s' % (res.name, res.walkin.name, location_key))],
                    limit=1).name
                if fail_pick_name:
                    pick_vals['name'] = fail_pick_name.split('-', 1)[0]
                stock_picking = self.env['stock.picking'].create(pick_vals)
                # TU DONG XUAT KHO
                stock_picking.with_context(exact_location=True).action_assign()  # ham check available trong inventory
                for move_line in stock_picking.move_ids_without_package: #set so luong done
                    for move_live_detail in move_line.move_line_ids:
                        move_live_detail.qty_done = move_live_detail.product_uom_qty
                    # move_line.quantity_done = move_line.product_uom_qty
                stock_picking.button_validate()  # ham tru product trong inventory

                # sua ngay hoan thanh
                for move_line in stock_picking.move_ids_without_package:
                    move_line.move_line_ids.write(
                        {'date': date_out})  # sửa ngày hoàn thành ở stock move line
                stock_picking.move_ids_without_package.write(
                    {'date': date_out})  # sửa ngày hoàn thành ở stock move

                stock_picking.date_done = date_out
                stock_picking.sci_date_done = date_out
                stock_picking.create_date = res.date

            #cập nhật vtth phiếu khám
            res.walkin.update_walkin_material()
        elif validate_str != '':
            raise ValidationError(_(
                "Các loại Thuốc sau đang không đủ số lượng tại tủ xuất:\n" + validate_str + "Hãy liên hệ với quản lý kho!"))

        res.write({'state': 'Đã xuất thuốc',
                   'date_out': date_out,
                   'isbuy': '1'
                   })

        return {'type': 'ir.actions.client',
            'tag': 'reload'}

    @api.multi
    def print_patient_prescription(self):
        return self.env.ref('shealth_all_in_one.action_sh_medical_report_patient_prescriptions').report_action(self)

    @api.multi
    def print_patient_huongdan(self):
        return self.env.ref('shealth_all_in_one.action_sh_medical_report_patient_prescriptions_huongdan').report_action(self)


class SHealthPrescriptionLines(models.Model):
    _name = 'sh.medical.prescription.line'
    _description = 'Prescription Lines'

    FREQUENCY_UNIT = [
        ('Seconds', 'Seconds'),
        ('Minutes', 'Minutes'),
        ('Hours', 'Hours'),
        ('Days', 'Days'),
        ('Weeks', 'Weeks'),
        ('When Required', 'When Required'),
    ]

    DURATION_UNIT = [
        ('Minutes', 'Minutes'),
        ('Hours', 'Hours'),
        ('Days', 'Days'),
        ('Months', 'Months'),
        ('Years', 'Years'),
        ('Indefinite', 'Indefinite'),
    ]

    prescription_id = fields.Many2one('sh.medical.prescription', string='Prescription Reference', required=True, ondelete='cascade', index=True)
    name = fields.Many2one('sh.medical.medicines', string='Medicines', help="Prescribed Medicines", domain=[('medicament_type','=','Medicine')], required=True)
    indication = fields.Many2one('sh.medical.pathology', string='Indication', help="Choose a disease for this medicament from the disease list. It can be an existing disease of the patient or a prophylactic.")
    dose = fields.Float(string='Dose', help="Amount of medicines (eg, 250 mg ) each time the patient takes it")
    # dose_unit_related = fields.Many2one('uom.uom', 'Unit of Measure')
    dose_unit_related = fields.Many2one('uom.uom',string='Dose',related='name.uom_id' ,help="Amount of medicines (eg, 250 mg ) each time the patient takes it")
    dose_unit = fields.Many2one('sh.medical.dose.unit', string='Dose Unit', help="Unit of measure for the medication to be taken")
    dose_route = fields.Many2one('sh.medical.drug.route', string='Administration Route', help="HL7 or other standard drug administration route code.")
    dose_form = fields.Many2one('sh.medical.drug.form','Form', help="Drug form, such as tablet or gel")
    qty = fields.Float(string='x', help="Quantity of units (eg, 2 capsules) of the medicament", default=lambda *a: 1.0)
    init_qty = fields.Float(string='x', help="Số lượng mặc định ban đầu theo dịch vụ", default=lambda *a: 0.0)
    common_dosage = fields.Many2one('sh.medical.dosage', string='Frequency', help="Common / standard dosage frequency for this medicines")
    frequency = fields.Integer('Frequency')
    frequency_unit = fields.Selection(FREQUENCY_UNIT, 'Unit', index=True)
    admin_times = fields.Char(string='Admin hours', size=128, help='Suggested administration hours. For example, at 08:00, 13:00 and 18:00 can be encoded like 08 13 18')
    duration = fields.Integer(string='Treatment duration')
    duration_period = fields.Selection(DURATION_UNIT, string='Treatment period', help="Period that the patient must take the medication. in minutes, hours, days, months, years or indefinately", index=True)
    start_treatment = fields.Datetime(string='Start of treatment')
    end_treatment = fields.Datetime('End of treatment')
    info = fields.Text('Comment')
    is_buy_out = fields.Boolean('Mua ngoài',default=False)
    patient = fields.Many2one('sh.medical.patient','Patient', help="Patient Name")

    national_prescription = fields.Boolean(string='Đơn thuốc quốc gia?', related='name.national_prescription')

    @api.onchange('qty', 'name')
    def onchange_qty(self):
        if self.qty <= 0 and self.name:
            raise UserError(_("Số lượng nhập phải lớn hơn 0!"))


# Vaccines Management
class SHealthVaccines(models.Model):
    _name = 'sh.medical.vaccines'
    _description = 'Vaccines'

    # Automatically detect logged in physician
    @api.multi
    def _get_physician(self):
        """Return default physician value"""
        therapist_obj = self.env['sh.medical.physician']
        domain = [('sh_user_id', '=', self.env.uid)]
        user_ids = therapist_obj.search(domain, limit=1)
        if user_ids:
            return user_ids.id or False
        else:
            return False

    name = fields.Many2one('sh.medical.medicines', string='Vaccine', domain=[('medicament_type','=','Vaccine')], required=True)
    patient = fields.Many2one('sh.medical.patient', string='Patient', help="Patient Name", required=True)
    doctor = fields.Many2one('sh.medical.physician', string='Physician', domain=[('is_pharmacist','=',False)], help="Current primary care / family doctor", required=True, default=_get_physician)
    date = fields.Datetime(string='Date', required=True, default=datetime.datetime.now())
    institution = fields.Many2one('sh.medical.health.center', string='Institution', help="Health Center where the patient is being or was vaccinated")
    dose = fields.Integer(string='Dose #', default=lambda *a: 1)
    info = fields.Text('Observation')

    @api.onchange('patient', 'name')
    def onchange_patient(self):
        res = {}
        if self.patient and self.name:
            dose = 0
            query = _("select max(dose) from sh_medical_vaccines where patient=%s and name=%s") % (str(self.patient.id), str(self.name.id))
            self.env.cr.execute(query)
            val = self.env.cr.fetchone()
            if val and val[0]:
                dose = int(val[0]) + 1
            else:
                dose = 1
            self.dose = dose
        return res

class Location(models.Model):
    _inherit = "stock.location"

    def should_bypass_reservation(self):
        self.ensure_one()
        # check neu mo lai phieu => trả vt về kho thì sẽ pass qua luôn để lấy dc data lô
        # flag = self.env.context.get('reopen_flag')
        # print(flag)
        # print(self.usage in ('supplier', 'inventory', 'production') or self.scrap_location)
        # if flag:
        #     return self.usage in ('supplier', 'inventory', 'production') or self.scrap_location
        # else:
        #     return self.usage in ('supplier', 'customer', 'inventory', 'production') or self.scrap_location
        # print(self.usage in ('supplier', 'inventory', 'production') or self.scrap_location)
        return self.usage in ('supplier', 'inventory', 'production') or self.scrap_location
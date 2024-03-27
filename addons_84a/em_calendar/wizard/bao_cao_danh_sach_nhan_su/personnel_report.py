from odoo import fields, api, models, _
from odoo.exceptions import ValidationError, UserError
from datetime import date, datetime, time
from openpyxl.styles import Font, borders, Alignment
import base64
from openpyxl import load_workbook, Workbook
from io import BytesIO
from openpyxl.styles import Font, borders, Alignment, PatternFill


thin = borders.Side(style='thin')
double = borders.Side(style='double')
all_border_thin = borders.Border(thin, thin, thin, thin)

class PersonnelReport(models.TransientModel):
    _name = 'personnel.report'
    _description = 'Báo cáo danh sách nhân sự'

    name = fields.Selection([('all', 'Tổng hợp nhân sự'), ('speciality', 'Theo chức vụ')], default="all")
    job_id = fields.Many2one(string='Chức vụ', comodel_name='sh.medical.speciality')

    def _get_data(self):
        datas = []
        """ Lấy ra tất cả thông tin nhân sự, bác sĩ, điều dưỡng viên """
        if self.name == 'speciality':
            moves = self.env['sh.medical.physician'].search([('speciality', '=', self.job_id.id)])
            em_physician = moves.mapped('department')
            # Phan loai sản phẩm
            dict_check = {
                'True': 'x',
                'False': None,
            }
            for rec in em_physician:
                product_lines = []  # for sorting
                product_index = {}
                index = 0
                physician_moves = moves.filtered(lambda m: rec in m.department)
                # total_physician = physician_moves.mapped('walkin')

                for physician_move in physician_moves:
                    if str(physician_move.id) not in product_index:
                        product_lines.append([None, index +1, physician_move.name or None, rec.name or None, physician_move.speciality.name or None, ', ' .join((rec.name for rec in physician_move.degree_id)) or None, physician_move.so_cchn or None, dict_check['%s' % physician_move.cc_qldd] or None, dict_check['%s' % physician_move.cc_qlbv] or None, dict_check['%s' % physician_move.cc_ksnk_48t] or None, dict_check['%s' % physician_move.cc_ksnk_3t] or None])
                        product_index[str(physician_move.id)] = index
                        index += 1

                service_lines = [rec.name]
                service_lines.extend(product_lines)
                datas.append(service_lines)

        else:
            moves = self.env['sh.medical.physician'].search([])
            em_physician = moves.mapped('department')
            # Phan loai sản phẩm
            dict_check = {
                'True': 'x',
                'False': None,
            }
            for rec in em_physician:
                product_lines = []  # for sorting
                product_index = {}
                index = 0
                physician_moves = moves.filtered(lambda m: rec in m.department)
                # total_physician = physician_moves.mapped('walkin')

                for physician_move in physician_moves:
                    if str(physician_move.id) not in product_index:
                        product_lines.append([None, index + 1, physician_move.name or None, rec.name or None,
                                              physician_move.speciality.name or None,
                                              ', '.join((rec.name for rec in physician_move.degree_id)) or None,
                                              physician_move.so_cchn or None,
                                              dict_check['%s' % physician_move.cc_qldd] or None,
                                              dict_check['%s' % physician_move.cc_qlbv] or None,
                                              dict_check['%s' % physician_move.cc_ksnk_48t] or None,
                                              dict_check['%s' % physician_move.cc_ksnk_3t] or None])
                        product_index[str(physician_move.id)] = index
                        index += 1

                service_lines = [rec.name]
                service_lines.extend(product_lines)
                datas.append(service_lines)

        return datas

    def create_report(self):
        inventory_attachment = self.env['ir.attachment'].sudo().browse(self.env.ref('em_calendar.bao_cao_danh_sach_nhan_su_attachment').id)
        decode = base64.b64decode(inventory_attachment.datas)
        wb = load_workbook(BytesIO(decode))
        ws = wb.active
        datas = self._get_data()


        line_font = Font(name='Times New Roman', size=12)
        line_font_st = Font(name='Times New Roman', size=12, bold=True)
        center_alm = Alignment(horizontal='center', vertical='center')
        row = 2
        end_col = 12
        for group in datas:
            ws.cell(row, 1).value, ws.cell(row, 1).font = group[0], line_font_st
            for col in range(1, end_col):
                ws.cell(row, col).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                ws.cell(row, 11).border = all_border_thin
            row += 1
            for i in range(1, len(group)):
                for col in range(1, end_col):
                    ws.cell(row, col).value, ws.cell(row, col).font = group[i][col-1], line_font
                for col in range(1, end_col):
                    ws.cell(row, col).border = all_border_thin
                for col in range(7, end_col):
                    ws.cell(row, col).alignment = center_alm
                row += 1

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({
            'name': 'Báo cáo danh sách nhân sự',
            'datas_fname': 'bao_cao_danh_sach_nhan_su.xlsx',
            'datas': report,
            'res_model': 'temp.creation',
            'public': True,
        })

        url = "/web/content/?model=ir.attachment&id=%s&filename_field=datas_fname&field=datas&download=true&filename=Báo cáo danh sách nhân sự.xlsx" \
              % (attachment.id)
        return {
            'name': 'ĐÁNH GIÁ CHẤT LƯỢNG DỊCH VỤ',
            'type': 'ir.actions.act_url',
            'url': url,
            'target': 'self',
        }

##############################################################################
#    Copyright (C) 2018 shealth (<http://scigroup.com.vn/>). All Rights Reserved
#    shealth, Hospital Management Solutions

# Odoo Proprietary License v1.0
#
# This software and associated files (the "Software") may only be used (executed,
# modified, executed after modifications) if you have purchased a valid license
# from the authors, typically via Odoo Apps, shealth.in, openerpestore.com, or if you have received a written
# agreement from the authors of the Software.
#
# You may develop Odoo modules that use the Software as a library (typically
# by depending on it, importing it and using its resources), but without copying
# any source code or material from the Software. You may distribute those
# modules under the license of your choice, provided that this license is
# compatible with the terms of the Odoo Proprietary License (For example:
# LGPL, MIT, or proprietary licenses similar to this one).
#
# It is forbidden to publish, distribute, sublicense, or sell copies of the Software
# or modified copies of the Software.
#
# The above copyright notice and this permission notice must be included in all
# copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT.
# IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM,
# DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE,
# ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER
# DEALINGS IN THE SOFTWARE.

##############################################################################
from datetime import date, datetime, time, timedelta
from dateutil.relativedelta import relativedelta
from odoo import api, fields, models, _
from odoo.exceptions import UserError,ValidationError,AccessError
from lxml import etree
from odoo.osv.orm import setup_modifiers
from odoo.tools.float_utils import float_round, float_compare, float_is_zero
from odoo.osv import expression
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.page import (
    PrintPageSetup,
    PageMargins,
    PrintOptions,
)
from openpyxl.styles import Font, borders, Alignment, PatternFill
from openpyxl.worksheet.pagebreak import Break
import base64
from io import BytesIO
import pytz

# Inherit Stock

class SHealthStockPicking(models.Model):
    _inherit = 'stock.location'

    location_institution_type = fields.Selection([('medicine', 'Tủ Thuốc'), ('supply', 'Vật tư')], string='Loại tủ')

class SHealthStockPicking(models.Model):
    _inherit = 'stock.picking'

    scheduled_date = fields.Datetime(default=lambda self: fields.Datetime.now())
    sci_date_done = fields.Datetime('Ngày hoàn thành', copy=False, help="Ngày hoàn thành nhập kho SCI", default=lambda self: fields.Datetime.now())
    hide_partner = fields.Boolean('Ẩn đối tác', compute="_compute_hide_partner")
    patient_id = fields.Many2one('sh.medical.patient', string='Bệnh nhân', help="Bệnh nhân")
    move_ids_without_package = fields.One2many(domain=['|', '&', ('state', '=', 'done'), ('quantity_done', '=', 0), '|', ('package_level_id', '=', False), ('picking_type_entire_packs', '=', False)])

    @api.multi
    def num2words_vnm(self, num):
        under_20 = ['không', 'một', 'hai', 'ba', 'bốn', 'năm', 'sáu', 'bảy', 'tám', 'chín', 'mười', 'mười một',
                    'mười hai', 'mười ba', 'mười bốn', 'mười lăm', 'mười sáu', 'mười bảy', 'mười tám', 'mười chín']
        tens = ['hai mươi', 'ba mươi', 'bốn mươi', 'năm mươi', 'sáu mươi', 'bảy mươi', 'tám mươi', 'chín mươi']
        above_100 = {100: 'trăm', 1000: 'nghìn', 1000000: 'triệu', 1000000000: 'tỉ'}

        num=int(num)
        if num < 20:
            return under_20[num].capitalize()

        elif num < 100:
            under_20[1], under_20[5] = 'mốt', 'lăm'  # thay cho một, năm
            result = tens[num // 10 - 2]
            if num % 10 > 0:  # nếu num chia 10 có số dư > 0 mới thêm ' ' và số đơn vị
                result += ' ' + under_20[num % 10]
            return result.capitalize()

        else:
            unit = max([key for key in above_100.keys() if key <= num])
            result = self.num2words_vnm(num // unit) + ' ' + above_100[unit]
            if num % unit != 0:
                if num > 1000 and num % unit < unit / 10:
                    result += ' không trăm'
                if 1 < num % unit < 10:
                    result += ' linh'
                result += ' ' + self.num2words_vnm(num % unit)
        return result.capitalize()

    @api.multi
    def button_validate(self):
        self.ensure_one()
        if not self.move_lines and not self.move_line_ids:
            raise UserError(_('Please add some items to move.'))
        # Thêm hàm check số lượng nhiều hơn khả dụng
        if self.move_ids_without_package:
            for rec in self.move_ids_without_package:
                if rec.check_quantity == True:
                    return {
                        'type': 'ir.actions.act_window',
                        'view_type': 'form',
                        'view_mode': 'form',
                        'res_model': 'stock.overprocessed.transfer',
                        'view_id': self.env.ref('shealth_all_in_one.view_overprocessed_transfer_inherited').id,
                        'target': 'new',
                        'context': self.env.context,
                    }

        # If no lots when needed, raise error
        picking_type = self.picking_type_id
        precision_digits = self.env['decimal.precision'].precision_get('Product Unit of Measure')
        no_quantities_done = all(float_is_zero(move_line.qty_done, precision_digits=precision_digits) for move_line in self.move_line_ids.filtered(lambda m: m.state not in ('done', 'cancel')))
        not_all_quantities_done = any(float_is_zero(move_line.qty_done, precision_digits=precision_digits) for move_line in self.move_line_ids.filtered(lambda m: m.state not in ('done', 'cancel')))
        no_reserved_quantities = all(float_is_zero(move_line.product_qty, precision_rounding=move_line.product_uom_id.rounding) for move_line in self.move_line_ids)
        if no_reserved_quantities and no_quantities_done:
            raise UserError(_('You cannot validate a transfer if no quantites are reserved nor done. To force the transfer, switch in edit more and encode the done quantities.'))

        if picking_type.use_create_lots or picking_type.use_existing_lots:
            lines_to_check = self.move_line_ids
            if not no_quantities_done:
                lines_to_check = lines_to_check.filtered(
                    lambda line: float_compare(line.qty_done, 0,
                                               precision_rounding=line.product_uom_id.rounding)
                )

            for line in lines_to_check:
                product = line.product_id
                if product and product.tracking != 'none':
                    if not line.lot_name and not line.lot_id:
                        raise UserError(_('You need to supply a Lot/Serial number for product %s.') % product.display_name)

        if not_all_quantities_done:
            view = self.env.ref('stock.view_immediate_transfer')
            wiz = self.env['stock.immediate.transfer'].create({'pick_ids': [(4, self.id)]})
            return {
                'name': _('Immediate Transfer?'),
                'type': 'ir.actions.act_window',
                'view_type': 'form',
                'view_mode': 'form',
                'res_model': 'stock.immediate.transfer',
                'views': [(view.id, 'form')],
                'view_id': view.id,
                'target': 'new',
                'res_id': wiz.id,
                'context': self.env.context,
            }

        if self._get_overprocessed_stock_moves() and not self._context.get('skip_overprocessed_check'):
            view = self.env.ref('stock.view_overprocessed_transfer')
            wiz = self.env['stock.overprocessed.transfer'].create({'picking_id': self.id})
            return {
                'type': 'ir.actions.act_window',
                'view_type': 'form',
                'view_mode': 'form',
                'res_model': 'stock.overprocessed.transfer',
                'views': [(view.id, 'form')],
                'view_id': view.id,
                'target': 'new',
                'res_id': wiz.id,
                'context': self.env.context,
            }

        # Check backorder should check for other barcodes
        if self._check_backorder():
            return self.action_generate_backorder_wizard()
        self.action_done()
        return

    @api.multi
    def action_assign(self):
        """Adjustment to bypass waiting state"""
        picks = self.filtered(lambda p: p.state == 'waiting')
        picks.mapped('move_lines').write({'move_orig_ids': [(5, 0, 0)], 'procure_method': 'make_to_stock'})
        return super(SHealthStockPicking, self).action_assign()

    @api.multi
    def action_cancel(self):
        # KHÔNG CHO HỦY PHIẾU CHƯA CÓ SẢN PHẨM NÀO DC CHỌN
        for picking in self:
            if picking.move_ids_without_package or self.env.context.get('no_check_quant') or self.env.uid == 1:
                return super(SHealthStockPicking, self).action_cancel()
            else:
                raise UserError(_('Phiếu của bạn không có sản phẩm nào để hủy bỏ!'))

    @api.multi
    def action_confirm(self):
        # KHÔNG CHO XÁC NHẬN PHIẾU CHƯA CÓ SẢN PHẨM NÀO DC CHỌN
        for picking in self:
            if picking.move_ids_without_package or self.env.context.get('no_check_quant') or self.env.uid == 1:
                return super(SHealthStockPicking, self).action_confirm()
            else:
                raise UserError(_('Phiếu của bạn không có sản phẩm nào để xác nhận!'))

    # @api.multi
    # def action_assign(self):
    #     res = super(SHealthStockPicking, self).action_assign()
    #     for picking in self:
    #         validate_str = ''
    #         for move in picking.move_ids_without_package:
    #             if move.reserved_availability == 0:
    #                 validate_str += '+ %s đang có số lượng khả dụng trong kho bằng 0\n' % move.product_id.name
    #
    #         if validate_str !='':
    #             raise UserError(_(validate_str))
    #     return res

    @api.multi
    def write(self, vals):
        move_vals = vals.get('move_ids_without_package')
        if move_vals:
            removed_move_ids = [move_val[1] for move_val in move_vals if move_val[0] == 2]
            removed_moves = self.env['stock.move'].browse(removed_move_ids)
            removed_moves._do_unreserve()
            removed_moves.write({'state': 'done', 'sequence': 999})
            remaining_move_vals = [move_val for move_val in move_vals if move_val[0] != 2]
            vals['move_ids_without_package'] = remaining_move_vals
        # raise UserError('Staphhh')
        return super(SHealthStockPicking, self).write(vals)

    @api.multi
    def download_excel(self):
        for record in self:
            if record.state != 'cancel':
                self.action_cancel()
        return {
            'name': 'Nhu cầu nhập hàng',
            'type': 'ir.actions.act_window',
            'view_type': 'form',
            'view_mode': 'form',
            'res_model': 'temp.wizard',
            'view_id': self.env.ref('ms_templates.report_wizard').id,
            'target': 'inline',
            'context': {'default_template_id': self.env.ref('shealth_all_in_one.nhu_cau_nhap_hang').id},
        }

    @api.depends('picking_type_id')
    def _compute_hide_partner(self):
        for pick in self:
            if pick.picking_type_id.id == self.env.ref('stock.warehouse0').int_type_id.id:
                pick.hide_partner = True
            else:
                pick.hide_partner = False

    @api.onchange('picking_type_id')
    def onchange_picking_type(self):
        super(SHealthStockPicking, self).onchange_picking_type()
        if self.env.context.get('view_for', True) == 'picking_int':
            self.location_dest_id = False
            self.location_id = self.env.ref(
                'shealth_all_in_one.sh_location_medicine_prescription_knhn').id
        elif self.env.context.get('view_for', True) == 'picking_int_vt':
            self.location_dest_id = False
            self.location_id = self.env.ref(
                'shealth_all_in_one.sh_location_supply_knhn').id
        elif self.env.context.get('view_for', True) == 'picking_int_return':
            self.location_dest_id = self.env.ref(
                'shealth_all_in_one.sh_location_medicine_prescription_knhn').id
            self.location_id = False

    @api.onchange('location_id')
    def onchange_location_id(self):
        if self.env.context.get('view_for', True) == 'picking_int_return':
            for rec in self:
                if rec.location_id.location_institution_type == 'medicine':
                    rec.location_dest_id = self.env.ref(
                        'shealth_all_in_one.sh_location_medicine_prescription_knhn').id
                elif rec.location_id.location_institution_type == 'supply':
                    rec.location_dest_id = self.env.ref('shealth_all_in_one.sh_location_supply_knhn').id
    @api.model
    def fields_view_get(self, view_id=None, view_type='form', toolbar=False, submenu=False):
        res = super(SHealthStockPicking, self).fields_view_get(view_id=view_id, view_type=view_type, toolbar=toolbar,
                                                      submenu=submenu)

        access_location = []
        # NẾU QUẢN LÝ KHO HOẶC ADMIN HOẶC NHẬP LIỆU => XEM TẤT CẢ LOCATION
        if self.env.user.has_group('shealth_all_in_one.group_sh_medical_stock_manager'):
            location_stock0 = self.env.ref('stock.warehouse0').lot_stock_id.id  # dia diem kho tổng
            access_location.append(location_stock0)
        else:
            grp_loc_dict = {
                'shealth_all_in_one.group_sh_medical_physician_subclinical_labtest': 'shealth_all_in_one.sh_labtest_dep_knhn',
                # 'shealth_all_in_one.group_sh_medical_physician_subclinical_imaging': 'shealth_all_in_one.sh_imaging_dep_knhn',
                'shealth_all_in_one.group_sh_medical_physician_surgery': 'shealth_all_in_one.sh_surgeries_dep_knhn',
                'shealth_all_in_one.group_sh_medical_physician_odontology': 'shealth_all_in_one.sh_rhm_dep_knhn',
                'shealth_all_in_one.group_sh_medical_physician_spa': 'shealth_all_in_one.sh_dalieu_dep_knhn'}
            for grp, loc in grp_loc_dict.items():
                if self.env.user.has_group(grp):
                    access_location.append(self.env.ref(loc).location_id.id)

            # quyen dieu duong
            if self.env.user.has_group('shealth_all_in_one.group_sh_medical_nurse') and self.env.user.physician_ids:
                physician_loc = self.env.user.physician_ids[0].department.mapped('location_id').ids
                access_location += physician_loc

        doc = etree.XML(res['arch'])

        for t in doc.xpath("//" + view_type):
            t.attrib['delete'] = 'false'
            t.attrib['duplicate'] = 'false'

        if self.env.context.get('view_for', True) == "picking_in":
            #ko cho tao phieu
            for t in doc.xpath("//" + view_type):
                t.attrib['create'] = 'false'

            for node in doc.xpath("//field[@name='date_done']"):
                node.set('string', 'Ngày nhập kho')
            for node in doc.xpath("//field[@name='partner_id']"):
                node.set('string', 'Nhà cung cấp')
            for node in doc.xpath("//field[@name='origin']"):
                node.set('invisible', '1')
                if view_type == 'tree':
                    setup_modifiers(node, res['fields']['origin'], in_tree_view=True)
                else:
                    setup_modifiers(node, res['fields']['origin'])
            # for node in doc.xpath("//field[@name='location_id']"):
            #     node.set('invisible', '1')
            #     if view_type == 'tree':
            #         setup_modifiers(node, res['fields']['location_id'], in_tree_view=True)
            #     else:
            #         setup_modifiers(node, res['fields']['location_id'])
            for node in doc.xpath("//field[@name='purchase_id']"):
                node.set('invisible', '0')
                setup_modifiers(node, res['fields']['purchase_id'])
        # elif self.env.context.get('view_for', True) in ["picking_int","picking_int_return"]:
        elif self.env.context.get('view_for', True) in ["picking_int","picking_int_vt"]:
            #neu quan ly kho duoc
            # if self.env.user.has_group('shealth_all_in_one.group_sh_medical_document') \
            #         or self.env.user.has_group('shealth_all_in_one.group_sh_medical_stock_manager'):
            #     #cho phép sửa phiếu
            #     for t in doc.xpath("//" + view_type):
            #         t.attrib['edit'] = 'true'
            # else:
            #     # ko cho phép sửa phiếu
            #     for t in doc.xpath("//" + view_type):
            #         t.attrib['edit'] = 'false'

            for node in doc.xpath("//field[@name='partner_id']"):
                node.set('invisible', '1')
                if view_type == 'tree':
                    setup_modifiers(node, res['fields']['partner_id'], in_tree_view=True)
            for node in doc.xpath("//field[@name='location_id']"):
                # node.set('readonly', 'True')
                if self.env.context.get('view_for', True) == 'picking_int':
                    node.set('domain', "[('id', '=', %d)]" % self.env.ref(
                        'shealth_all_in_one.sh_location_medicine_prescription_knhn').id)
                else:
                    node.set('domain', "[('id', '=', %d)]" % self.env.ref(
                        'shealth_all_in_one.sh_location_supply_knhn').id)
                setup_modifiers(node, res['fields']['location_id'])
            for node in doc.xpath("//field[@name='location_dest_id']"):
                if self.env.context.get('view_for', True) == 'picking_int':
                    node.set('domain', "[('location_id', 'child_of', %s),('name', 'ilike', 'Thuốc')]" % str(access_location))
                else:
                    node.set('domain', "[('location_id', 'child_of', %s),('name', 'ilike', 'VT')]" % str(access_location))
                setup_modifiers(node, res['fields']['location_dest_id'])
        elif self.env.context.get('view_for', True) in ["picking_int_return"]:
            for node in doc.xpath("//field[@name='partner_id']"):
                node.set('invisible', '1')
                if view_type == 'tree':
                    setup_modifiers(node, res['fields']['partner_id'], in_tree_view=True)
            for node in doc.xpath("//field[@name='location_id']"):
                node_domain = "[('location_id', 'child_of', %s),('location_id.usage', '=', 'internal'),('name', 'ilike', 'Tủ')]" % (
                    str(access_location))
                node.set("domain", node_domain)
                setup_modifiers(node, res['fields']['location_id'])
            for node in doc.xpath("//field[@name='location_dest_id']"):
                node_domain = "[('id', 'in', {})]".format((self.env.ref(
                    'shealth_all_in_one.sh_location_medicine_prescription_knhn').id, self.env.ref(
                    'shealth_all_in_one.sh_location_supply_knhn').id))
                node.set("domain", node_domain)
                # node.set('readonly', '1')
                setup_modifiers(node, res['fields']['location_dest_id'])
        elif self.env.context.get('view_for', True) == "picking_out_bn":
            for t in doc.xpath("//" + view_type):
                t.attrib['create'] = 'false'

            for node in doc.xpath("//field[@name='partner_id']"):
                node.set('invisible', '1')
                if view_type == 'tree':
                    setup_modifiers(node, res['fields']['partner_id'], in_tree_view=True)
                else:
                    setup_modifiers(node, res['fields']['partner_id'])
            for node in doc.xpath("//field[@name='patient_id']"):
                node.set('invisible', '0')
                setup_modifiers(node, res['fields']['patient_id'])
                # if view_type == 'tree':
                #     setup_modifiers(node, res['fields']['patient_id'], in_tree_view=True)
            for node in doc.xpath("//field[@name='location_dest_id']"):
                node.set('invisible', '1')
                if view_type == 'tree':
                    setup_modifiers(node, res['fields']['location_dest_id'], in_tree_view=True)

        res['arch'] = etree.tostring(doc, encoding='unicode')
        return res

    #overide hàm này để ghi nhận lại ngày hoàn thành
    @api.multi
    def action_done(self):
        res = super(SHealthStockPicking, self).action_done()

        for stock_picking in self:
            if stock_picking.sci_date_done:
                for move_line in stock_picking.move_ids_without_package:
                    move_line.move_line_ids.write({'date': stock_picking.sci_date_done})  # sửa ngày hoàn thành ở stock move line
                stock_picking.move_ids_without_package.write(
                    {'date': stock_picking.sci_date_done})  # sửa ngày hoàn thành ở stock move

                stock_picking.write({'date_done': stock_picking.sci_date_done})
            else:
                stock_picking.write({'sci_date_done': stock_picking.date_done})

        return res

    def view_current_stock(self):
        #mặc định xem tủ kê đơn và cấp cứu vs tất cả các quyền
        access_location = []

        # NẾU QUẢN LÝ KHO HOẶC ADMIN HOẶC NHẬP LIỆU => XEM TẤT CẢ LOCATION
        if self.env.user.has_group('shealth_all_in_one.group_sh_medical_stock_manager'):
            location_stock0 = self.env.ref('stock.warehouse0').lot_stock_id.id  # dia diem kho tổng
            access_location.append(location_stock0)
        else:
            grp_loc_dict = {'shealth_all_in_one.group_sh_medical_physician_subclinical_labtest': 'shealth_all_in_one.sh_labtest_dep_knhn',
                            # 'shealth_all_in_one.group_sh_medical_physician_subclinical_imaging': 'shealth_all_in_one.sh_imaging_dep_knhn',
                            'shealth_all_in_one.group_sh_medical_physician_surgery': 'shealth_all_in_one.sh_surgeries_dep_knhn',
                            'shealth_all_in_one.group_sh_medical_physician_odontology': 'shealth_all_in_one.sh_rhm_dep_knhn',
                            'shealth_all_in_one.group_sh_medical_physician_spa': 'shealth_all_in_one.sh_dalieu_dep_knhn'}
            for grp, loc in grp_loc_dict.items():
                if self.env.user.has_group(grp):
                    access_location.append(self.env.ref(loc).location_id.id)

            # quyen bác sĩ chung
            if self.env.user.has_group('shealth_all_in_one.group_sh_medical_physician'):
                access_location.append(self.env.ref('shealth_all_in_one.sh_location_medicine_prescription_knhn').id)
                access_location.append(self.env.ref('shealth_all_in_one.sh_location_medicine_emergency_knhn').id)

            #quyen dieu duong
            if self.env.user.has_group('shealth_all_in_one.group_sh_medical_nurse') and self.env.user.physician_ids:
                physician_loc = self.env.user.physician_ids[0].department.mapped('location_id').ids
                access_location += physician_loc

        return {
            'name': 'Dự trữ hiện tại',
            'type': 'ir.actions.act_window',
            'view_type': 'form',
            'view_mode': 'tree',
            'domain': [('location_id', 'child_of', access_location)],
            'context': {'search_default_locationgroup': True, 'search_default_productgroup': True,'view_only_name':True},
            'res_model': 'stock.quant',
            'view_id': self.env.ref('shealth_all_in_one.sci_current_stock_quant_tree').id,
            'target': 'current'
        }

class SHealthProductionLot(models.Model):
    _inherit = 'stock.production.lot'

    product_init_qty = fields.Float('Initial quantity')
    lot_price = fields.Float('Lot price')
    product_qty = fields.Float('Số lượng thực tế')
    reserved_qty = fields.Float('Số lượng giữ trước', compute="_compute_quant_ids")

    api.depends('quant_ids')
    def _compute_quant_ids(self):
        for lot in self:
            if lot.quant_ids:
                lot.reserved_qty = sum(lot.quant_ids.mapped('reserved_quantity'))

class SHealthStockMove(models.Model):
    _inherit = 'stock.move'

    check_quantity = fields.Boolean('Check sô lượng tồn kho', default=False, compute='compute_check_quantity', store=True)
    # Check số lượng khả dụng
    @api.depends('move_line_ids')
    def compute_check_quantity(self):
        for rec in self:
            quantity = rec.move_line_ids
            for record in quantity:
                if record.qty_done > record.product_uom_qty:
                    rec.check_quantity = True

    def _get_domain_product_id(self):
        if self.env.context.get('view_for', True) in ["picking_int", "picking_int_vt"]:
            if self.env.context.get('view_for', True) == 'picking_int':
                return [('categ_id', '=', self.env.ref('shealth_all_in_one.sh_medicines').id)]
            else:
                return [('categ_id', '=', self.env.ref('shealth_all_in_one.sh_supplies').id)]
        else:
            return [('type', 'in', ['product', 'consu'])]

    product_id = fields.Many2one(domain=lambda self: self._get_domain_product_id())

    @api.onchange('product_uom_qty', 'product_id')
    def onchange_product_uom_qty(self):
        if self.product_id:
            if self.product_uom_qty <= 0:
                self.product_uom_qty = 1

            if self.product_uom_qty <= 0:
                raise UserError(_("Số lượng nhập phải lớn hơn 0!"))

# class SHealthStockLocation(models.Model):
#     _inherit = 'stock.location'
#
#     def _get_location_id_by_user(self):
#         print(self.env.user.name)
#         access_location = []
#         # NẾU QUẢN LÝ KHO HOẶC ADMIN HOẶC NHẬP LIỆU => XEM TẤT CẢ LOCATION
#         if self.env.user.has_group('shealth_all_in_one.group_sh_medical_document') \
#                 or self.env.user.has_group('shealth_all_in_one.group_sh_medical_stock_manager'):
#             location_stock0 = self.env.ref('stock.warehouse0').lot_stock_id.id  # dia diem kho tổng
#             access_location.append(location_stock0)
#         else:
#             grp_loc_dict = {
#                 'shealth_all_in_one.group_sh_medical_physician_subclinical_labtest': 'shealth_all_in_one.sh_labtest_dep_knhn',
#                 'shealth_all_in_one.group_sh_medical_physician_subclinical_imaging': 'shealth_all_in_one.sh_imaging_dep_knhn',
#                 'shealth_all_in_one.group_sh_medical_physician_surgery': 'shealth_all_in_one.sh_surgeries_dep_knhn',
#                 'shealth_all_in_one.group_sh_medical_physician_odontology': 'shealth_all_in_one.sh_rhm_dep_knhn',
#                 'shealth_all_in_one.group_sh_medical_physician_spa': 'shealth_all_in_one.sh_dalieu_dep_knhn'}
#             for grp, loc in grp_loc_dict.items():
#                 if self.env.user.has_group(grp):
#                     access_location.append(self.env.ref(loc).location_id.id)
#
#             # quyen dieu duong
#             if self.env.user.has_group('shealth_all_in_one.group_sh_medical_nurse') and self.env.user.physician_ids:
#                 physician_loc = self.env.user.physician_ids[0].department.mapped('location_id').ids
#                 access_location += physician_loc
#         return access_location

class SHealthStockMoveLine(models.Model):
    _inherit = 'stock.move.line'

    def _action_done(self):
        """ This method is called during a move's `action_done`. It'll actually move a quant from
        the source location to the destination location, and unreserve if needed in the source
        location.

        This method is intended to be called on all the move lines of a move. This method is not
        intended to be called when editing a `done` move (that's what the override of `write` here
        is done.
        """
        Quant = self.env['stock.quant']

        # First, we loop over all the move lines to do a preliminary check: `qty_done` should not
        # be negative and, according to the presence of a picking type or a linked inventory
        # adjustment, enforce some rules on the `lot_id` field. If `qty_done` is null, we unlink
        # the line. It is mandatory in order to free the reservation and correctly apply
        # `action_done` on the next move lines.
        ml_to_delete = self.env['stock.move.line']
        for ml in self:
            # Check here if `ml.qty_done` respects the rounding of `ml.product_uom_id`.
            uom_qty = float_round(ml.qty_done, precision_rounding=ml.product_uom_id.rounding, rounding_method='HALF-UP')
            precision_digits = self.env['decimal.precision'].precision_get('Product Unit of Measure')
            qty_done = float_round(ml.qty_done, precision_digits=precision_digits, rounding_method='HALF-UP')
            if float_compare(uom_qty, qty_done, precision_digits=precision_digits) != 0:
                raise UserError(_('The quantity done for the product "%s" doesn\'t respect the rounding precision \
                                  defined on the unit of measure "%s". Please change the quantity done or the \
                                  rounding precision of your unit of measure.') % (ml.product_id.display_name, ml.product_uom_id.name))

            qty_done_float_compared = float_compare(ml.qty_done, 0, precision_rounding=ml.product_uom_id.rounding)
            if qty_done_float_compared > 0:
                if ml.product_id.tracking != 'none':
                    picking_type_id = ml.move_id.picking_type_id
                    if picking_type_id:
                        if picking_type_id.use_create_lots:
                            # If a picking type is linked, we may have to create a production lot on
                            # the fly before assigning it to the move line if the user checked both
                            # `use_create_lots` and `use_existing_lots`.
                            if ml.lot_name and not ml.lot_id:
                                lot_price = 0
                                if ml.picking_id.purchase_id:
                                    lot_price = ml.picking_id.purchase_id.order_line.filtered(lambda o: o.product_id == ml.product_id)[0].price_subtotal
                                lot = self.env['stock.production.lot'].create(
                                    {'name': ml.lot_name, 'product_id': ml.product_id.id,
                                     'product_init_qty': ml.qty_done, 'lot_price': lot_price}
                                )
                                ml.write({'lot_id': lot.id})
                        elif not picking_type_id.use_create_lots and not picking_type_id.use_existing_lots:
                            # If the user disabled both `use_create_lots` and `use_existing_lots`
                            # checkboxes on the picking type, he's allowed to enter tracked
                            # products without a `lot_id`.
                            continue
                    elif ml.move_id.inventory_id:
                        # If an inventory adjustment is linked, the user is allowed to enter
                        # tracked products without a `lot_id`.
                        continue

                    if not ml.lot_id:
                        raise UserError(_('You need to supply a Lot/Serial number for product %s.') % ml.product_id.display_name)
            elif qty_done_float_compared < 0:
                raise UserError(_('No negative quantities allowed'))
            else:
                ml_to_delete |= ml
        ml_to_delete.unlink()

        # Now, we can actually move the quant.
        done_ml = self.env['stock.move.line']
        for ml in self - ml_to_delete:
            if ml.product_id.type == 'product':
                rounding = ml.product_uom_id.rounding

                # if this move line is force assigned, unreserve elsewhere if needed
                if not ml.location_id.should_bypass_reservation() and float_compare(ml.qty_done, ml.product_qty, precision_rounding=rounding) > 0:
                    extra_qty = ml.qty_done - ml.product_qty
                    ml._free_reservation(ml.product_id, ml.location_id, extra_qty, lot_id=ml.lot_id, package_id=ml.package_id, owner_id=ml.owner_id, ml_to_ignore=done_ml)
                # unreserve what's been reserved
                if not ml.location_id.should_bypass_reservation() and ml.product_id.type == 'product' and ml.product_qty:
                    try:
                        Quant._update_reserved_quantity(ml.product_id, ml.location_id, -ml.product_qty, lot_id=ml.lot_id, package_id=ml.package_id, owner_id=ml.owner_id, strict=True)
                    except UserError:
                        Quant._update_reserved_quantity(ml.product_id, ml.location_id, -ml.product_qty, lot_id=False, package_id=ml.package_id, owner_id=ml.owner_id, strict=True)

                # move what's been actually done
                quantity = ml.product_uom_id._compute_quantity(ml.qty_done, ml.move_id.product_id.uom_id, rounding_method='HALF-UP')
                available_qty, in_date = Quant._update_available_quantity(ml.product_id, ml.location_id, -quantity, lot_id=ml.lot_id, package_id=ml.package_id, owner_id=ml.owner_id)
                if available_qty < 0 and ml.lot_id:
                    # see if we can compensate the negative quants with some untracked quants
                    untracked_qty = Quant._get_available_quantity(ml.product_id, ml.location_id, lot_id=False, package_id=ml.package_id, owner_id=ml.owner_id, strict=True)
                    if untracked_qty:
                        taken_from_untracked_qty = min(untracked_qty, abs(quantity))
                        Quant._update_available_quantity(ml.product_id, ml.location_id, -taken_from_untracked_qty, lot_id=False, package_id=ml.package_id, owner_id=ml.owner_id)
                        Quant._update_available_quantity(ml.product_id, ml.location_id, taken_from_untracked_qty, lot_id=ml.lot_id, package_id=ml.package_id, owner_id=ml.owner_id)
                Quant._update_available_quantity(ml.product_id, ml.location_dest_id, quantity, lot_id=ml.lot_id, package_id=ml.result_package_id, owner_id=ml.owner_id, in_date=in_date)
            done_ml |= ml
        # Reset the reserved quantity as we just moved it to the destination location.
        (self - ml_to_delete).with_context(bypass_reservation_update=True).write({
            'product_uom_qty': 0.00,
            'date': fields.Datetime.now(),
        })

#BIẾN CỤC BỘ STYLE EXCEL
thin = borders.Side(style='thin')
dotted = borders.Side(style='hair')
gray_thin = borders.Side(style='thin', color='808080')
all_border_thin = borders.Border(left=thin, right=thin, top=thin, bottom=thin)
all_border_gray = borders.Border(left=gray_thin, right=gray_thin, top=gray_thin, bottom=gray_thin)
dotted_top_bot = borders.Border(left=thin, right=thin, top=dotted, bottom=dotted)

# KIỂM KÊ
class SHealthInventory(models.Model):
    _inherit = "stock.inventory"

    @api.model
    def fields_view_get(self, view_id=None, view_type='form', toolbar=False, submenu=False):
        res = super(SHealthInventory, self).fields_view_get(view_id=view_id, view_type=view_type, toolbar=toolbar,
                                                               submenu=submenu)

        doc = etree.XML(res['arch'])

        for node in doc.xpath("//field[@name='category_id']"):
            node.set('domain',
                         "[('id', 'in', [%d,%d])]" % (self.env.ref('shealth_all_in_one.sh_medicines').id, self.env.ref('shealth_all_in_one.sh_supplies').id))
            setup_modifiers(node, res['fields']['category_id'])

        res['arch'] = etree.tostring(doc, encoding='unicode')
        return res

    def action_start(self):
        for inventory in self.filtered(lambda x: x.state not in ('done','cancel')):
            vals = {'state': 'confirm', 'date': inventory.date}
            if (inventory.filter != 'partial') and not inventory.line_ids:
                vals.update({'line_ids': [(0, 0, line_values) for line_values in inventory._get_inventory_lines_values()]})
            inventory.write(vals)
        return True

    def _action_done(self):
        res = super(SHealthInventory, self)._action_done()

        for stock_inventory in self:
            if stock_inventory.date:
                for move_line in stock_inventory.move_ids:
                    move_line.move_line_ids.write(
                        {'date': stock_inventory.date})  # sửa ngày hoàn thành ở stock move line
                stock_inventory.move_ids.write(
                    {'date': stock_inventory.date})  # sửa ngày hoàn thành ở stock move

        return res

    # IN PHIẾU KIỂM KÊ KHO
    # @api.multi
    def print_stock_inventory(self):
        wb = Workbook()
        ws = wb.active

        row = 1
        for inventory in self:
            # inventory_attachment = self.env['ir.attachment'].browse(
            #     self.env.ref('shealth_all_in_one.sci_stock_inventory_report_attachment').id)
            # decode = base64.b64decode(inventory_attachment.datas)

            main_font = Font(name='Times New Roman', size=12)
            font_11 = Font(name='Times New Roman', size=11)
            font_14 = Font(name='Times New Roman', size=14)
            font_14_bold = Font(name='Times New Roman', size=14, bold=True)
            font_12_bold = Font(name='Times New Roman', size=12, bold=True)
            alignment_center = Alignment(horizontal='center', vertical='center')

            health_center = self.env['sh.medical.health.center'].sudo().search([('company_id', '=',
                                                                                 inventory.company_id.id)], limit=1)
            ws.cell(row=row, column=2).value, ws.cell(row=row, column=2).font = \
                inventory.company_id.name.upper() or 'CTY CP BVTM KANGNAM HÀ NỘI', font_11
            ws.cell(row=row, column=10).value, ws.cell(row=row, column=10).font = 'MS: 11D/BV-01', font_11
            ws.cell(row=row+1, column=2).value, ws.cell(row=row+1, column=2).font = \
                health_center[0].name.upper() or 'BỆNH VIỆN THẨM MỸ KANGNAM HÀ NỘI', font_11
            ws.cell(row=row + 1, column=10).value, ws.cell(row=row + 1, column=10).font = 'Số', font_11

            ws.merge_cells(start_row=row+2, start_column=2, end_row=row+2, end_column=11)
            ws.cell(row=row + 2, column=2).value, ws.cell(row=row + 2, column=2).font, ws.cell(row=row + 2, column=2).alignment = 'BIÊN BẢN KIỂM KÊ %s %s' % (
            inventory.location_id.name.upper(), inventory.date.strftime('THÁNG %m NĂM %Y')), Font(name='Times New Roman', size=18, bold=True),alignment_center

            ws.cell(row=row + 4, column=3).value, ws.cell(row=row + 4, column=3).font = 'Tổ kiểm kê gồm có', font_14_bold
            ws.cell(row=row + 5, column=2).value, ws.cell(row=row + 5, column=2).font = 1, font_14
            ws.cell(row=row + 5, column=3).value, ws.cell(row=row + 5, column=3).font = \
                self.env.ref('__import__.data_physician_duoc').name if self.env.ref('__import__.data_physician_duoc',
                                                                                    False) else 'DS. Trưởng khoa dược', font_14
                # self.env.ref('__import__.data_physician_duoc').name or 'DS. Trưởng khoa dược', font_14
            ws.merge_cells(start_row=row + 5, start_column=6, end_row=row + 5, end_column=7)
            ws.cell(row=row + 5, column=5).value, ws.cell(row=row + 5, column=5).font = 'Chức danh:', font_14
            ws.cell(row=row + 5, column=6).value, ws.cell(row=row + 5, column=6).font = \
                self.env.ref('__import__.data_physician_duoc').job_id.name if self.env.ref('__import__.data_physician_duoc',
                                                                                    False) else 'DS. Trưởng khoa dược', font_14
                # self.env.ref('__import__.data_physician_duoc').job_id.name or 'Trưởng Khoa Dược', font_14

            ws.cell(row=row + 6, column=2).value, ws.cell(row=row + 6, column=2).font = 2, font_14
            ws.cell(row=row + 6, column=3).value, ws.cell(row=row + 6, column=3).font = \
                self.env.ref('__import__.data_physician_duoc_si').name if self.env.ref('__import__.data_physician_duoc_si',
                                                                                    False) else'DS. Dược sĩ viên', font_14
            ws.merge_cells(start_row=row + 6, start_column=6, end_row=row + 6, end_column=7)
            ws.cell(row=row + 6, column=5).value, ws.cell(row=row + 6, column=5).font = 'Chức danh:', font_14
            ws.cell(row=row + 6, column=6).value, ws.cell(row=row + 6, column=6).font = 'Dược sĩ', font_14

            ws.merge_cells(start_row=row + 8, start_column=3, end_row=row + 8, end_column=11)
            ws.cell(row=row + 8, column=3).value, ws.cell(row=row + 8, column=3).font = 'Đã kiểm kê tại %s %s từ %s' % \
                             (inventory.location_id.name, health_center[0].name,
                              inventory.date.strftime('%H giờ %M ngày %d tháng %m năm %Y')), font_14

            ws.cell(row=row + 10, column=3).value, ws.cell(row=row + 10,
                                                          column=3).font = 'Kết quả như sau:', font_14_bold

            ws.merge_cells(start_row=row + 12, start_column=2, end_row=row + 13, end_column=2)
            ws.cell(row=row+12, column=2).value, ws.cell(row=row+12, column=2).font = 'STT', font_12_bold
            ws.cell(row=row+12, column=2).border, ws.cell(row=row+12, column=2).alignment = all_border_thin, alignment_center
            ws.cell(row=row+13, column=2).border = all_border_thin

            ws.merge_cells(start_row=row + 12, start_column=3, end_row=row + 13, end_column=3)
            ws.cell(row=row + 12, column=3).value, ws.cell(row=row + 12, column=3).font = 'Tên thuốc, nồng độ, hàm lượng', font_12_bold
            ws.cell(row=row + 12, column=3).border, ws.cell(row=row + 12,
                                                            column=3).alignment = all_border_thin, alignment_center

            ws.merge_cells(start_row=row + 12, start_column=4, end_row=row + 13, end_column=4)
            ws.cell(row=row + 12, column=4).value, ws.cell(row=row + 12,
                                                           column=4).font = 'Đơn vị', font_12_bold
            ws.cell(row=row + 12, column=4).border, ws.cell(row=row + 12,
                                                            column=4).alignment = all_border_thin, alignment_center
            ws.cell(row=row + 13, column=4).border = all_border_thin

            ws.merge_cells(start_row=row + 12, start_column=5, end_row=row + 13, end_column=5)
            ws.cell(row=row + 12, column=5).value, ws.cell(row=row + 12,
                                                           column=5).font = 'Số Lô/Sê-ri', font_12_bold
            ws.cell(row=row + 12, column=5).border, ws.cell(row=row + 12,
                                                            column=5).alignment = all_border_thin, alignment_center
            ws.cell(row=row + 13, column=5).border = all_border_thin

            ws.merge_cells(start_row=row + 12, start_column=6, end_row=row + 13, end_column=6)
            ws.cell(row=row + 12, column=6).value, ws.cell(row=row + 12,
                                                           column=6).font = 'Hãng, nước sản xuất', font_12_bold
            ws.cell(row=row + 12, column=6).border, ws.cell(row=row + 12,
                                                            column=6).alignment = all_border_thin, Alignment(
                    horizontal='center', vertical='center', wrap_text=True)
            ws.cell(row=row + 13, column=6).border = all_border_thin

            ws.merge_cells(start_row=row + 12, start_column=7, end_row=row + 13, end_column=7)
            ws.cell(row=row + 12, column=7).value, ws.cell(row=row + 12,
                                                           column=7).font = 'Hạn dùng', font_12_bold
            ws.cell(row=row + 12, column=7).border, ws.cell(row=row + 12,
                                                            column=7).alignment = all_border_thin, alignment_center

            ws.merge_cells(start_row=row + 12, start_column=8, end_row=row + 12, end_column=9)
            ws.cell(row=row + 12, column=8).value, ws.cell(row=row + 12,
                                                           column=8).font = 'Số lượng', font_12_bold
            ws.cell(row=row + 12, column=8).border, ws.cell(row=row + 12,
                                                            column=8).alignment = all_border_thin, alignment_center
            ws.cell(row=row + 12, column=9).border = all_border_thin

            ws.cell(row=row + 13, column=8).value, ws.cell(row=row + 13,
                                                           column=8).font = 'Sổ sách', font_12_bold
            ws.cell(row=row + 13, column=8).border, ws.cell(row=row + 13,
                                                            column=8).alignment = all_border_thin, alignment_center
            ws.cell(row=row + 13, column=9).value, ws.cell(row=row + 13,
                                                           column=9).font = 'Thực tế', font_12_bold
            ws.cell(row=row + 13, column=9).border, ws.cell(row=row + 13,
                                                            column=9).alignment = all_border_thin, alignment_center

            ws.merge_cells(start_row=row + 12, start_column=10, end_row=row + 13, end_column=10)
            ws.cell(row=row + 12, column=10).value, ws.cell(row=row + 12,
                                                           column=10).font = 'Hỏng vỡ', font_12_bold
            ws.cell(row=row + 12, column=10).border, ws.cell(row=row + 12,
                                                            column=10).alignment = all_border_thin, alignment_center
            ws.cell(row=row + 13, column=10).border = all_border_thin

            ws.merge_cells(start_row=row + 12, start_column=11, end_row=row + 13, end_column=11)
            ws.cell(row=row + 12, column=11).value, ws.cell(row=row + 12,
                                                           column=11).font = 'Ghi chú', font_12_bold
            ws.cell(row=row + 12, column=11).border, ws.cell(row=row + 12,
                                                            column=11).alignment = all_border_thin, alignment_center
            ws.cell(row=row + 13, column=11).border = all_border_thin

            row += 14
            for line in inventory.line_ids:
                ws.cell(row=row, column=2).value, ws.cell(row=row, column=2).font = row - 14, main_font
                ws.cell(row=row, column=2).border, ws.cell(row=row, column=2).alignment = all_border_thin, alignment_center

                ws.cell(row=row, column=3).value, ws.cell(row=row, column=3).font = line.product_id.name, main_font
                ws.cell(row=row, column=3).border, ws.cell(row=row, column=3).alignment = all_border_thin, Alignment(
                    horizontal='left', vertical='center', wrap_text=True)
                max_length_product = len(line.product_id.name) if line.product_id.name else 0

                ws.cell(row=row, column=4).value, ws.cell(row=row, column=4).font = line.product_uom_id.name, main_font
                ws.cell(row=row, column=4).border, ws.cell(row=row,
                                                           column=4).alignment = all_border_thin, alignment_center
                ws.cell(row=row, column=5).value, ws.cell(row=row, column=5).font = line.prod_lot_id.name or '', main_font
                ws.cell(row=row, column=5).border, ws.cell(row=row,
                                                           column=5).alignment = all_border_thin, alignment_center
                max_length_lot = len(line.prod_lot_id.name) if line.prod_lot_id.name else 0

                medicine_details = self.env['sh.medical.medicines'].sudo().search([('default_code', '=',
                                                                                    line.product_id.default_code)], limit=1)
                if medicine_details:
                    ws.cell(row=row, column=6).value, ws.cell(row=row, column=6).font =\
                        medicine_details[0].origin.name or '', main_font

                ws.cell(row=row, column=6).border, ws.cell(row=row, column=6).alignment = \
                    all_border_thin, alignment_center

                ws.cell(row=row, column=7).value, ws.cell(row=row, column=7).font = \
                    line.prod_lot_id.removal_date.strftime('%d/%m/%Y') if line.prod_lot_id.removal_date else '', main_font
                ws.cell(row=row, column=7).border, ws.cell(row=row, column=7).alignment = all_border_thin, alignment_center

                ws.cell(row=row, column=8).value, ws.cell(row=row, column=8).font = line.theoretical_qty, main_font
                ws.cell(row=row, column=8).border, ws.cell(row=row, column=8).alignment = all_border_thin, alignment_center
                ws.cell(row=row, column=9).value, ws.cell(row=row, column=9).font = line.product_qty, main_font
                ws.cell(row=row, column=9).border, ws.cell(row=row, column=9).alignment = all_border_thin, alignment_center
                ws.cell(row=row, column=10).border = all_border_thin
                ws.cell(row=row, column=11).border = all_border_thin
                row += 1

                max_length = max_length_product if max_length_product > max_length_lot else max_length_lot
                adjusted_height = (max_length + 2) * 1.2
                if adjusted_height > 0:
                    ws.row_dimensions[row].height = adjusted_height

                col_widths = [('a', 1.57), ('b', 5), ('c', 30.71), ('d', 6.71), ('e', 19), ('f', 15.86), ('g', 16),
                              ('h', 9.86), ('i', 10.14), ('j', 7.86), ('k', 9.86)]
                for value in col_widths:
                    ws.column_dimensions[value[0]].width = value[1]

            # Đề xuất
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
            ws.cell(row=row, column=2).value, ws.cell(row=row, column=2).font = 'Ý kiến đề xuất:', font_14_bold
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='left')
            ws.merge_cells(start_row=row+1, start_column=2, end_row=row+1, end_column=11)
            ws.cell(row=row+1, column=2).alignment, ws.cell(row=row+1, column=2).font = Alignment(
                horizontal='left', wrap_text=True), main_font

            #Ký
            ws.merge_cells(start_row=row+3, start_column=2, end_row=row+3, end_column=4)
            ws.merge_cells(start_row=row+3, start_column=8, end_row=row+3, end_column=11)
            ws.cell(row=row+3, column=2).value, ws.cell(row=row+3, column=2).font = 'THÀNH VIÊN HĐ KIỂM KÊ', font_14_bold
            ws.cell(row=row + 3, column=6).value, ws.cell(row=row + 3, column=6).font = 'THƯ KÝ', font_14_bold
            ws.cell(row=row + 3, column=8).value, ws.cell(row=row + 3, column=8).font = 'CHỦ TỊCH HĐ KIỂM KÊ', font_14_bold
            ws.cell(row=row+3, column=2).alignment = alignment_center
            ws.cell(row=row+3, column=6).alignment = alignment_center
            ws.cell(row=row+3, column=8).alignment = alignment_center

            row += 7
            ws.row_breaks.append(Break(id=row))
            row += 1

        row -= 1
        # ws.sheet_view.showGridLines = False
        ws.print_area = 'A1:K%s' % str(row)
        ws.set_printer_settings(paper_size=9, orientation='landscape')
        ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.5, bottom=0.5, header=0.3, footer=0.3)

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        report = base64.encodebytes((fp.read()))
        fp.close()
        attachment = self.env['ir.attachment'].sudo().create({'name': 'MS report',
                                                              'datas_fname': 'BB_KIEM_KE.xlsx',
                                                              'datas': report,
                                                              'res_model': 'temp.creation',
                                                              'public': True})

        url = "/web/content/?model=ir.attachment&id=%s&filename_field=datas_fname&field=datas&download=true&filename=BB_KIEM_KE.xlsx" \
              % (attachment.id)
        cron_clean_attachment = self.env.ref('ms_templates.clean_attachments')
        cron_clean_attachment.sudo().nextcall = fields.Datetime.now() + relativedelta(seconds=10)
        return {'name': 'BIÊN BẢN KIỂM KÊ',
                'type': 'ir.actions.act_url',
                'url': url,
                'target': 'self',
        }

        return {'name': 'BIÊN BẢN KIỂM KÊ',
                'type': 'ir.actions.act_window',
                'res_model': 'temp.wizard',
                'view_mode': 'form',
                'view_type': 'form',
                'target': 'inline',
                'view_id': self.env.ref('ms_templates.report_wizard').id,
                'context': {'attachment_id': attachment.id}
        }

#TIÊU HỦY THUỐC HẾT HẠN
class SHealthStockScrap(models.Model):
    _inherit = 'stock.scrap'

    sci_date_done = fields.Datetime('Ngày hoàn thành', copy=False, help="Ngày ghi nhận hoàn thành SCI", states={'done': [('readonly', True)]}, default=lambda self: fields.Datetime.now())
    note = fields.Text('Lý do chi tiết', copy=False, help="Ghi nhận lý do", states={'done': [('readonly', True)]})

    room_use = fields.Many2one('sh.medical.health.center.ot',copy=False, string='Phòng sử dụng', help="Xuất cho phòng nào sử dụng", states={'done': [('readonly', True)]})
    date_expected = fields.Datetime(states={'done': [('readonly', True)]})
    origin = fields.Char(states={'done': [('readonly', True)]})

    @api.onchange('product_id','location_id')
    def onchange_product_id(self):
        if self.product_id:
            self.product_uom_id = self.product_id.uom_id

            #báo lỗi khi chưa chọn tủ xuất
            if not self.location_id:
                raise ValidationError(_('Bạn phải chọn tủ xuất!'))

            #lấy tất cả các lô của sp ở địa điểm tủ xuất đã chọn và số lượng tồn > 0
            data_lot_inlocation = self.env['stock.quant'].search([('product_id', '=', self.product_id.id), ('location_id', '=', self.location_id.id), ('quantity', '>=', 1)])

            #nếu có data lô thì set trường lô hàng là lô sắp hết hạn
            if data_lot_inlocation:
                self.lot_id = self.env['stock.production.lot'].sudo().search(
                    [('id', 'in', data_lot_inlocation.mapped('lot_id').ids)], limit=1,
                    order='removal_date asc, create_date asc')
            else:
                self.lot_id = False
            #nếu trường lô hàng ko dc set (khi sp tracking by lot) thì thông báo ko khà dụng tại tủ xuất này
            if self.product_id.tracking == 'lot' and not self.lot_id:
                raise ValidationError(_("%s không có số lượng khả dụng tại Tủ xuất này!" % self.product_id.name))
            #nếu số lượng > 0 thì kiểm tra số lượng đủ xuất ko thì thông báo
            elif self.scrap_qty > 0:
                quantity_on_hand = self.env['stock.quant']._get_available_quantity(self.product_id,
                                                                                   self.location_id,
                                                                                   self.lot_id)  # check quantity trong location và lot
                if quantity_on_hand < self.scrap_qty:
                    raise ValidationError(_(
                        "Không đủ số lượng sản phẩm! \nSản phẩm ""%s"" với mã Lô hàng ""%s"" chỉ còn %d %s tại ""%s""." % (
                            self.product_id.name, self.lot_id.name, quantity_on_hand, self.product_id.uom_id.name, self.location_id.name)))
            elif self.scrap_qty <= 0:
                raise ValidationError(_("Số lượng nhập phải lớn hơn 0!"))
            else:
                raise ValidationError(_("Bạn phải chọn Tủ xuất!"))

            return {'domain': {'lot_id': [('id', 'in', data_lot_inlocation.mapped('lot_id').ids)]}}

    @api.onchange('lot_id','scrap_qty')
    def onchange_lot_id(self):
        if self.product_id and self.location_id:
            if self.scrap_qty <= 0:
                raise ValidationError(_("Số lượng nhập phải lớn hơn 0!"))

            quantity_on_hand = self.env['stock.quant']._get_available_quantity(self.product_id,
                                                                                   self.location_id,
                                                                                   self.lot_id)  # check quantity trong location và lot
            if quantity_on_hand < self.scrap_qty:
                raise ValidationError(_(
                    "Không đủ số lượng sản phẩm! \nSản phẩm ""%s"" với mã Lô hàng ""%s"" chỉ còn %d %s tại ""%s""." % (
                        self.product_id.name, self.lot_id.name, quantity_on_hand, self.product_id.uom_id.name,
                        self.location_id.name)))

    # @api.onchange('scrap_qty')
    # def onchange_scrap_qty(self):
    #     if self.scrap_qty <= 0:
    #         raise ValidationError(_("Số lượng nhập phải lớn hơn 0!"))

    # chỉ hiển thị default stock khi có quyền khoa dược
    def _get_default_location_id(self):
        company_user = self.env.user.company_id
        warehouse = self.env['stock.warehouse'].search([('company_id', '=', company_user.id)], limit=1)
        if warehouse:
            #quyền quản ls kho dược
            if self.env.user.has_group('shealth_all_in_one.group_sh_medical_stock_manager'):
                return warehouse.lot_stock_id.id
            # quyen dieu duong
            elif self.env.user.has_group('shealth_all_in_one.group_sh_medical_nurse') and self.env.user.physician_ids:
                physician_loc = self.env.user.physician_ids[0].department.mapped('location_id').ids
                return physician_loc[0].child_ids[0].id or False
            else:
                #quyền bác sĩ
                grp_loc_dict = {
                    'shealth_all_in_one.group_sh_medical_physician_subclinical_labtest': 'shealth_all_in_one.sh_labtest_dep_knhn',
                    # 'shealth_all_in_one.group_sh_medical_physician_subclinical_imaging': 'shealth_all_in_one.sh_imaging_dep_knhn',
                    'shealth_all_in_one.group_sh_medical_physician_surgery': 'shealth_all_in_one.sh_surgeries_dep_knhn',
                    'shealth_all_in_one.group_sh_medical_physician_odontology': 'shealth_all_in_one.sh_rhm_dep_knhn',
                    'shealth_all_in_one.group_sh_medical_physician_spa': 'shealth_all_in_one.sh_dalieu_dep_knhn'}
                flag = False
                for grp, loc in grp_loc_dict.items():
                    if self.env.user.has_group(grp):
                        flag = True
                        return self.env.ref(loc).location_id.child_ids[0].id or False

                if not flag:
                    return False
        return None

    location_id = fields.Many2one('stock.location', default=_get_default_location_id)

    @api.model
    def fields_view_get(self, view_id=None, view_type='form', toolbar=False, submenu=False):
        res = super(SHealthStockScrap, self).fields_view_get(view_id=view_id, view_type=view_type, toolbar=toolbar,
                                                               submenu=submenu)

        access_location = []
        # NẾU QUẢN LÝ KHO HOẶC ADMIN HOẶC NHẬP LIỆU => XEM TẤT CẢ LOCATION
        if self.env.user.has_group('shealth_all_in_one.group_sh_medical_stock_manager'):
            location_stock0 = self.env.ref('stock.warehouse0').lot_stock_id.id  # dia diem kho tổng
            access_location.append(location_stock0)
        else:
            grp_loc_dict = {
                'shealth_all_in_one.group_sh_medical_physician_subclinical_labtest': 'shealth_all_in_one.sh_labtest_dep_knhn',
                # 'shealth_all_in_one.group_sh_medical_physician_subclinical_imaging': 'shealth_all_in_one.sh_imaging_dep_knhn',
                'shealth_all_in_one.group_sh_medical_physician_surgery': 'shealth_all_in_one.sh_surgeries_dep_knhn',
                'shealth_all_in_one.group_sh_medical_physician_odontology': 'shealth_all_in_one.sh_rhm_dep_knhn',
                'shealth_all_in_one.group_sh_medical_physician_spa': 'shealth_all_in_one.sh_dalieu_dep_knhn'}
            for grp, loc in grp_loc_dict.items():
                if self.env.user.has_group(grp):
                    access_location.append(self.env.ref(loc).location_id.id)

            # quyen dieu duong
            if self.env.user.has_group('shealth_all_in_one.group_sh_medical_nurse') and self.env.user.physician_ids:
                physician_loc = self.env.user.physician_ids[0].department.mapped('location_id').ids
                access_location += physician_loc

        doc = etree.XML(res['arch'])

        for t in doc.xpath("//" + view_type):
            t.attrib['delete'] = 'false'
            t.attrib['duplicate'] = 'false'

        for node in doc.xpath("//field[@name='location_id']"):
            # if self.env.user.has_group('shealth_all_in_one.group_sh_medical_stock_manager'):
            #     node.set('domain',"['|',('id', '=', %d),('name', 'ilike', 'Tủ')]" % (self.env.ref('stock.stock_location_stock').id))
            # else:
            #     # default_location_id =
            #     node.set('domain',
            #              "[('location_id', 'child_of', %s),('name', 'ilike', 'Tủ')]" % (str(access_location)))
            # setup_modifiers(node, res['fields']['location_id'])

            node.set('domain', "[('location_id', 'child_of', %s),'|',('name', 'ilike', 'Tủ'),('name', 'ilike', 'dược')]" % (str(access_location)))
            setup_modifiers(node, res['fields']['location_id'])

        res['arch'] = etree.tostring(doc, encoding='unicode')
        return res

    # overide hàm này để ghi nhận lại ngày hoàn thành hủy
    @api.multi
    def do_scrap(self):
        for scrap in self:
            move = self.env['stock.move'].create(scrap._prepare_move_values())
            # master: replace context by cancel_backorder
            move.with_context(is_scrap=True)._action_done()
            #ghi nhận lại ngày hoàn thành stock.move
            move.write({'date': scrap.sci_date_done or fields.Datetime.now()})
            # ghi nhận lại ngày hoàn thành stock.move.line
            move.move_line_ids.write({'date': scrap.sci_date_done or fields.Datetime.now()})

            scrap.write({'move_id': move.id, 'state': 'done'})
        return True

    @api.model
    def create(self, vals):  # ghi nhận mã sequence khác nhau giữa xuất sd phòng và tiêu hủy
        if 'name' not in vals or vals['name'] == _('New'):
            if vals.get('scrap_location_id') == self._get_default_scrap_location_id():
                vals['name'] = self.env['ir.sequence'].next_by_code('stock.scrap') or _('New')
            else:
                vals['name'] = self.env['ir.sequence'].next_by_code('stock.scrap.room.use') or _('New')
                if not vals.get('lot_id'):
                    # lấy tất cả các lô của sp ở địa điểm tủ xuất đã chọn và số lượng tồn > 0
                    data_lot_inlocation = self.env['stock.quant'].search(
                        [('product_id', '=', vals.get('product_id')), ('location_id', '=', vals.get('location_id')),
                         ('quantity', '>', 0)])

                    # nếu có data lô thì set trường lô hàng là lô sắp hết hạn
                    if data_lot_inlocation:
                        vals['lot_id'] = self.env['stock.production.lot'].sudo().search(
                            [('id', 'in', data_lot_inlocation.mapped('lot_id').ids)], limit=1,
                            order='removal_date asc, create_date asc').id

        scrap = super(SHealthStockScrap, self).create(vals)
        return scrap

#ĐƠN MUA HÀNG NCC
class SHealthPurchaseOrder(models.Model):
    _inherit = 'purchase.order'

    date_year = fields.Char('Năm')

    @api.onchange('date_order', 'partner_id')
    def _get_year(self):
        if self.date_order:
            date_order = self.date_order.strftime('%d%m%Y')
            self.date_year = date_order

    def _update_purchase(self):
        employee_ids = self.search([])
        for rec in employee_ids:
            print('chạy')
            rec.date_year = str(rec.date_order.strftime('%d%m%Y'))

    @api.model_cr_context
    def _auto_init(self):
        print('Tạo constraints mới')
        super(SHealthPurchaseOrder, self)._auto_init()
        self._sql_constraints += [
            ('name_sci_year_uniq',
             'unique(name,partner_id,date_year)',
             'Mã đơn hàng của nhà cung cấp này đã tồn tại trong hệ thống'),
        ]
        self._add_sql_constraints()

    @api.onchange('order_line')
    def onchange_date_order(self):
        for purchase in self:
            for line in purchase.order_line:
                line.date_planned = purchase.date_order

    @api.multi
    def button_cancel(self):
        # KHÔNG CHO HỦY ĐƠN HÀNG CHƯA CÓ SẢN PHẨM NÀO DC CHỌN
        for purchase in self:
            if not purchase.order_line:
                raise UserError(_('Đơn hàng của bạn không có sản phẩm nào để hủy bỏ!'))

        return super(SHealthPurchaseOrder, self).button_cancel()

    @api.multi
    def button_confirm(self):
        # KHÔNG CHO XÁC NHẬN ĐƠN HÀNG CHƯA CÓ SẢN PHẨM NÀO DC CHỌN
        for purchase in self:
            if not purchase.order_line:
                raise UserError(_('Bạn phải chọn ít nhất 1 sản phẩm trước khi xác nhận đơn hàng này!'))

        res = super(SHealthPurchaseOrder, self).button_confirm()
        #khi xác nhận đơn mua NCC sẽ xác nhận phiếu nhập kho luôn theo ngày đặt hàng
        for purchase in self:
            picking_related = purchase.picking_ids[0]
            picking_related.sci_date_done = purchase.date_order

            # set hoàn thành
            for move_line in picking_related.move_ids_without_package:
                for move_line_detail in move_line.move_line_ids:
                    #check nếu là thuốc hay vật tư thì chuyển thẳng về tủ tương ứng
                    if move_line_detail.product_id.categ_id == self.env.ref('shealth_all_in_one.sh_medicines'):
                        move_line_detail.location_dest_id = self.env.ref('shealth_all_in_one.sh_location_medicine_prescription_knhn')
                    elif move_line_detail.product_id.categ_id == self.env.ref('shealth_all_in_one.sh_supplies'):
                        move_line_detail.location_dest_id = self.env.ref(
                            'shealth_all_in_one.sh_location_supply_knhn')
                    # move_line_detail.qty_done = move_line_detail.product_uom_qty
            # KO TỰ ĐỘNG XÁC NHẬN PHIẾU NHẬP KHO TỪ NHÀ CUNG CẤP NỮA
            # picking_related.button_validate()

        return res

#ĐƠN MUA HÀNG NCC
class SHealthPurchaseOrderLine(models.Model):
    _inherit = 'purchase.order.line'

    @api.onchange('product_qty', 'product_id')
    def onchange_product_qty(self):
        if self.product_qty <= 0 and self.product_id:
            raise UserError(_("Số lượng nhập phải lớn hơn 0!"))


class wizardMultiProductPurchase(models.TransientModel):
    _name = 'wizard.multi.product.purchase'

    product_ids = fields.Many2many('product.product', 'sci_product_product_purchase_rel', 'purchase_id', 'product_id', string="Sản phẩm")
    date_order = fields.Datetime('Ngày đặt hàng', default=lambda self: fields.Datetime.now())

    @api.multi
    def add_product(self):
        for line in self.product_ids:
            self.env['purchase.order.line'].create({
                'product_id': line.id,
                'name':line.name,
                'product_qty':1,
                'price_unit':line.standard_price,
                'order_id': self._context.get('active_id'),
                'date_planned': self.date_order,
                'product_uom': line.uom_po_id.id or line.uom_id.id
            })
        return


class SCIStockQuant(models.Model):
    _inherit = 'stock.quant'

    def _gather(self, product_id, location_id, lot_id=None, package_id=None, owner_id=None, strict=False):
        removal_strategy = self._get_removal_strategy(product_id, location_id)
        removal_strategy_order = self._get_removal_strategy_order(removal_strategy)
        domain = [
            ('product_id', '=', product_id.id),
        ]
        if not strict:
            if lot_id:
                domain = expression.AND([[('lot_id', '=', lot_id.id)], domain])
            if package_id:
                domain = expression.AND([[('package_id', '=', package_id.id)], domain])
            if owner_id:
                domain = expression.AND([[('owner_id', '=', owner_id.id)], domain])
            if self.env.context.get('exact_location'):
                domain = expression.AND([[('location_id', '=', location_id.id)], domain])
            else:
                domain = expression.AND([[('location_id', 'child_of', location_id.id)], domain])
            # print(domain)
        else:
            domain = expression.AND([[('lot_id', '=', lot_id and lot_id.id or False)], domain])
            domain = expression.AND([[('package_id', '=', package_id and package_id.id or False)], domain])
            domain = expression.AND([[('owner_id', '=', owner_id and owner_id.id or False)], domain])
            domain = expression.AND([[('location_id', '=', location_id.id)], domain])

        # Copy code of _search for special NULLS FIRST/LAST order
        self.sudo(self._uid).check_access_rights('read')
        query = self._where_calc(domain)
        self._apply_ir_rules(query, 'read')
        from_clause, where_clause, where_clause_params = query.get_sql()
        where_str = where_clause and (" WHERE %s" % where_clause) or ''
        query_str = 'SELECT "%s".id FROM ' % self._table + from_clause + where_str + " ORDER BY "+ removal_strategy_order
        self._cr.execute(query_str, where_clause_params)
        res = self._cr.fetchall()
        # No uniquify list necessary as auto_join is not applied anyways...
        return self.browse([x[0] for x in res])


class SCIStockImmediateTransfer(models.TransientModel):
    _inherit = 'stock.immediate.transfer'

    def process(self):
        pick_to_backorder = self.env['stock.picking']
        pick_to_do = self.env['stock.picking']
        for picking in self.pick_ids:
            # If still in draft => confirm and assign
            if picking.state == 'draft':
                picking.action_confirm()
                if picking.state != 'assigned':
                    picking.action_assign()
                    if picking.state != 'assigned':
                        raise UserError(_("Could not reserve all requested products. Please use the \'Mark as Todo\' button to handle the reservation manually."))
            for move in picking.move_lines.filtered(lambda m: m.state not in ['done', 'cancel'] and m.quantity_done == 0):
                for move_line in move.move_line_ids:
                    move_line.qty_done = move_line.product_uom_qty
            if picking._check_backorder():
                pick_to_backorder |= picking
                continue
            pick_to_do |= picking
        # Process every picking that do not require a backorder, then return a single backorder wizard for every other ones.
        if pick_to_do:
            pick_to_do.action_done()
        if pick_to_backorder:
            return pick_to_backorder.action_generate_backorder_wizard()
        return False